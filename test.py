import pymysql
import datetime
from datetime import datetime
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tqdm import tqdm
import time
import config
import dateutil.relativedelta
import re
from openpyxl import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import requests
from slacker import Slacker
from reqStatus import requestStaus

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

rex = re.compile('_F[0-9]+')

wb = Workbook()

ws = wb.active
no = 2

cancelState = f'''
        select `channel_order_number`,`fcode`, `product_name`,
        `product_option`, `state`, `cancel_reason`, `cancel_detail_reason`
        from `11st_cancel`;
        '''
cursor.execute(cancelState)
cancelNowTotal = cursor.fetchall()
for cancelNow in cancelNowTotal:
    channel_order_number = cancelNow[0]
    fcode = cancelNow[1]
    product_name = cancelNow[2]
    product_option = cancelNow[3]
    state = cancelNow[4]
    cancelReason = cancelNow[5]
    cancelDetailReason = cancelNow[6]
    print(channel_order_number, fcode)
    bflowStatus = requestStaus(channel_order_number, fcode)

    if bflowStatus['success'] is True:

        product_order_number = bflowStatus['message']['orderItemOptionId']
        order_number = bflowStatus['message']['orderCode']
        orderState = bflowStatus['message']['status']
        channel = bflowStatus['message']['channel']
        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

            if claimType is None:
                claim = None
            else:
                claim = claimType + "/" + claimStatus
        else:
            claim = None
    else:
        product_order_number = None
        order_number = None
        orderState = None
        claimType = None
        claimStatus = None
        claim = None


    ws.cell(row=1, column=1).value = '상품주문번호'
    ws.cell(row=1, column=2).value = '주문번호'
    ws.cell(row=1, column=3).value = '외부채널주문번호'
    ws.cell(row=1, column=4).value = '상품명'
    ws.cell(row=1, column=5).value = '상품옵션'
    ws.cell(row=1, column=6).value = '브리치 클레임상태'
    ws.cell(row=1, column=7).value = '브리치 주문상태'
    ws.cell(row=1, column=8).value = '11번가 상태'
    ws.cell(row=1, column=9).value = '11번가 클레임이'
    ws.cell(row=1, column=10).value = '11번가 클레임상세이유'
    print(orderState, state)
    if orderState == '결제취소' and state == '취소완료':
        print('skip')
        continue
    else:
        ws.cell(row=no, column=1).value = product_order_number
        ws.cell(row=no, column=2).value = order_number
        ws.cell(row=no, column=3).value = channel_order_number
        ws.cell(row=no, column=4).value = product_name
        ws.cell(row=no, column=5).value = product_option
        ws.cell(row=no, column=6).value = claim
        ws.cell(row=no, column=7).value = orderState
        ws.cell(row=no, column=8).value = state
        ws.cell(row=no, column=9).value = cancelReason
        ws.cell(row=no, column=10).value = cancelDetailReason

        no += 1

result = config.ST_LOGIN['excelPath'] + 'CancelResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()

sql = '''
    select `id`, `product_option` from `channel_order`  where channel = '11st'
    '''
cursor.execute(sql)
optionRows = cursor.fetchall()

for optionRow in optionRows:
    idNo = optionRow[0]
    fcodeText = optionRow[1]
    if fcodeText is None:
        fcode = None
    elif fcodeText is not None:
        fcodeText = rex.search(optionRow[1])
        fcode = fcodeText.group()

    updateSql = '''
                update `channel_order` set fcode = %s where id = %s
                '''
    updateValue = (
        fcode,
        idNo
    )
    cursor.execute(updateSql, updateValue)
    print(updateSql, updateValue)

ebayFcode = f'''
            select c.`id`, c.`channel_order_number`, s.`fcode` 
            from  `channel_order` as c join `sell` as s
            on c.`channel_order_number` = s.`channel_order_number`
            where c.`channel` in ('gmarket', 'auction');
            '''
cursor.execute(ebayFcode)
ebayFcodes = cursor.fetchall()

for Fcode in ebayFcodes:
    idNo = Fcode[0]
    channelOrderNumber = Fcode[1]
    fcode = Fcode[2]

    FcodeUpdate = '''
                update `channel_order` set fcode = %s where id = %s
                '''
    updateValue = (
        fcode,
        idNo
    )
    cursor.execute(FcodeUpdate, updateValue)
    print(FcodeUpdate, updateValue)

print('FcodeUpdate')

ebayOrderList = f'''
            select `channel_order_number`,`fcode`, `product_name`, `product_option`, `state` 
            from `channel_order` where `payment_at` >= {endNow} and `channel` in ('gmarket', 'auction')
            and state not in ('입금대기', '판매자송금', '구매결정완료'); 
            '''

cursor.execute(ebayOrderList)
ebayOrderRows = cursor.fetchall()

wb = Workbook()

ws = wb.active
no = 2

for ebayOrderRow in ebayOrderRows:
    channelOrderNumber = ebayOrderRow[0]
    fcode = ebayOrderRow[1]
    productName = ebayOrderRow[2]
    productOption = ebayOrderRow[3]
    state = ebayOrderRow[4]


    bflowStatus = requestStaus(channelOrderNumber, fcode)

    if bflowStatus['success'] is True:

        productOrderNumber = bflowStatus['message']['orderItemOptionId']
        orderState = bflowStatus['message']['status']
        channel = bflowStatus['message']['channel']
        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

            if claimType is None:
                claim = None
            else:
                claim = claimType + "/" + claimStatus
        else:
            claim = None
    else:
        productOrderNumber = None
        orderState = None
        claimType = None
        claimStatus = None
        claim = None

    ws.cell(row=1, column=1).value = '상품주문번호'
    ws.cell(row=1, column=2).value = '외부채널주문번호'
    ws.cell(row=1, column=3).value = '상품명'
    ws.cell(row=1, column=4).value = '상품옵션'
    ws.cell(row=1, column=5).value = '브리치 주문상태'
    ws.cell(row=1, column=6).value = '브리치 클레임상태'
    ws.cell(row=1, column=7).value = '채널 상태'
    ws.cell(row=1, column=8).value = '채널명'

    # 비플로우 결제취소 / 채널상태 취소요청 , 취소중 , 반품완료  => 불필요
    # 비플로우 교환 / 채널상태 배송중, 교환요청 , 구매결정완료 => 불필요
    # 비플로우 반품 / 채널상태 반품보류 , 반품요청 = > 불필요
    # 비플로우 배송준비 / 채널 상태 배송지연 / 발송예정 = > 불필요
    # 비플로우 배송지연 / 채널상태 배송지연 / 발송예정  = > 불필요
    if state == '교환수거완료'\
            or state == '교환수거중'\
            or state == '교환완료'\
            or state == '배송중'\
            or state == '교환요청'\
            or state == '구매결정완료'\
            and orderState == '교환':
        print('skip')
        continue
    elif state == '반품수거완료'\
            or state == '반품수거중'\
            or state == '반품완료'\
            or state == '반품보류'\
            or state == '반품요청'\
            and orderState == '반품':
        print('skip')
        continue
    elif state == '입금확인'\
            or state == '주문확인'\
            or state == '배송지연/발송예정'\
            and orderState == '배송준비'\
            or orderState == '결제확인':
        print('skip')
        continue
    elif state == '취소완료'\
            or state == '환불완료'\
            or state == '취소요청'\
            or state == '취소중'\
            or state == '반품완료'\
            and orderState == '결제취소':
        print('skip')
        continue
    elif state == '배송중' and orderState == '출고완료' or orderState == '배송중':
        print('skip')
        continue
    elif state == '주문확인' or state == '배송지연/발송예정' and orderState == '배송지연':
        print('skip')
        continue
    elif state == '배송완료' or state == '구매결정완료' and orderState == '배송완료':
        print('skip')
        continue
    elif state == '미입금구매취소':
        print('skip')
        continue
    elif state == '입금대기':
        print('skip')
        continue
    elif state == '판매자송금':
        print('skip')
        continue
    elif state == '반품보류' or state == '미수취신고':
        ws.cell(row=no, column=1).value = productOrderNumber
        ws.cell(row=no, column=2).value = channelOrderNumber
        ws.cell(row=no, column=3).value = productName
        ws.cell(row=no, column=4).value = productOption
        ws.cell(row=no, column=5).value = claim
        ws.cell(row=no, column=6).value = orderState
        ws.cell(row=no, column=7).value = state
        ws.cell(row=no, column=8).value = channel
        no += 1
    else:
        ws.cell(row=no, column=1).value = productOrderNumber
        ws.cell(row=no, column=2).value = channelOrderNumber
        ws.cell(row=no, column=3).value = productName
        ws.cell(row=no, column=4).value = productOption
        ws.cell(row=no, column=5).value = claim
        ws.cell(row=no, column=6).value = orderState
        ws.cell(row=no, column=7).value = state
        ws.cell(row=no, column=8).value = channel
        no += 1

result = config.ST_LOGIN['excelPath'] + 'ebayOrderResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()
print(result)
cursor.close()
db.close()
