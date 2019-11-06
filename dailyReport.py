import pymysql
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils.cell import get_column_letter
import numpy as np
import config

wb = Workbook()

ws = wb.active

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

reportMonth = 10, 11

weekSql =f'''
        SELECT 
        week,
        min(date),
        max(date),
        sum(brich_total_amount),
        sum(brich_qty),
        sum(brich_sales),
        sum(brich_cogs),
        sum(brich_refund_amount),
        sum(brich_refund_qty),
        sum(gmarket_total_amount),
        sum(gmarket_qty),
        sum(gmarket_sales),
        sum(gmarket_cogs),
        sum(gmarket_refund_amount),
        sum(gmarket_refund_qty),
        sum(auction_total_amount),
        sum(auction_qty),
        sum(auction_sales),
        sum(auction_cogs),
        sum(auction_refund_amount),
        sum(auction_refund_qty),
        sum(11st_total_amount),
        sum(11st_qty),
        sum(11st_sales),
        sum(11st_cogs),
        sum(11st_refund_amount),
        sum(11st_refund_qty),
        sum(g9_total_amount),
        sum(g9_qty),
        sum(g9_sales),
        sum(g9_cogs),
        sum(g9_refund_amount),
        sum(g9_refund_qty),
        sum(interpark_total_amount),
        sum(interpark_qty),
        sum(interpark_sales),
        sum(interpark_cogs),
        sum(interpark_refund_amount),
        sum(interpark_refund_qty),
        sum(wemakeprice_total_amount),
        sum(wemakeprice_qty),
        sum(wemakeprice_sales),
        sum(wemakeprice_cogs),
        sum(wemakeprice_refund_amount),
        sum(wemakeprice_refund_qty),
        sum(coupang_total_amount),
        sum(coupang_qty),
        sum(coupang_sales),
        sum(coupang_cogs),
        sum(coupang_refund_amount),
        sum(coupang_refund_qty),
        sum(tmon_total_amount),
        sum(tmon_qty),
        sum(tmon_sales),
        sum(tmon_cogs),
        sum(tmon_refund_amount),
        sum(tmon_refund_qty),
        sum(ssg_total_amount),
        sum(ssg_qty),
        sum(ssg_sales),
        sum(ssg_cogs),
        sum(ssg_refund_amount),
        sum(ssg_refund_qty)
        FROM sell_to_channel where month in ({reportMonth[0]},{reportMonth[1]}) GROUP BY week
        '''

cursor.execute(weekSql)
weekAmounts = cursor.fetchall()


def intNone (text):
    if text is None:
        return 0
    else:
        text = int(text)
        return text

no = 2

for weekAmount in weekAmounts:
    week = weekAmount[0]
    minDate = datetime.strftime(weekAmount[1], '%Y-%m-%d')
    maxDate = datetime.strftime(weekAmount[2], '%Y-%m-%d')
    brich_total_amount = intNone(weekAmount[3])
    brich_total_qty = intNone(weekAmount[4])
    brich_total_sales = intNone(weekAmount[5])
    brich_total_cogs = intNone(weekAmount[6])
    brich_total_refund_amount = intNone(weekAmount[7])
    brich_total_refund_qty = intNone(weekAmount[8])
    gmarket_total_amount = intNone(weekAmount[9])
    gmarket_total_qty = intNone(weekAmount[10])
    gmarket_total_sales = intNone(weekAmount[11])
    gmarket_total_cogs = intNone(weekAmount[12])
    gmarket_total_refund_amount = intNone(weekAmount[13])
    gmarket_total_refund_qty = intNone(weekAmount[14])
    auction_total_amount = intNone(weekAmount[15])
    auction_total_qty = intNone(weekAmount[16])
    auction_total_sales = intNone(weekAmount[17])
    auction_total_cogs = intNone(weekAmount[18])
    auction_total_refund_amount = intNone(weekAmount[19])
    auction_total_refund_qty = intNone(weekAmount[20])
    st_total_amount = intNone(weekAmount[21])
    st_total_qty = intNone(weekAmount[22])
    st_total_sales = intNone(weekAmount[23])
    st_total_cogs = intNone(weekAmount[24])
    st_total_refund_amount = intNone(weekAmount[25])
    st_total_refund_qty = intNone(weekAmount[26])
    g9_total_amount = intNone(weekAmount[27])
    g9_total_qty = intNone(weekAmount[28])
    g9_total_sales = intNone(weekAmount[29])
    g9_total_cogs = intNone(weekAmount[30])
    g9_total_refund_amount = intNone(weekAmount[31])
    g9_total_refund_qty = intNone(weekAmount[32])
    interpark_total_amount = intNone(weekAmount[33])
    interpark_total_qty = intNone(weekAmount[34])
    interpark_total_sales = intNone(weekAmount[35])
    interpark_total_cogs = intNone(weekAmount[36])
    interpark_total_refund_amount = intNone(weekAmount[37])
    interpark_total_refund_qty = intNone(weekAmount[38])
    wemakeprice_total_amount = intNone(weekAmount[39])
    wemakeprice_total_qty = intNone(weekAmount[40])
    wemakeprice_total_sales = intNone(weekAmount[41])
    wemakeprice_total_cogs = intNone(weekAmount[42])
    wemakeprice_total_refund_amount = intNone(weekAmount[43])
    wemakeprice_total_refund_qty = intNone(weekAmount[44])
    coupang_total_amount = intNone(weekAmount[45])
    coupang_total_qty = intNone(weekAmount[46])
    coupang_total_sales = intNone(weekAmount[47])
    coupang_total_cogs = intNone(weekAmount[48])
    coupang_total_refund_amount = intNone(weekAmount[49])
    coupang_total_refund_qty = intNone(weekAmount[50])
    tmon_total_amount = intNone(weekAmount[51])
    tmon_total_qty = intNone(weekAmount[52])
    tmon_total_sales = intNone(weekAmount[53])
    tmon_total_cogs = intNone(weekAmount[54])
    tmon_total_refund_amount = intNone(weekAmount[55])
    tmon_total_refund_qty = intNone(weekAmount[56])
    ssg_total_amount = intNone(weekAmount[57])
    ssg_total_qty = intNone(weekAmount[58])
    ssg_total_sales = intNone(weekAmount[59])
    ssg_total_cogs = intNone(weekAmount[60])
    ssg_total_refund_amount = intNone(weekAmount[61])
    ssg_total_refund_qty = intNone(weekAmount[62])
    openmarket_total_amount = gmarket_total_amount + auction_total_amount + st_total_amount + g9_total_amount + interpark_total_amount
    openmarket_total_qty = gmarket_total_qty + auction_total_qty + st_total_qty + g9_total_qty + interpark_total_qty
    openmarket_total_sales = gmarket_total_sales + auction_total_sales + st_total_sales + g9_total_sales + interpark_total_sales
    openmarket_total_cogs = gmarket_total_cogs + auction_total_cogs + st_total_cogs + g9_total_cogs + interpark_total_cogs
    openmarket_total_refund_amount = (
        gmarket_total_refund_amount + auction_total_refund_amount + st_total_refund_amount
        + g9_total_refund_amount + interpark_total_refund_amount
    )
    openmarket_total_refund_qty = (
        gmarket_total_refund_qty + auction_total_refund_qty + st_total_refund_qty
        + g9_total_refund_qty + interpark_total_refund_qty
    )
    social_total_amount = wemakeprice_total_amount + coupang_total_amount + tmon_total_amount
    social_total_qty = wemakeprice_total_qty + coupang_total_qty + tmon_total_qty
    social_total_sales = wemakeprice_total_sales + coupang_total_sales + tmon_total_sales
    social_total_cogs = wemakeprice_total_cogs + coupang_total_cogs + tmon_total_cogs
    social_total_refund_amount = (
        wemakeprice_total_refund_amount
        + coupang_total_refund_amount
        + tmon_total_refund_amount
    )
    social_total_refund_qty = (
        wemakeprice_total_refund_qty
        + coupang_total_refund_qty
        + tmon_total_refund_qty
    )
    multi_total_amount = ssg_total_amount
    multi_total_qty = ssg_total_qty
    multi_total_sales = ssg_total_sales
    multi_total_cogs = ssg_total_cogs
    multi_total_refund_amount = ssg_total_refund_amount
    multi_total_refund_qty = ssg_total_refund_qty

    ws.cell(row=1, column=1).value = '주차'
    ws.cell(row=1, column=2).value = '일자'
    ws.cell(row=1, column=3).value = '채널'
    ws.cell(row=1, column=4).value = '실거래액'
    ws.cell(row=1, column=5).value = '거래액'
    ws.cell(row=1, column=6).value = '매출'
    ws.cell(row=1, column=7).value = '매출원가'
    ws.cell(row=1, column=8).value = '공헌이익'
    ws.cell(row=1, column=9).value = '반품금액'
    ws.cell(row=1, column=10).value = '반품율'

    ws.cell(row=no, column=1).value = week
    ws.cell(row=no, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=no, column=3).value = '브리치'
    ws.cell(row=no, column=4).value = brich_total_amount - brich_total_refund_amount
    ws.cell(row=no, column=5).value = brich_total_amount
    ws.cell(row=no, column=6).value = brich_total_sales
    ws.cell(row=no, column=7).value = brich_total_cogs
    ws.cell(row=no, column=8).value = brich_total_sales - brich_total_cogs
    ws.cell(row=no, column=9).value = brich_total_refund_amount
    ws.cell(row=no, column=10).value = round(brich_total_refund_amount / brich_total_amount, 2)

    ws.cell(row=no + 1, column=1).value = week
    ws.cell(row=no + 1, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=no + 1, column=3).value = '오픈마켓'
    ws.cell(row=no + 1, column=4).value = openmarket_total_amount - openmarket_total_refund_amount
    ws.cell(row=no + 1, column=5).value = openmarket_total_amount
    ws.cell(row=no + 1, column=6).value = openmarket_total_sales
    ws.cell(row=no + 1, column=7).value = openmarket_total_cogs
    ws.cell(row=no + 1, column=8).value = openmarket_total_sales - openmarket_total_cogs
    ws.cell(row=no + 1, column=9).value = openmarket_total_refund_amount
    ws.cell(row=no + 1, column=10).value = round(openmarket_total_refund_amount / openmarket_total_amount, 2)

    ws.cell(row=no + 2, column=1).value = week
    ws.cell(row=no + 2, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=no + 2, column=3).value = '소셜커머스'
    ws.cell(row=no + 2, column=4).value = social_total_amount - social_total_refund_amount
    ws.cell(row=no + 2, column=5).value = social_total_amount
    ws.cell(row=no + 2, column=6).value = social_total_sales
    ws.cell(row=no + 2, column=7).value = social_total_cogs
    ws.cell(row=no + 2, column=8).value = social_total_sales - social_total_cogs
    ws.cell(row=no + 2, column=9).value = social_total_refund_amount
    ws.cell(row=no + 2, column=10).value = round(social_total_refund_amount / social_total_amount, 2)

    ws.cell(row=no + 3, column=1).value = week
    ws.cell(row=no + 3, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=no + 3, column=3).value = '종합몰'
    ws.cell(row=no + 3, column=4).value = multi_total_amount - multi_total_refund_amount
    ws.cell(row=no + 3, column=5).value = multi_total_amount
    ws.cell(row=no + 3, column=6).value = multi_total_sales
    ws.cell(row=no + 3, column=7).value = multi_total_cogs
    ws.cell(row=no + 3, column=8).value = multi_total_sales - multi_total_cogs
    ws.cell(row=no + 3, column=9).value = multi_total_refund_amount
    ws.cell(row=no + 3, column=10).value = round(multi_total_refund_amount / multi_total_amount, 2)

    ws.cell(row=no + 4, column=4).value = f'''=sum(d{no}:d{no + 3})'''
    ws.cell(row=no + 4, column=5).value = f'''=sum(e{no}:e{no + 3})'''
    ws.cell(row=no + 4, column=6).value = f'''=sum(f{no}:f{no + 3})'''
    ws.cell(row=no + 4, column=7).value = f'''=sum(g{no}:g{no + 3})'''
    ws.cell(row=no + 4, column=8).value = f'''=sum(h{no}:h{no + 3})'''
    ws.cell(row=no + 4, column=9).value = f'''=sum(i{no}:i{no + 3})'''
    ws.cell(row=no + 4, column=10).value = f'''=i{no+4}/e{no+4}'''
    no += 6

monthSql =f'''
        SELECT 
        month,
        min(date),
        max(date),
        sum(brich_total_amount),
        sum(brich_qty),
        sum(brich_sales),
        sum(brich_cogs),
        sum(brich_refund_amount),
        sum(brich_refund_qty),
        sum(gmarket_total_amount),
        sum(gmarket_qty),
        sum(gmarket_sales),
        sum(gmarket_cogs),
        sum(gmarket_refund_amount),
        sum(gmarket_refund_qty),
        sum(auction_total_amount),
        sum(auction_qty),
        sum(auction_sales),
        sum(auction_cogs),
        sum(auction_refund_amount),
        sum(auction_refund_qty),
        sum(11st_total_amount),
        sum(11st_qty),
        sum(11st_sales),
        sum(11st_cogs),
        sum(11st_refund_amount),
        sum(11st_refund_qty),
        sum(g9_total_amount),
        sum(g9_qty),
        sum(g9_sales),
        sum(g9_cogs),
        sum(g9_refund_amount),
        sum(g9_refund_qty),
        sum(interpark_total_amount),
        sum(interpark_qty),
        sum(interpark_sales),
        sum(interpark_cogs),
        sum(interpark_refund_amount),
        sum(interpark_refund_qty),
        sum(wemakeprice_total_amount),
        sum(wemakeprice_qty),
        sum(wemakeprice_sales),
        sum(wemakeprice_cogs),
        sum(wemakeprice_refund_amount),
        sum(wemakeprice_refund_qty),
        sum(coupang_total_amount),
        sum(coupang_qty),
        sum(coupang_sales),
        sum(coupang_cogs),
        sum(coupang_refund_amount),
        sum(coupang_refund_qty),
        sum(tmon_total_amount),
        sum(tmon_qty),
        sum(tmon_sales),
        sum(tmon_cogs),
        sum(tmon_refund_amount),
        sum(tmon_refund_qty),
        sum(ssg_total_amount),
        sum(ssg_qty),
        sum(ssg_sales),
        sum(ssg_cogs),
        sum(ssg_refund_amount),
        sum(ssg_refund_qty)
        FROM sell_to_channel where month in ({reportMonth[0]},{reportMonth[1]}) GROUP BY month
        '''

cursor.execute(monthSql)
monthAmounts = cursor.fetchall()


def intNone (text):
    if text is None:
        return 0
    else:
        text = int(text)
        return text


monthNo = ws.max_row + 2
monthFirstNo = ws.max_row + 1

for monthAmount in monthAmounts:
    week = monthAmount[0]
    minDate = datetime.strftime(monthAmount[1], '%Y-%m-%d')
    maxDate = datetime.strftime(monthAmount[2], '%Y-%m-%d')
    brich_total_amount = intNone(monthAmount[3])
    brich_total_qty = intNone(monthAmount[4])
    brich_total_sales = intNone(monthAmount[5])
    brich_total_cogs = intNone(monthAmount[6])
    brich_total_refund_amount = intNone(monthAmount[7])
    brich_total_refund_qty = intNone(monthAmount[8])
    gmarket_total_amount = intNone(monthAmount[9])
    gmarket_total_qty = intNone(monthAmount[10])
    gmarket_total_sales = intNone(monthAmount[11])
    gmarket_total_cogs = intNone(monthAmount[12])
    gmarket_total_refund_amount = intNone(monthAmount[13])
    gmarket_total_refund_qty = intNone(monthAmount[14])
    auction_total_amount = intNone(monthAmount[15])
    auction_total_qty = intNone(monthAmount[16])
    auction_total_sales = intNone(monthAmount[17])
    auction_total_cogs = intNone(monthAmount[18])
    auction_total_refund_amount = intNone(monthAmount[19])
    auction_total_refund_qty = intNone(monthAmount[20])
    st_total_amount = intNone(monthAmount[21])
    st_total_qty = intNone(monthAmount[22])
    st_total_sales = intNone(monthAmount[23])
    st_total_cogs = intNone(monthAmount[24])
    st_total_refund_amount = intNone(monthAmount[25])
    st_total_refund_qty = intNone(monthAmount[26])
    g9_total_amount = intNone(monthAmount[27])
    g9_total_qty = intNone(monthAmount[28])
    g9_total_sales = intNone(monthAmount[29])
    g9_total_cogs = intNone(monthAmount[30])
    g9_total_refund_amount = intNone(monthAmount[31])
    g9_total_refund_qty = intNone(monthAmount[32])
    interpark_total_amount = intNone(monthAmount[33])
    interpark_total_qty = intNone(monthAmount[34])
    interpark_total_sales = intNone(monthAmount[35])
    interpark_total_cogs = intNone(monthAmount[36])
    interpark_total_refund_amount = intNone(monthAmount[37])
    interpark_total_refund_qty = intNone(monthAmount[38])
    wemakeprice_total_amount = intNone(monthAmount[39])
    wemakeprice_total_qty = intNone(monthAmount[40])
    wemakeprice_total_sales = intNone(monthAmount[41])
    wemakeprice_total_cogs = intNone(monthAmount[42])
    wemakeprice_total_refund_amount = intNone(monthAmount[43])
    wemakeprice_total_refund_qty = intNone(monthAmount[44])
    coupang_total_amount = intNone(monthAmount[45])
    coupang_total_qty = intNone(monthAmount[46])
    coupang_total_sales = intNone(monthAmount[47])
    coupang_total_cogs = intNone(monthAmount[48])
    coupang_total_refund_amount = intNone(monthAmount[49])
    coupang_total_refund_qty = intNone(monthAmount[50])
    tmon_total_amount = intNone(monthAmount[51])
    tmon_total_qty = intNone(monthAmount[52])
    tmon_total_sales = intNone(monthAmount[53])
    tmon_total_cogs = intNone(monthAmount[54])
    tmon_total_refund_amount = intNone(monthAmount[55])
    tmon_total_refund_qty = intNone(monthAmount[56])
    ssg_total_amount = intNone(monthAmount[57])
    ssg_total_qty = intNone(monthAmount[58])
    ssg_total_sales = intNone(monthAmount[59])
    ssg_total_cogs = intNone(monthAmount[60])
    ssg_total_refund_amount = intNone(monthAmount[61])
    ssg_total_refund_qty = intNone(monthAmount[62])
    openmarket_total_amount = gmarket_total_amount + auction_total_amount + st_total_amount + g9_total_amount + interpark_total_amount
    openmarket_total_qty = gmarket_total_qty + auction_total_qty + st_total_qty + g9_total_qty + interpark_total_qty
    openmarket_total_sales = gmarket_total_sales + auction_total_sales + st_total_sales + g9_total_sales + interpark_total_sales
    openmarket_total_cogs = gmarket_total_cogs + auction_total_cogs + st_total_cogs + g9_total_cogs + interpark_total_cogs
    openmarket_total_refund_amount = (
        gmarket_total_refund_amount + auction_total_refund_amount + st_total_refund_amount
        + g9_total_refund_amount + interpark_total_refund_amount
    )
    openmarket_total_refund_qty = (
        gmarket_total_refund_qty + auction_total_refund_qty + st_total_refund_qty
        + g9_total_refund_qty + interpark_total_refund_qty
    )
    social_total_amount = wemakeprice_total_amount + coupang_total_amount + tmon_total_amount
    social_total_qty = wemakeprice_total_qty + coupang_total_qty + tmon_total_qty
    social_total_sales = wemakeprice_total_sales + coupang_total_sales + tmon_total_sales
    social_total_cogs = wemakeprice_total_cogs + coupang_total_cogs + tmon_total_cogs
    social_total_refund_amount = (
        wemakeprice_total_refund_amount
        + coupang_total_refund_amount
        + tmon_total_refund_amount
    )
    social_total_refund_qty = (
        wemakeprice_total_refund_qty
        + coupang_total_refund_qty
        + tmon_total_refund_qty
    )
    multi_total_amount = ssg_total_amount
    multi_total_qty = ssg_total_qty
    multi_total_sales = ssg_total_sales
    multi_total_cogs = ssg_total_cogs
    multi_total_refund_amount = ssg_total_refund_amount
    multi_total_refund_qty = ssg_total_refund_qty

    ws.cell(row=monthNo, column=1).value = '월'
    ws.cell(row=monthNo, column=2).value = '일자'
    ws.cell(row=monthNo, column=3).value = '채널'
    ws.cell(row=monthNo, column=4).value = '실거래액'
    ws.cell(row=monthNo, column=5).value = '거래액'
    ws.cell(row=monthNo, column=6).value = '매출'
    ws.cell(row=monthNo, column=7).value = '매출원가'
    ws.cell(row=monthNo, column=8).value = '공헌이익'
    ws.cell(row=monthNo, column=9).value = '반품금액'
    ws.cell(row=monthNo, column=10).value = '반품율'

    ws.cell(row=monthNo, column=1).value = week
    ws.cell(row=monthNo, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=monthNo, column=3).value = '브리치'
    ws.cell(row=monthNo, column=4).value = brich_total_amount - brich_total_refund_amount
    ws.cell(row=monthNo, column=5).value = brich_total_amount
    ws.cell(row=monthNo, column=6).value = brich_total_sales
    ws.cell(row=monthNo, column=7).value = brich_total_cogs
    ws.cell(row=monthNo, column=8).value = brich_total_sales - brich_total_cogs
    ws.cell(row=monthNo, column=9).value = brich_total_refund_amount
    ws.cell(row=monthNo, column=10).value = round(brich_total_refund_amount / brich_total_amount, 2)

    ws.cell(row=monthNo + 1, column=1).value = week
    ws.cell(row=monthNo + 1, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=monthNo + 1, column=3).value = '오픈마켓'
    ws.cell(row=monthNo + 1, column=4).value = openmarket_total_amount - openmarket_total_refund_amount
    ws.cell(row=monthNo + 1, column=5).value = openmarket_total_amount
    ws.cell(row=monthNo + 1, column=6).value = openmarket_total_sales
    ws.cell(row=monthNo + 1, column=7).value = openmarket_total_cogs
    ws.cell(row=monthNo + 1, column=8).value = openmarket_total_sales - openmarket_total_cogs
    ws.cell(row=monthNo + 1, column=9).value = openmarket_total_refund_amount
    ws.cell(row=monthNo + 1, column=10).value = round(openmarket_total_refund_amount / openmarket_total_amount, 2)

    ws.cell(row=monthNo + 2, column=1).value = week
    ws.cell(row=monthNo + 2, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=monthNo + 2, column=3).value = '소셜커머스'
    ws.cell(row=monthNo + 2, column=4).value = social_total_amount - social_total_refund_amount
    ws.cell(row=monthNo + 2, column=5).value = social_total_amount
    ws.cell(row=monthNo + 2, column=6).value = social_total_sales
    ws.cell(row=monthNo + 2, column=7).value = social_total_cogs
    ws.cell(row=monthNo + 2, column=8).value = social_total_sales - social_total_cogs
    ws.cell(row=monthNo + 2, column=9).value = social_total_refund_amount
    ws.cell(row=monthNo + 2, column=10).value = round(social_total_refund_amount / social_total_amount, 2)

    ws.cell(row=monthNo + 3, column=1).value = week
    ws.cell(row=monthNo + 3, column=2).value = minDate + ' - ' + maxDate
    ws.cell(row=monthNo + 3, column=3).value = '종합몰'
    ws.cell(row=monthNo + 3, column=4).value = multi_total_amount - multi_total_refund_amount
    ws.cell(row=monthNo + 3, column=5).value = multi_total_amount
    ws.cell(row=monthNo + 3, column=6).value = multi_total_sales
    ws.cell(row=monthNo + 3, column=7).value = multi_total_cogs
    ws.cell(row=monthNo + 3, column=8).value = multi_total_sales - multi_total_cogs
    ws.cell(row=monthNo + 3, column=9).value = multi_total_refund_amount
    ws.cell(row=monthNo + 3, column=10).value = round(multi_total_refund_amount / multi_total_amount, 2)

    ws.cell(row=monthNo + 4, column=4).value = f'''=sum(d{monthNo}:d{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=5).value = f'''=sum(e{monthNo}:e{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=6).value = f'''=sum(f{monthNo}:f{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=7).value = f'''=sum(g{monthNo}:g{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=8).value = f'''=sum(h{monthNo}:h{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=9).value = f'''=sum(i{monthNo}:i{monthNo + 3})'''
    ws.cell(row=monthNo + 4, column=10).value = f'''=i{monthNo+4}/e{monthNo+4}'''
    monthNo += 6


for col in ws.columns:
    max_length = 0
    columnIndex = col[0].column
    column = get_column_letter(columnIndex)
    for cell in col:
        if max_length < len(str(cell.value)) < 30:
            max_length = len(str(cell.value))
        else:
            pass
    ws.column_dimensions[column].width = (max_length + 1) * 1.2
makeToday = datetime.today()

now = makeToday.strftime("%m%d_%H%M")
result = '2019_데일리_' + now + '.xlsx'
print(result)
wb.save(result)
