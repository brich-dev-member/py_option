from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import date
from datetime import datetime
import pyexcel as p
import pymysql
from openpyxl import load_workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook
from slacker import Slacker
from pyvirtualdisplay import Display

display = Display(visible=0, size=(1200, 900))
display.start()


def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)


def findSelect(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# f코드 정규
rex = re.compile('_F[0-9]+')
rexSpace = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

# 셀레니움 셋
options = Options()
# options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

# driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
# params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "/path/to/download/dir"}}
# command_result = driver.execute("send_command", params)

countSleep(1, 5)
driver.get('https://login.11st.co.kr/auth/front/selleroffice/login.tmall')

driver.find_element_by_id('user-id').send_keys(config.ST_LOGIN['id'])
driver.find_element_by_id('passWord').send_keys(config.ST_LOGIN['password'])
driver.find_element_by_xpath('/html/body/div/form[1]/fieldset/button').click()
time.sleep(5)
driver.get('https://soffice.11st.co.kr/escrow/AuthSellerClaimManager.tmall?method=getClaimList&clm=01&searchVer=02')
time.sleep(2)
windowLists = driver.window_handles
for windowList in windowLists[1:]:
    driver.switch_to.window(driver.window_handles[-1])
    driver.close()

driver.switch_to.window(driver.window_handles[0])

driver.find_element_by_xpath('//*[@id="totalClmCountA"]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="ext-gen7"]/div[2]/div[1]/div[6]/div/a').click()
countSleep(1, 10)


def changeFileToXlsx(originalName, resultName):
    stOriExcel = config.ST_LOGIN['excelPath'] + originalName
    stResultExcel = config.ST_LOGIN['excelPath'] + resultName + now + '.xls'
    stResultXlsx = config.ST_LOGIN['excelPath'] + resultName + now + '.xlsx'

    os.rename(stOriExcel, stResultExcel)

    p.save_book_as(file_name=stResultExcel, dest_file_name=stResultXlsx)
    return stResultXlsx


returnFile = changeFileToXlsx('claimGoodsList.xls', '11st_return_')
print(returnFile)

findSelect('//*[@id="sltDuration"]', 'RECENT_MONTH')
findSelect('//*[@id="clmStat"]', '106')

driver.find_element_by_xpath('//*[@id="ext-gen7"]/div[2]/div[1]/div[4]/div[2]/form/div/div[2]/div/button[1]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="ext-gen7"]/div[2]/div[1]/div[6]/div/a').click()
countSleep(1, 10)

returnCompleteFile = changeFileToXlsx('claimGoodsList.xls', '11st_return_Complete_')
print(returnCompleteFile)


channelSql = '''
    INSERT INTO `bflow`.`channel_returns` (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at
        ) VALUES (
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s
       )
        ON DUPLICATE KEY UPDATE 
        security_refund = %s,
        security_refund_at = %s,
        return_request_at = %s,
        return_complete_at = %s,
        return_delivery_case = %s,
        return_delivery_fees = %s,
        return_request_case = %s,
        channel_delivery_fees = %s,
        return_respons = %s,
        payment_case = %s,
        return_qty = %s,
        refund_state = %s,
        delivery_company = %s,
        delivery_code = %s,
        return_delivery_arrive_at = %s,
        return_hold_at = %s,
        return_delivery_complete_at = %s,
        fcode = %s,
        claim_state = %s
    '''

path = returnFile

wb = load_workbook(path)

ws = wb.active

maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=7, max_row=maxRow):
    order_number = replacenone(row[3].value)
    order_number_line = replacenone(row[4].value)
    return_number = None
    channel = '11st'
    security_refund = replacenone(row[2].value)
    security_refund_at = row[7].value
    return_request_at = row[5].value
    return_complete_at = row[6].value
    return_delivery_case = replacenone(row[29].value)
    return_delivery_fees = replaceint(row[28].value)
    return_request_case = replacenone(row[22].value)
    channel_delivery_fees = replaceint(row[30].value)
    return_respons = replacenone(row[31].value)
    payment_case = replacenone(row[32].value)
    return_qty = replaceint(row[14].value)
    refund_state = replacenone(row[1].value)
    product_name = replacenone(row[9].value)
    product_option = replacenone(row[10].value)
    if product_option is not None:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()

    delivery_company = replacenone(row[37].value)
    delivery_code = replacenone(row[38].value)
    return_delivery_arrive_at = row[39].value
    return_hold_at = row[40].value
    return_delivery_complete_at = row[41].value
    claim_state = '반품'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

path = returnCompleteFile

wb = load_workbook(path)

ws = wb.active

maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=7, max_row=maxRow):
    order_number = replacenone(row[3].value)
    order_number_line = replacenone(row[4].value)
    return_number = None
    channel = '11st'
    security_refund = replacenone(row[2].value)
    security_refund_at = row[7].value
    return_request_at = row[5].value
    return_complete_at = row[6].value
    return_delivery_case = replacenone(row[29].value)
    return_delivery_fees = replaceint(row[28].value)
    return_request_case = replacenone(row[22].value)
    channel_delivery_fees = replaceint(row[30].value)
    return_respons = replacenone(row[31].value)
    payment_case = replacenone(row[32].value)
    return_qty = replaceint(row[14].value)
    refund_state = replacenone(row[1].value)
    product_name = replacenone(row[9].value)
    product_option = replacenone(row[10].value)
    if product_option is not None:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()

    delivery_company = replacenone(row[37].value)
    delivery_code = replacenone(row[38].value)
    return_delivery_arrive_at = row[39].value
    return_hold_at = row[40].value
    return_delivery_complete_at = row[41].value
    claim_state = '반품'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

# 11번가 교환 데이터 수집

driver.get('https://soffice.11st.co.kr/escrow/AuthSellerClaimManager.tmall?method=getClaimList&clm=02&searchVer=02')

driver.find_element_by_xpath('//*[@id="totalClmCountA"]').click()
driver.find_element_by_xpath('//*[@id="sch_box_wrap"]/div[2]/div[2]/div/button[1]').click()
countSleep(1, 3)
driver.find_element_by_xpath('//*[@id="ext-gen6"]/form[1]/div/div[1]/div[6]/div/a').click()
countSleep(1, 3)
returnRequestFile = changeFileToXlsx('claimGoodsList.xls', '11st_return_Request_')

print(returnRequestFile)

path = returnRequestFile

wb = load_workbook(path)

ws = wb.active

maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=7, max_row=maxRow):
    order_number = replacenone(row[2].value)
    order_number_line = replacenone(row[3].value)
    return_number = None
    channel = '11st'
    security_refund = None
    security_refund_at = None
    return_request_at = row[4].value
    return_complete_at = row[5].value
    return_delivery_case = replacenone(row[23].value)
    return_delivery_fees = replaceint(row[20].value)
    return_request_case = replacenone(row[16].value)
    channel_delivery_fees = replaceint(row[21].value)
    return_respons = replacenone(row[23].value)
    payment_case = replacenone(row[24].value)
    return_qty = replaceint(row[10].value)
    refund_state = replacenone(row[1].value)
    product_name = replacenone(row[6].value)
    product_option = replacenone(row[7].value)
    if product_option is not None:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()

    delivery_company = replacenone(row[28].value)
    delivery_code = replacenone(row[29].value)
    return_delivery_arrive_at = row[30].value
    return_hold_at = None
    return_delivery_complete_at = None
    claim_state = '교환'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

driver.quit()

display.stop()

