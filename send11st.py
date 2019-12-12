import smtplib
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from slacker import Slacker
import config
from datetime import date
from datetime import datetime
import os
import dateutil.relativedelta
from openpyxl import load_workbook
import time

makeToday = datetime.now()
makeWeek = datetime.weekday(makeToday)
makeTime = datetime.time(makeToday).strftime('%H:%M')
now = makeToday.strftime("%m-%d_%H-%M-%S")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

fileResults = os.listdir(config.ST_LOGIN['excelPath'])
print(sorted(fileResults, reverse=True))
cancelResultLists = []
stOrderResultLists = []
eabyOrderResultLists = []
stReturnResultLists = []
newReturnResultLsits = []


def findFile(filename, listName):
    if filename == a[0]:
        if a[1] == totalNow:
            listName.append(result)


for fileResult in fileResults:
    result = os.path.join(config.ST_LOGIN['excelPath'], fileResult)
    # result = os.path.abspath(fileResult)
    ext = os.path.split(result)
    a = ext[-1].split("_")
    b = now.split("_")
    findFile('CancelResult', cancelResultLists)
    findFile('11stOrderResult', stOrderResultLists)
    findFile('ebayOrderResult', eabyOrderResultLists)
    findFile('channelReturnResult', stReturnResultLists)
    findFile('channelReturnMissMatch', newReturnResultLsits)

slack = Slacker(config.SLACK_API['token'])


def checkFileToSend(resultList, sendName):
    maxRow = max(resultList)
    wb = load_workbook(maxRow)
    ws = wb.active
    resultRow = str(ws.max_row - 1)
    sendResult = open(maxRow, 'rb')
    sendTitle = sendName + now + '총' + resultRow + '건'
    print(sendResult, sendTitle)
    slack.files.upload(
        file_=sendResult,
        channels=config.SLACK_API['channels'],
        title=sendTitle,
    )

try:
    checkFileToSend(cancelResultLists, 'CancelResult_')
except Exception as ex:
    print(ex)

try:
    checkFileToSend(stOrderResultLists, '11stChannelOrderResult_')
except Exception as ex:
    print(ex)
try:
    checkFileToSend(eabyOrderResultLists, 'ebayChannelOrderResult_')
except Exception as ex:
    print(ex)
try:
    checkFileToSend(stReturnResultLists, 'channelReturnResult_')
except Exception as ex:
    print(ex)
try:
    checkFileToSend(newReturnResultLsits, 'channelReturnMissMatch_')
except Exception as ex:
    print(ex)

# print(cancelResultLists, stOrderResultLists, eabyOrderResultLists)
# maxCancel = max(cancelResultLists)
# maxStOrder = max(stOrderResultLists)
# maxEbayOrder = max(eabyOrderResultLists)
# print(maxCancel, maxStOrder, maxEbayOrder)
# wb = load_workbook(maxCancel)
# ws = wb.active
# cancelRow = str(ws.max_row - 1)
# wb = load_workbook(maxStOrder)
# ws = wb.active
# stOrderRow = str(ws.max_row - 1)
# wb = load_workbook(maxEbayOrder)
# ws = wb.active
# ebayOrderRow = str(ws.max_row - 1)
# cancelSendResult = open(maxCancel, 'rb')
# stOrderSendResult = open(maxStOrder, 'rb')
# ebayOrderSendResult = open(maxEbayOrder, 'rb')
# cancelTitle = 'CancelResult_' + now + '총' + cancelRow + '건'
# stOrderTitle = '11stChannelOrderResult_' + now + '총' + stOrderRow + '건'
# ebayOrderTitle = 'ebayChannelOrderResult_' + now + '총' + ebayOrderRow + '건'
# print(cancelSendResult)
# print(stOrderSendResult)
# print(ebayOrderSendResult)
#
#
# slack.files.upload(
#     file_=cancelSendResult,
#     channels=config.SLACK_API['channels'],
#     title=cancelTitle,
# )
# time.sleep(1)
# slack.files.upload(
#     file_=stOrderSendResult,
#     channels=config.SLACK_API['channels'],
#     title=stOrderTitle
# )
# time.sleep(1)
# slack.files.upload(
#     file_=ebayOrderSendResult,
#     channels=config.SLACK_API['channels'],
#     title=ebayOrderTitle
# )
# server = smtplib.SMTP('smtp.gmail.com', 587)
# server.starttls()
# server.login(config.MAIL_LOGIN['account2'], config.MAIL_LOGIN['password2'])
#
# msg = MIMEBase('multipart', 'mixed')
# msg['Subject'] = '11번가 취소완료 리스트' + totalNow
# msg['from'] = config.MAIL_LOGIN['account2']
# msg['to'] = 'ashyrion@naver.com'
#
# body = f'''
#     11번가 취소 완료 리스트입니다.
#     완료일시 : {totalNow}
#     파일작성시간 : {now}
#     '''
# msg.attach(MIMEText(body, 'plain'))
#
#
# path = config.ST_LOGIN['excelPath'] + os.path.basename(cancelResultLists[0])
# print(path)
# part = MIMEBase('application', 'octet-stream')
# part.set_payload(open(path, 'rb').read())
# encoders.encode_base64(part)
# part.add_header('Content-Disposition', 'attachment;filename=' + path)
# msg.attach(part)
# server.sendmail(config.MAIL_LOGIN['account2'], 'ashyrion@naver.com', msg.as_string())
# server.quit()
