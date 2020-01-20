import os
import re
import time
from datetime import datetime
from datetime import datetime, timedelta
import json
import dateutil.relativedelta
from openpyxl import load_workbook, Workbook
from slacker import Slacker
import requests
import config
import hashlib

# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
weekNow = (makeToday - timedelta(weeks=1)).strftime("%Y-%m-%d")

wb = Workbook()

ws = wb.active

no = 2

# var password = {
#     hash: function (password) {
#         // 모든 비밀번호는 소문자 화
#         password = password.toLowerCase();
#         var passwordHash;
#         CommonAjax.basic({url:'/salt.json', method:'GET', async:false, callbackFunc:function(res) {
#             var salt = res.data.salt;
#             var substrSalt = salt.substr(1,1) + salt.substr(4,1) + salt.substr(8,1) + salt.substr(12,1);
#             passwordHash = $.sha1(substrSalt+$.sha1(password))+substrSalt;
#         }});
#         return passwordHash;
#     }
# }

def hashPassword(salt, wmpPassword):
    password = wmpPassword.lower()
    substrSalt = salt[1:2] + salt[4:5] + salt[8:9] + salt[12:13]
    h = hashlib.sha1(password.encode('utf-8'))
    hHex = h.hexdigest()
    pHash = hashlib.sha1((substrSalt + hHex).encode('utf-8'))
    pHex = pHash.hexdigest()
    return pHex + substrSalt

wb = Workbook()

ws = wb.active

no = 2

with requests.Session() as s:
    getSalt = s.get('https://wpartner.wemakeprice.com/salt.json?_=')
    jsonSalt = json.loads(getSalt.text)
    salt = jsonSalt['data']['salt']
    password = config.WMP_LOGIN['password']

    wpPassword = hashPassword(salt, password)
    print(wpPassword)
    print(config.WMP_LOGIN['id'])

    loginUrl = 'https://wpartner.wemakeprice.com/login.json'
    loginHeader = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json, text/javascript, */*; q=0.01'
    }
    loginData = {
        'userId' : config.WMP_LOGIN['id'],
        'userPassword': wpPassword,
        'userIdSaveYn':''

    }
    loginRequest = s.post(url=loginUrl, data=loginData, headers=loginHeader)
    print(loginRequest)
    productHeaders = {
        'Content-Type': 'application/json;charset=UTF-8'
    }
    productUrl = 'https://wpartner.wemakeprice.com/product/getProdList.json'
    productParams = {
        'schDate': 'chgDate',
        'schStartDate': '2019-10-09',
        'schEndDate': '2020-01-09',
        'schLcateCd': '',
        'schMcateCd': '',
        'schScateCd': '',
        'schDcateCd': '',
        'schProdStatus[0]': 'I',
        'schProdStatus[1]': 'ST',
        'schProdStatus[2]': 'A',
        'schProdStatus[3]': 'S',
        'schProdStatus[4]': 'F',
        'schType': 'prodNm',
        'schValue': '',
        'schDispYn': '',
        'dispOnlyDealYn': '',
        'epYn': '',
        'bundleKind': '',
        'shipType': '',
        'shipMethod': '',
        'schMallType': '',
        '_': '1578563721845'
    }
    producRequest = s.get(url=productUrl, params=productParams)
    print(producRequest.url)
    productJson = json.loads(producRequest.text)


    for product in productJson['list']:
        productNo = product['prodNo']
        optionList = s.get(f'https://wpartner.wemakeprice.com/product/getProdOptionItemList.json?prodNo={productNo}&_=1578564568880')
        optionJson = json.loads(optionList.text)
        print(optionJson)
        productName = product['prodNm']
        productPrice = product['salePrice']
        brandName = product['brandNm']
        shipType = product['shipTypeNm']
        productDisplay = product['dispYnNm']

        for option in optionJson:
            optionFirst = option['opt1Val']
            optionsecond = option['opt2Val']
            saleStatusNm = option['saleStatusNm'] #판매상태
            dispYnNm = option['dispYnNm'] #노출여부

            ws.cell(row=1, column=1).value = '브랜드'
            ws.cell(row=1, column=2).value = '상품명'
            ws.cell(row=1, column=3).value = '상품가격'
            ws.cell(row=1, column=4).value = '배송타입'
            ws.cell(row=1, column=5).value = '노출여부'
            ws.cell(row=1, column=6).value = '옵션1'
            ws.cell(row=1, column=7).value = '옵션2'
            ws.cell(row=1, column=8).value = '판매상태'
            ws.cell(row=1, column=9).value = '노출여부'

            ws.cell(row=no, column=1).value = brandName
            ws.cell(row=no, column=2).value = productName
            ws.cell(row=no, column=3).value = productPrice
            ws.cell(row=no, column=4).value = shipType
            ws.cell(row=no, column=5).value = productDisplay
            ws.cell(row=no, column=6).value = optionFirst
            ws.cell(row=no, column=7).value = optionsecond
            ws.cell(row=no, column=8).value = saleStatusNm
            ws.cell(row=no, column=9).value = dispYnNm

            no += 1


result = config.ST_LOGIN['excelPath'] + 'wmpProductList' +  "_" + now + '.xlsx'
print(result)
wb.save(result)
wb.close()
    




# 상품조회
# https://wpartner.wemakeprice.com/product/getProdList.json?schDate=chgDate&schStartDate=2019-10-08&schEndDate=2020-01-08&schLcateCd=&schMcateCd=&schScateCd=&schDcateCd=&schProdStatus%5B0%5D=I&schProdStatus%5B1%5D=ST&schProdStatus%5B2%5D=A&schProdStatus%5B3%5D=S&schProdStatus%5B4%5D=F&schType=prodNm&schValue=&schDispYn=&dispOnlyDealYn=&epYn=&bundleKind=&shipType=&shipMethod=&schMallType=&_=1578478038452
# 상품 상세
# https://wpartner.wemakeprice.com/product/getProdOptionItemList.json?prodNo=476128448&_=1578477924174


