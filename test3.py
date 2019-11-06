from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import date
from datetime import datetime
import pyexcel as p
import pymysql
from openpyxl import load_workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook

def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


options = Options()
# options.add_argument('--headless')
# options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/Users/daegukim/py_option/chromedriver', options=options)

# driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
# params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "/path/to/download/dir"}}
# command_result = driver.execute("send_command", params)

driver.get('https://www.esmplus.com/Member/SignIn/LogOn')
driver.find_element_by_xpath('//*[@id="rdoSiteSelect" and @value="GMKT"]').click()
driver.find_element_by_xpath('//*[@id="SiteId"]').send_keys(config.ESM_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="SitePassword"]').send_keys(config.ESM_LOGIN['password'])
driver.find_element_by_xpath('//*[@id="btnSiteLogOn"]').click()
time.sleep(3)
windowLists = driver.window_handles
for windowList in windowLists[1:]:
    driver.switch_to.window(driver.window_handles[-1])
    driver.close()

driver.switch_to.window(driver.window_handles[0])
driver.get('http://www.esmplus.com/Member/CustomerService/FindCustomer?menuCode=TDM144')
driver.find_element_by_xpath('//*[@id="contents"]/div/div[2]/ul/li[1]/span/span[5]/a').click()
driver.find_element_by_xpath('//*[@id="btnSearch"]').click()
driver.find_element_by_xpath('//*[@id="contents"]/div/div[3]/a').click()
time.sleep(10)

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")


def changeFileToXlsx(originalName, resultName):
    stOriExcel = config.ST_LOGIN['excelPath'] + originalName
    stResultExcel = config.ST_LOGIN['excelPath'] + resultName + now + '.xls'
    stResultXlsx = config.ST_LOGIN['excelPath'] + resultName + now + '.xlsx'

    os.rename(stOriExcel, stResultExcel)

    p.save_book_as(file_name=stResultExcel, dest_file_name=stResultXlsx)
    return stResultXlsx


gmarketFile = changeFileToXlsx('findcustomer_' + totalNow +'.xls', 'gmarket_state_')

path = gmarketFile

wb = load_workbook(path)

ws = wb.active
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

delOrder = '''DELETE From `excel`.`channel_order` where channel in ("gmarket", "auction")'''
cursor.execute(delOrder)
orderSql = '''
        INSERT INTO `excel`.`channel_order` (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE state = %s
        '''
maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=2):
    state = replacenone(row[12].value)
    channel_order_number = replacenone(row[2].value)
    channel_order_list = replacenone(row[3].value)
    product_name = replacenone(row[4].value)
    product_option = None
    quantity = replaceint(row[6].value)
    payment_at = replacenone(row[7].value)
    if channel_order_list[0] == 'B':
        channel = 'auction'
    else:
        channel = 'gmarket'

    orderValues = (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel,
        state
    )
    cursor.execute(orderSql, orderValues)
    print(orderSql, orderValues)
db.close()




