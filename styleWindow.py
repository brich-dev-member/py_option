import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

# 스타일윈도우크롤링
# UV / ORDER / REVIEW
viewType = "UV"
# MENS2 / ITGIRL2 / WOMEN
gender = "MENS2"
# 전체 10000031
cateId = "10000031"
# DAILY / WEEKLY
dayWeek = "WEEKLY"
url = "https://m.shopping.naver.com/v1/best/window/products?aggregateType=" + viewType + "&gender=" + gender + "&bestCategoryId=" + cateId + "&aggregatePeriod=" + dayWeek
print(url)
response = requests.get(url)
results = response.json()
print(results)
no = 1


def subInfo(shopUrl):
    req = requests.get(shopUrl)
    html = req.text

    soup = BeautifulSoup(html, 'html.parser')
    info = soup.select('script')
    a = str(info[0])
    b = a.lstrip('<script>window.__PRELOADED_STATE__=').rstrip('</script>')
    c = json.loads(b)

    try:
        zzim = c['storeKeep']['A']['zzimCount']
        addressInfo = c['store']['A']['channel']['businessAddressInfo']['fullAddressInfo']
        telInfo = c['store']['A']['channel']['contactInfo']['telNo']['formattedNumber']
    except KeyError:
        telInfo = None

    print(zzim, addressInfo, telInfo)
    return [zzim, addressInfo, telInfo]


for result in results:
    no += 1
    productRank = result['rank']
    productName = result['name']
    productPrice = result['mobileDiscountPrice']
    productUrl = "https://m.shopping.naver.com" + result['productUrlWithTrCode']
    productImg = result['imageUrl']
    providerName = result['channelName']
    providerUrl = "https://m.shopping.naver.com/style/style/stores/" + result['channelId'] + "/about"
    print(providerUrl)
    moreInfo = subInfo(providerUrl)

    ws.cell(row=1, column=1).value = '순위'
    ws.cell(row=1, column=2).value = '상품명'
    ws.cell(row=1, column=3).value = '상품가격'
    ws.cell(row=1, column=4).value = '상품URL'
    ws.cell(row=1, column=5).value = '상품IMG'
    ws.cell(row=1, column=6).value = '판매자명'
    ws.cell(row=1, column=7).value = '판매자URL'
    ws.cell(row=1, column=8).value = '스토어찜수'
    ws.cell(row=1, column=9).value = '주소'
    ws.cell(row=1, column=10).value = '전화번호'

    ws.cell(row=no, column=1).value = productRank
    ws.cell(row=no, column=2).value = productName
    ws.cell(row=no, column=3).value = productPrice
    ws.cell(row=no, column=4).value = productUrl
    ws.cell(row=no, column=5).value = productImg
    ws.cell(row=no, column=6).value = providerName
    ws.cell(row=no, column=7).value = providerUrl
    ws.cell(row=no, column=8).value = moreInfo[0]
    ws.cell(row=no, column=9).value = moreInfo[1]
    ws.cell(row=no, column=10).value = moreInfo[2]

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
wb.save("SW_" + now + "_" + viewType + "_" + gender + "_" + dayWeek + ".xlsx")