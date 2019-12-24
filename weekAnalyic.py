import pymysql
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils.cell import get_column_letter
import config

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

wb = Workbook()

ws = wb.active
wa = wb.create_sheet('상품등록셀러')

month = 12
year = 2019

weekSql = f'''
        select week from sell where month = {month} and year= {year} group by week;
        '''
cursor.execute(weekSql)
weeks = cursor.fetchall()

weekList = list()
for week in weeks:
    a = int(str(week).replace("(", "").replace(",", "").replace(")", ""))
    weekList.append(a)

listSql = f'''
        select distinct(`provider_name`) FROM `sell` where month = {month} and year = {year} or
        week in ({weekList[0]},{weekList[1]},{weekList[2]},{weekList[3]})
        group by `provider_name`
        order by sum(total_amount) desc;
        '''
# ,{weekList[2]},{weekList[3]},{weekList[4]}
cursor.execute(listSql)
providerLists = cursor.fetchall()

no = 2

for providerList in providerLists:
    provider_name = providerList[0]

    amountSql = f'''
        select sum(`total_amount`),
        (select sum(`total_amount`) from sell where week={weekList[0]} and provider_name = '{provider_name}'),
        (select sum(`total_amount`) from sell where week={weekList[1]} and provider_name = '{provider_name}'),
        (select sum(`total_amount`) from sell where week={weekList[2]} and provider_name = '{provider_name}'),
        (select sum(`total_amount`) from sell where week={weekList[3]} and provider_name = '{provider_name}')

        from sell where provider_name = '{provider_name}' and month = {month};
    '''
    #         (select sum(`total_amount`) from sell where week={weekList[2]} and provider_name = '{provider_name}'),
    #         (select sum(`total_amount`) from sell where week={weekList[3]} and provider_name = '{provider_name}'),
    #         (select sum(`total_amount`) from sell where week={weekList[4]} and provider_name = '{provider_name}')
    cursor.execute(amountSql)
    providerAmountLists = cursor.fetchall()

    amount1st = providerAmountLists[0][0]
    amount2nd = providerAmountLists[0][1]
    amount3rd = providerAmountLists[0][2]
    amount4st = providerAmountLists[0][3]
    amount5st = providerAmountLists[0][4]
    # amount6st = providerAmountLists[0][5]

    ws.cell(row=1, column=1).value = '셀러명'
    ws.cell(row=1, column=2).value = f'{month}월 판매'
    ws.cell(row=1, column=3).value = f'{weekList[0]}주차 판매'
    ws.cell(row=1, column=4).value = f'{weekList[1]}주차 판매'
    ws.cell(row=1, column=5).value = f'{weekList[2]}주차 판매'
    ws.cell(row=1, column=6).value = f'{weekList[3]}주차 판매'
    # ws.cell(row=1, column=7).value = f'{weekList[4]}주차 판매'

    ws.cell(row=no, column=1).value = provider_name
    ws.cell(row=no, column=2).value = amount1st
    ws.cell(row=no, column=3).value = amount2nd
    ws.cell(row=no, column=4).value = amount3rd
    ws.cell(row=no, column=5).value = amount4st
    ws.cell(row=no, column=6).value = amount5st
    # ws.cell(row=no, column=7).value = amount6st

    no += 1

productListSql = f'''
        select distinct(`provider_number`), `provider_name` FROM `product` where month = {month} and year = {year} or
        week in ({weekList[0]},{weekList[1]},{weekList[2]},{weekList[3]})
        group by `provider_number`
        order by count(product_number) desc;
        '''
# ,{weekList[2]},{weekList[3]},{weekList[4]}
cursor.execute(productListSql)
productProviderLists = cursor.fetchall()

pno = 2


def noZeroCount(text):
    if text is 0:
        return None
    else:
        return int(text)


for productProviderList in productProviderLists:
    product_provider_number = productProviderList[0]
    product_provider_name = productProviderList[1]

    productSql = f'''
            select count(`product_number`),
            (select count(`product_number`) from product where week={weekList[0]} and provider_number = '{product_provider_number}'),
            (select count(`product_number`) from product where week={weekList[1]} and provider_number = '{product_provider_number}'),
            (select count(`product_number`) from product where week={weekList[2]} and provider_number = '{product_provider_number}'),
            (select count(`product_number`) from product where week={weekList[3]} and provider_number = '{product_provider_number}')
            from product where provider_number = '{product_provider_number}' and month = {month};
        '''

    # (select count(`product_number`) from product where week={weekList[2]} and provider_name = '{product_provider_name}'),
    # (select count(`product_number`) from product where week={weekList[3]} and provider_name = '{product_provider_name}'),
    # (select count(`product_number`) from product where week={weekList[4]} and provider_name = '{product_provider_name}')

    cursor.execute(productSql)
    providerProductLists = cursor.fetchall()

    productCount1st = noZeroCount(providerProductLists[0][0])
    productCount2nd = noZeroCount(providerProductLists[0][1])
    productCount3rd = noZeroCount(providerProductLists[0][2])
    productCount4st = noZeroCount(providerProductLists[0][3])
    productCount5st = noZeroCount(providerProductLists[0][4])
    # productCount6st = noZeroCount(providerProductLists[0][5])

    wa.cell(row=1, column=1).value = '셀러명'
    wa.cell(row=1, column=2).value = f'{month}월 등록수'
    wa.cell(row=1, column=3).value = f'{weekList[0]}주차 등록수'
    wa.cell(row=1, column=4).value = f'{weekList[1]}주차 등록수'
    wa.cell(row=1, column=5).value = f'{weekList[2]}주차 등록수'
    wa.cell(row=1, column=6).value = f'{weekList[3]}주차 등록수'
    # wa.cell(row=1, column=7).value = f'{weekList[4]}주차 등록수'

    wa.cell(row=pno, column=1).value = product_provider_name
    wa.cell(row=pno, column=2).value = productCount1st
    wa.cell(row=pno, column=3).value = productCount2nd
    wa.cell(row=pno, column=4).value = productCount3rd
    wa.cell(row=pno, column=5).value = productCount4st
    wa.cell(row=pno, column=6).value = productCount5st
    # wa.cell(row=pno, column=7).value = productCount6st

    pno += 1

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
result = '2019_상품등록_판매_지표' + now + '.xlsx'
print(result)
wb.save(result)
cursor.close()
db.close()
