from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from glob import glob
from datetime import date
from datetime import datetime
import pyexcel as p
import pymysql
from openpyxl import load_workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook
from slacker import Slacker
import requests
import json
from pyvirtualdisplay import Display

display = Display(visible=0, size=(1200, 900))
display.start()


def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()

def replaceTextToInt(text):
    if text is None:
        return
    else:
        text = str(text)
        a = text.split(",")
        b = "".join(a)
        text = int(b)
        return text


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)


def findSelect(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# f코드 정규
rex = re.compile('_F[0-9]+')
rexSpace = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
nowTime = makeToday.strftime("%H_%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

def changeFileToXlsx(originalName, resultName):
    stOriExcel = config.ST_LOGIN['excelPath'] + originalName
    stResultExcel = config.ST_LOGIN['excelPath'] + resultName + now + '.xls'
    stResultXlsx = config.ST_LOGIN['excelPath'] + resultName + now + '.xlsx'

    os.rename(stOriExcel, stResultExcel)

    p.save_book_as(file_name=stResultExcel, dest_file_name=stResultXlsx)
    return stResultXlsx

# 셀레니움 셋
options = Options()
# options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

# 지마켓 시작
driver.switch_to.window(driver.window_handles[0])
driver.get('https://www.esmplus.com/Member/SignIn/LogOn')
driver.find_element_by_xpath('//*[@id="rdoSiteSelect" and @value="GMKT"]').click()
driver.find_element_by_xpath('//*[@id="SiteId"]').send_keys(config.ESM_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="SitePassword"]').send_keys(config.ESM_LOGIN['password'])
driver.find_element_by_xpath('//*[@id="btnSiteLogOn"]').click()
countSleep(1, 3)
print('로그인')

windowLists = driver.window_handles
for windowList in windowLists[1:]:
    driver.switch_to.window(driver.window_handles[-1])
    driver.close()

driver.switch_to.window(driver.window_handles[0])
countSleep(1, 3)
print('팝업 제거')
driver.get('https://www.esmplus.com/Escrow/Claim/ReturnRequestManagement?menuCode=TDM118')
countSleep(1, 3)
print('메뉴 이동')
driver.find_element_by_xpath('//*[@id="divSearch"]/div/div[1]/table/tbody/tr[1]/td[1]/span[5]/a').click()
driver.find_element_by_xpath('//*[@id="btnSearch"]').click()
countSleep(1, 5)
print('검색 결과')
driver.find_element_by_xpath('//*[@id="contents"]/div[1]/form/div[4]/span/span/a').click()
countSleep(1, 5)
print('엑셀 다운')
alert = driver.switch_to.alert
alert.accept()
countSleep(1, 10)
print('엑셀 다운 완료')

findEbayFile = glob(config.ST_LOGIN['excelPath'] + "ReturnRequest_*.xls")
print(findEbayFile)
fileResult = max(findEbayFile).split("/")
print(fileResult)
ebayReturnFileName = fileResult[-1]
print(ebayReturnFileName)

ebayResult = changeFileToXlsx(ebayReturnFileName, 'ebayRetuneRequst_')

print('파일 변경 완료')

channelSql = '''
    INSERT INTO `bflow`.`channel_returns` (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at
        ) VALUES (
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s
       )
        ON DUPLICATE KEY UPDATE 
        security_refund = %s,
        security_refund_at = %s,
        return_request_at = %s,
        return_complete_at = %s,
        return_delivery_case = %s,
        return_delivery_fees = %s,
        return_request_case = %s,
        channel_delivery_fees = %s,
        return_respons = %s,
        payment_case = %s,
        return_qty = %s,
        refund_state = %s,
        delivery_company = %s,
        delivery_code = %s,
        return_delivery_arrive_at = %s,
        return_hold_at = %s,
        return_delivery_complete_at = %s,
        fcode = %s,
        claim_state = %s
    '''

path = ebayResult

wb = load_workbook(path)

ws = wb.active

maxRow = ws.max_row

for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=maxRow )):
    print(idx, "/" , maxRow)
    order_number = replacenone(row[3].value)
    order_number_line = None
    return_number = None
    mallId = replacenone(row[1].value)
    searchMall = mallId.split()
    if searchMall[1] == '(brich_07)':
        channel = 'gmarket'
    elif searchMall[1] == '(brich)':
        channel = 'auction'
    elif searchMall[1] == '(brichmall)':
        channel = 'g9'
    security = replacenone(row[0].value)
    if security == 'Y':
        security_refund = '빠른환불'
    else:
        security_refund = None
    security_refund_at = None
    return_request_at = row[10].value
    return_complete_at = row[15].value
    return_delivery_case = replacenone(row[27].value)
    return_delivery_fees = replaceTextToInt(row[8].value)
    return_request_case = None
    channel_delivery_fees = None
    return_respons = replacenone(row[5].value)
    payment_case = replacenone(row[9].value)
    return_qty = replaceint(row[21].value)
    refund_state = replacenone(row[2].value)
    product_name = replacenone(row[17].value)
    product_option = replacenone(row[22].value)
    detailCode = replacenone(row[49].value)
    if detailCode is not None:
        changeCode = "_" + detailCode + "_"
        makeCode = rex.search(changeCode)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()

    delivery_company = replacenone(row[29].value)
    delivery_code = replacenone(row[30].value)
    return_delivery_arrive_at = row[13].value
    return_hold_at = row[14].value
    return_delivery_complete_at = None
    claim_state = '반품'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

db.close()
display.stop()
driver.quit()
# returnUrl = 'https://www.esmplus.com/Escrow/Claim/ReturnManagementSearch'
# returnParms = {
#     'page': '1',
#     'limit': '200',
#     'siteGbn': '1',
#     'searchAccount': 'TA^278815',
#     'searchDateType': 'ODD',
#     'searchSDT': '2019-08-27',
#     'searchEDT': '2019-11-27',
#     'searchType': 'RR',
#     'searchKey': 'ON',
#     'searchKeyword':'', 
#     'orderByType': '',
#     'excelInfo': '',
#     'searchStatus': 'RR',
#     'searchAllYn': 'N',
#     'tabGbn': '1',
#     'SortFeild': 'PayDate',
#     'SortType': 'Desc',
#     'start': '0',
#     'searchDistrType': 'AL',
#     'searchRewardStatus': 'NN',
#     'searchFastRefundYn': '',
# }
# returnHeader = {'Content-Type' : 'application/x-www-form-urlencoded',
#                 'Referer': 'https://www.esmplus.com/Escrow/Claim/ReturnRequestManagement?menuCode=TDM118',
#                 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# returnCookies = {'ESM_AUTH': 'B044D833077F80351AC6DCE532CFB728D47244CB65B62B3F34030FF8E03C1AE2C23398EB38289CC329DFD6A993367ED223EF303D71393DFD43995312D0446B28992A1E68EF0FBC13874B36E64981384B128058B1CE46A65998FD479E669941142039F898C85EB8100AD2FCADD11D721598FC6D45'}
# returnRequest = requests.get(returnUrl, params=returnParms, headers=returnHeader)

# returnToJson = returnRequest.json
# print(returnRequest.url)
# print(returnRequest.request)
# print(returnRequest.text)
# print(returnRequest.content)
# print(returnRequest.request)
# print(returnToJson)

# https://www.esmplus.com/Escrow/Claim/ReturnRequestManagement?menuCode=TDM118