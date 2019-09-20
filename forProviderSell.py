import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tqdm import tqdm

Tk().withdraw()
filename = askopenfilename()

path = filename

wb = load_workbook(path)

ws = wb.active

db = pymysql.connect(host='127.0.0.1', user='root', password='root', db='excel', charset='utf8')
cursor = db.cursor()

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)


def replacedate(text):
    if text is None:
        return
    else:
        text = str(text)[0:10]
        return text.strip()


no = 2


for row in rows:
    provider_number = row[3].value
    month = 8
    # 9월 시작 주 35
    week = 36

    sql = f'''
        select sum(`total_amount`),
        (select sum(`total_amount`) from sell where week={week} and provider_number = '{provider_number}'),
        (select sum(`total_amount`) from sell where week={week} + 1 and provider_number = '{provider_number}'),
        (select sum(`total_amount`) from sell where week={week} + 2 and provider_number = '{provider_number}'),
        (select sum(`total_amount`) from sell where week={week} + 3 and provider_number = '{provider_number}')
        from sell where provider_number = '{provider_number}' and month = {month};    
    '''
    cursor.execute(sql)
    results = cursor.fetchall()

    for result in results:
        ws.cell(row=no, column=8).value = result[0]
        ws.cell(row=no, column=9).value = result[1]
        ws.cell(row=no, column=10).value = result[2]
        ws.cell(row=no, column=11).value = result[3]
        ws.cell(row=no, column=12).value = result[4]
        ws.cell(row=no, column=13).value = f'''=sum(I{no}:L{no})'''

        no += 1

for col in ws.columns:
    max_length = 0
    columnIndex = col[0].column
    column = get_column_letter(columnIndex)
    for cell in col:
        if max_length < len(str(cell.value)) < 30:
            max_length = len(str(cell.value))
        else:
            pass
    ws.column_dimensions[column].width = (max_length + 1) * 1.2

result = filename
wb.save(result)

