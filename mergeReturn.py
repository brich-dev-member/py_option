import config
from openpyxl import Workbook
import pymysql
from datetime import date
from datetime import datetime, timedelta
import dateutil.relativedelta
from dateutil.parser import parse
from reqStatus import requestStaus, requestStausChannel
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
weekNow = (makeToday - timedelta(weeks=1)).strftime("%Y-%m-%d")


# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

# 클레임 상태값 필드가 변경됨. refund_state 처리완료, 수거중 
# claim_state 교환:1 , 반품:1 
mergeSql = '''
            select `order_number`, `fcode`, `return_request_at`, `refund_state`, `return_delivery_fees`,
            `return_respons`, `payment_case`, `delivery_company`, `delivery_code`, `return_delivery_arrive_at`,
            `return_delivery_complete_at`, `channel`
            from `channel_returns`;
            '''
cursor.execute(mergeSql)
returnMerges = cursor.fetchall()

wb = Workbook()

ws = wb.active

no = 2

for returnData in returnMerges:
    channel_order_number = returnData[0]
    fcode = returnData[1]
    channel_return_request_at = returnData[2]
    refund_state = returnData[3]
    return_delivery_fees = returnData[4]
    return_respons = returnData[5]
    payment_case = returnData[6]
    delivery_company = returnData[7]
    delivery_code = returnData[8]
    return_delivery_arrive_at = returnData[9]
    return_delivery_complete_at = returnData[10]
    channel = returnData[11]
    if channel == 'gmarket':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    elif channel == 'auction':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    elif channel == 'g9':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    else:
        bflowStatus = requestStaus(channel_order_number, fcode)
        print(bflowStatus['message'])
        
    if bflowStatus['success'] is True:
        product_order_number = bflowStatus['message']['orderItemOptionId']
        channel = bflowStatus['message']['channel']
        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

            return_number = bflowStatus['message']['claims'][0]['claimId']
            return_request_at = bflowStatus['message']['claims'][0]['createdAt']
            if claimType is None:
                claim_state = None
            elif claimType == 'cancel':
                claimType = '취소'
                claim_state = claimType + ":" + claimStatus
            elif claimType == 'return':
                claimType = '반품'
                claim_state = claimType + ":" + claimStatus
            elif claimType == 'exchange':
                claimType = '교환'
                claim_state = claimType + ":" + claimStatus
            else:
                claim_state = claimType + ":" + claimStatus
        else:
            claim_state = bflowStatus['message']['status']
            return_number = None
            return_request_at = None
    

        print(claim_state, refund_state)

        

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '반품번호'
        ws.cell(row=1, column=3).value = '채널 주문번호'
        ws.cell(row=1, column=4).value = '채널'
        ws.cell(row=1, column=5).value = '브리치 반품신청일'
        ws.cell(row=1, column=6).value = '브리치 반품상태'
        ws.cell(row=1, column=7).value = '채널 반품신청일'
        ws.cell(row=1, column=8).value = '채널 상태'
        ws.cell(row=1, column=9).value = '반품 배송비'
        ws.cell(row=1, column=10).value = '반품 책임자'
        ws.cell(row=1, column=11).value = '반품비용 처리'
        ws.cell(row=1, column=12).value = '반품택배사'
        ws.cell(row=1, column=13).value = '송장번호'
        ws.cell(row=1, column=14).value = '반품도착일'
        ws.cell(row=1, column=15).value = '반품완료일'

    # 채널 상태값이 교환거부 / 교환철회 / 반품보류 / 반품철회 
    # 11번가는 반품취소철회 메뉴 데이터
    # 11번가 반품 비용처리 payment_case = 박스에동봉 / 판매자에게 직접송급은 별도의 리스트 관리
        lastWeek = parse(weekNow)
        refundDate = channel_return_request_at
        print(lastWeek.date(), refundDate.date())



        if claim_state == '반품:수거중' and refund_state == '반품신청' or refund_state == '반품요청':
            if refundDate.date() > lastWeek.date():
                pass
            else:
                ws.cell(row=no, column=1).value = product_order_number
                ws.cell(row=no, column=2).value = return_number
                ws.cell(row=no, column=3).value = channel_order_number
                ws.cell(row=no, column=4).value = channel
                ws.cell(row=no, column=5).value = return_request_at
                ws.cell(row=no, column=6).value = claim_state
                ws.cell(row=no, column=7).value = channel_return_request_at
                ws.cell(row=no, column=8).value = refund_state
                ws.cell(row=no, column=9).value = return_delivery_fees
                ws.cell(row=no, column=10).value = return_respons
                ws.cell(row=no, column=11).value = payment_case
                ws.cell(row=no, column=12).value = delivery_company
                ws.cell(row=no, column=13).value = delivery_code
                ws.cell(row=no, column=14).value = return_delivery_arrive_at
                ws.cell(row=no, column=15).value = return_delivery_complete_at

                no += 1
        elif claim_state == '교환:수거중' and refund_state == '교환신청' or refund_state == '교환요청':
            if refundDate.date() > lastWeek.date():
                pass
            else:
                ws.cell(row=no, column=1).value = product_order_number
                ws.cell(row=no, column=2).value = return_number
                ws.cell(row=no, column=3).value = channel_order_number
                ws.cell(row=no, column=4).value = channel
                ws.cell(row=no, column=5).value = return_request_at
                ws.cell(row=no, column=6).value = claim_state
                ws.cell(row=no, column=7).value = channel_return_request_at
                ws.cell(row=no, column=8).value = refund_state
                ws.cell(row=no, column=9).value = return_delivery_fees
                ws.cell(row=no, column=10).value = return_respons
                ws.cell(row=no, column=11).value = payment_case
                ws.cell(row=no, column=12).value = delivery_company
                ws.cell(row=no, column=13).value = delivery_code
                ws.cell(row=no, column=14).value = return_delivery_arrive_at
                ws.cell(row=no, column=15).value = return_delivery_complete_at

                no += 1
        elif claim_state == '반품:처리완료' and refund_state == '환불승인완료' or refund_state == '반품완료':
            pass
        elif channel == 'wemakprice' and refund_state == '반품철회' or refund_state == '교환거부' or refund_state == '교환철회':
            ws.cell(row=no, column=1).value = product_order_number
            ws.cell(row=no, column=2).value = return_number
            ws.cell(row=no, column=3).value = channel_order_number
            ws.cell(row=no, column=4).value = channel
            ws.cell(row=no, column=5).value = return_request_at
            ws.cell(row=no, column=6).value = claim_state
            ws.cell(row=no, column=7).value = channel_return_request_at
            ws.cell(row=no, column=8).value = refund_state
            ws.cell(row=no, column=9).value = return_delivery_fees
            ws.cell(row=no, column=10).value = return_respons
            ws.cell(row=no, column=11).value = payment_case
            ws.cell(row=no, column=12).value = delivery_company
            ws.cell(row=no, column=13).value = delivery_code
            ws.cell(row=no, column=14).value = return_delivery_arrive_at
            ws.cell(row=no, column=15).value = return_delivery_complete_at

            no += 1
        else:
            ws.cell(row=no, column=1).value = product_order_number
            ws.cell(row=no, column=2).value = return_number
            ws.cell(row=no, column=3).value = channel_order_number
            ws.cell(row=no, column=4).value = channel
            ws.cell(row=no, column=5).value = return_request_at
            ws.cell(row=no, column=6).value = claim_state
            ws.cell(row=no, column=7).value = channel_return_request_at
            ws.cell(row=no, column=8).value = refund_state
            ws.cell(row=no, column=9).value = return_delivery_fees
            ws.cell(row=no, column=10).value = return_respons
            ws.cell(row=no, column=11).value = payment_case
            ws.cell(row=no, column=12).value = delivery_company
            ws.cell(row=no, column=13).value = delivery_code
            ws.cell(row=no, column=14).value = return_delivery_arrive_at
            ws.cell(row=no, column=15).value = return_delivery_complete_at

            no += 1
    elif bflowStatus['success'] is False:
        pass
result = config.ST_LOGIN['excelPath'] + 'channelReturnResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()
db.close()
print(result)