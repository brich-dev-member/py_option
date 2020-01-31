#
# 11번가에서 취소 목록을 불러와서 별도의 엑셀파일로 정리 후, 슬랙으로 던져주는 코드
#
# 1. 외부채널: 취소신청 / 비플로우: 결제후 출고전 단계인 값은 취소요청으로 변경
# 2. 비플로우 취소요청이 수집이후 24시간 경과 17시가 되는경우에는 일괄 취소완료로 변경
# 3. 비플로우 : 취소완료 -> 환불완료 / 외부채널 : 취소, 환불완료로 변경
#

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import datetime
import pyexcel as p
import pymysql
from openpyxl import load_workbook
import dateutil.relativedelta
import re
from openpyxl import Workbook
from pyvirtualdisplay import Display
from reqStatus import requestStaus, requestStausChannel

display = Display(visible=0, size=(1200, 900))
display.start()


def replace_date(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replace_none(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replace_int(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


rex = re.compile('_F[0-9]+')

# 날짜 관련
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")


def change_file_to_xlsx(original_name, result_name):
    st_ori_excel = config.ST_LOGIN['excelPath'] + original_name
    st_result_excel = config.ST_LOGIN['excelPath'] + result_name + now + '.xls'
    st_result_xlsx = config.ST_LOGIN['excelPath'] + result_name + now + '.xlsx'

    os.rename(st_ori_excel, st_result_excel)

    p.save_book_as(file_name=st_result_excel, dest_file_name=st_result_xlsx)
    os.remove(st_result_excel)
    return st_result_xlsx


def find_select(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)


options = Options()

options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver: WebDriver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

driver.get('https://login.11st.co.kr/auth/front/selleroffice/login.tmall')

driver.find_element_by_id('user-id').send_keys(config.ST_LOGIN['id'])
driver.find_element_by_id('passWord').send_keys(config.ST_LOGIN['password'])
driver.find_element_by_xpath('/html/body/div/form[1]/fieldset/button').click()
time.sleep(5)
print('login')
driver.get('https://soffice.11st.co.kr/escrow/OrderCancelManageList2nd.tmall')
time.sleep(5)

find_select('//select[@id="key"]', '02')
find_select('//select[@id="shDateType"]', '07')
find_select('//select[@id="sltDuration"]', 'TODAY')
time.sleep(2)
driver.find_element_by_xpath('//*[@id="search_area"]/form/div[1]/div[2]/div[2]/div[2]/div/button[1]').click()
time.sleep(3)

driver.find_element_by_xpath('//*[@id="search_area"]/form/div[3]/div/a').click()
time.sleep(2)
driver.get('https://soffice.11st.co.kr/escrow/OrderingLogistics.tmall')
time.sleep(2)
findUp = driver.find_element_by_id('goDlvTmpltPopup')
driver.switch_to.frame(findUp.find_element_by_tag_name('iframe'))
driver.find_element_by_xpath('//*[@id="ext-gen6"]/div/button').click()
time.sleep(1)
driver.switch_to.default_content()
driver.find_element_by_xpath('//*[@id="order_good_301"]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="searchform"]/div/div[1]/div[6]/div/a[2]').click()
time.sleep(4)
print(driver.window_handles)
driver.switch_to.window(driver.window_handles[1])
driver.find_element_by_xpath('/html/body/div/div[2]/div[4]/div/a[1]').click()
time.sleep(60)
driver.close()

# 취소 신청, 완료 목록 -> 둘 다 필요함
cancelFile = change_file_to_xlsx('39731068_sellListlistType.xls', '11st_Cancel_')
# 배송준비 목록 - 외부채널에서 배송준비중인데 우리쪽에서 취소 또는 클레임을 찾으려고... - 발송처리할 내역
logiFile = change_file_to_xlsx('39731068_logistics.xls', '11st_logi_')

path = cancelFile

wb = load_workbook(path)

ws = wb.active

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

delOrder = '''DELETE From  `bflow`.`11st_cancel`'''
cursor.execute(delOrder)
sql = '''INSERT INTO `bflow`.`11st_cancel` (
        state,
        channel_order_number,
        channel_order_list,
        claim_request,
        claim_complete,
        product_name,
        product_option,
        quantity,
        order_amount,
        cancel_reason,
        cancel_detail_reason,
        cancel_response,
        add_delivery_fees,
        cancel_complete_date,
        payment_at,
        product_amount,
        product_option_amount,
        cancel_complete_user,
        fcode
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE state = %s, claim_request = %s, claim_complete = %s, cancel_reason = %s,
        cancel_detail_reason = %s, cancel_response =%s, add_delivery_fees = %s, cancel_complete_date = %s,
        cancel_complete_user = %s, fcode = %s
         '''
maxRow = ws.max_row - 2
print(maxRow)

for row in ws.iter_rows(min_row=7, max_row=maxRow):
    state = replace_none(row[1].value)
    channel_order_number = replace_none(row[2].value)
    channel_order_list = replace_int(row[3].value)
    claim_request = row[4].value
    claim_complete = row[5].value
    product_name = replace_none(row[6].value)
    product_option = replace_none(row[7].value)
    quantity = replace_int(row[8].value)
    order_amount = replace_int(row[13].value)
    cancel_reason = replace_none(row[16].value)
    cancel_detail_reason = replace_none(row[17].value)
    cancel_response = replace_none(row[18].value)
    add_delivery_fees = replace_int(row[19].value)
    cancel_complete_date = replace_date(row[20].value)
    payment_at = row[31].value
    product_amount = replace_int(row[33].value)
    product_option_amount = replace_int(row[34].value)
    cancel_complete_user = replace_none(row[37].value)
    if product_option is None:
        fcode = None
    else:
        try:
            fcode = rex.search(product_option).group()
        except Exception as ex:
            fcode = None
            print(product_option)
            print(ex)

    values = (
        state,
        channel_order_number,
        channel_order_list,
        claim_request,
        claim_complete,
        product_name,
        product_option,
        quantity,
        order_amount,
        cancel_reason,
        cancel_detail_reason,
        cancel_response,
        add_delivery_fees,
        cancel_complete_date,
        payment_at,
        product_amount,
        product_option_amount,
        cancel_complete_user,
        fcode,
        state,
        claim_request,
        claim_complete,
        cancel_reason,
        cancel_detail_reason,
        cancel_response,
        add_delivery_fees,
        cancel_complete_date,
        cancel_complete_user,
        fcode
    )
    print(sql, values)
    cursor.execute(sql, values)
os.remove(cancelFile)

path = logiFile

wb = load_workbook(path)

ws = wb.active

delOrder = '''DELETE From `bflow`.`channel_order` where channel = "11st"'''
cursor.execute(delOrder)
orderSql = '''
        INSERT INTO `bflow`.`channel_order` (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel,
        fcode,
        channel_amount,
        channel_calculate
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE state = %s
        '''
maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=3, max_row=maxRow):
    state = replace_none(row[1].value)
    channel_order_number = replace_none(row[2].value)
    channel_order_list = replace_int(row[3].value)
    product_name = replace_none(row[6].value)
    product_option = replace_none(row[7].value)
    quantity = replace_int(row[8].value)
    payment_at = row[5].value
    channel = '11st'
    channel_product_amount = row[10].value.replace(',', '')
    channel_option_amount = row[11].value.replace(',', '')
    channel_amount = str(int(channel_product_amount) + int(channel_option_amount))

    channel_calculate = row[17].value
    if product_option is None:
        fcode = None
    else:
        try:
            fcode = rex.search(product_option).group()
        except Exception as ex:
            fcode = None
            print(product_option)
            print(ex)
    orderValues = (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel,
        fcode,
        channel_amount,
        channel_calculate,
        state

    )
    cursor.execute(orderSql, orderValues)
    print(orderSql, orderValues)
os.remove(logiFile)
# 11번가 끝

# 지마켓 시작
driver.switch_to.window(driver.window_handles[0])
driver.get('https://www.esmplus.com/Member/SignIn/LogOn')
driver.find_element_by_xpath('//*[@id="rdoSiteSelect" and @value="GMKT"]').click()
driver.find_element_by_xpath('//*[@id="SiteId"]').send_keys(config.ESM_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="SitePassword"]').send_keys(config.ESM_LOGIN['password'])
driver.find_element_by_xpath('//*[@id="btnSiteLogOn"]').click()
time.sleep(3)
windowLists = driver.window_handles
for windowList in windowLists[1:]:
    driver.switch_to.window(driver.window_handles[-1])
    driver.close()

driver.switch_to.window(driver.window_handles[0])
driver.get('http://www.esmplus.com/Member/CustomerService/FindCustomer?menuCode=TDM144')
driver.find_element_by_xpath('//*[@id="contents"]/div/div[2]/ul/li[1]/span/span[5]/a').click()
driver.find_element_by_xpath('//*[@id="btnSearch"]').click()
driver.find_element_by_xpath('//*[@id="contents"]/div/div[3]/a').click()
time.sleep(15)

gmarketFile = change_file_to_xlsx('findcustomer_' + totalNow + '.xls', 'gmarket_state_')

path = gmarketFile

wb = load_workbook(path)

ws = wb.active

delOrder = '''DELETE From `bflow`.`channel_order` where channel = "gmarket" or channel = "auction"'''
cursor.execute(delOrder)
orderSql = '''
        INSERT INTO `bflow`.`channel_order` (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel,
        fcode
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE state = %s
        '''
maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=2):
    state = replace_none(row[12].value)
    channel_order_number = replace_none(row[2].value)
    channel_order_list = replace_none(row[3].value)
    product_name = replace_none(row[4].value)
    product_option = None
    quantity = replace_int(row[6].value)
    payment_at = replace_none(row[7].value)
    if channel_order_list[0] == 'B':
        channel = 'auction'
    else:
        channel = 'gmarket'
    if product_option is None:
        fcode = None
    else:
        fcode = rex.search(product_option).group()

    orderValues = (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        channel,
        fcode,
        state

    )
    cursor.execute(orderSql, orderValues)
    print(orderSql, orderValues)
os.remove(gmarketFile)
# 지마켓 끝

wb = Workbook()

ws = wb.active
no = 2

cancelState = f'''
        select `channel_order_number`,`fcode`, `product_name`,
        `product_option`, `state`, `cancel_reason`, `cancel_detail_reason`
        from `11st_cancel`;
        '''
cursor.execute(cancelState)
cancelNowTotal = cursor.fetchall()
for cancelNow in cancelNowTotal:
    channel_order_number = cancelNow[0]
    fcode = cancelNow[1]
    product_name = cancelNow[2]
    product_option = cancelNow[3]
    state = cancelNow[4]
    cancelReason = cancelNow[5]
    cancelDetailReason = cancelNow[6]
    print(channel_order_number, fcode)
    bflowStatus = requestStaus(channel_order_number, fcode)

    if bflowStatus['success'] is True:

        product_order_number = bflowStatus['message']['orderItemOptionId']
        order_number = bflowStatus['message']['orderCode']
        orderState = bflowStatus['message']['status']

        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']
            if claimType is None:
                claim_state = None
            elif claimType is 'cancel':
                claimType = '취소'
                claim_state = claimType + ":" + claimStatus
            elif claimType is 'return':
                claimType = '반품'
                claim_state = claimType + ":" + claimStatus
            elif claimType is 'exchange':
                claimType = '교환'
                claim_state = claimType + ":" + claimStatus
            else:
                claim_state = claimType + ":" + claimStatus
        else:
            claim_state = None

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '주문번호'
        ws.cell(row=1, column=3).value = '외부채널주문번호'
        ws.cell(row=1, column=4).value = '상품명'
        ws.cell(row=1, column=5).value = '상품옵션'
        ws.cell(row=1, column=6).value = '브리치 클레임상태'
        ws.cell(row=1, column=7).value = '브리치 주문상태'
        ws.cell(row=1, column=8).value = '11번가 상태'
        ws.cell(row=1, column=9).value = '11번가 클레임이'
        ws.cell(row=1, column=10).value = '11번가 클레임상세이유'
        print(orderState, state)
        if orderState == '결제취소' and state == '취소완료':
            print('skip')
            continue
        else:
            ws.cell(row=no, column=1).value = product_order_number
            ws.cell(row=no, column=2).value = order_number
            ws.cell(row=no, column=3).value = channel_order_number
            ws.cell(row=no, column=4).value = product_name
            ws.cell(row=no, column=5).value = product_option
            ws.cell(row=no, column=6).value = claim_state
            ws.cell(row=no, column=7).value = orderState
            ws.cell(row=no, column=8).value = state
            ws.cell(row=no, column=9).value = cancelReason
            ws.cell(row=no, column=10).value = cancelDetailReason

            no += 1
    else:
        pass
result = config.ST_LOGIN['excelPath'] + 'CancelResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()

sql = '''
    select `id`, `product_option`, `channel_order_number` from `channel_order`  where channel = '11st'
    '''
cursor.execute(sql)
optionRows = cursor.fetchall()

for optionRow in optionRows:
    idNo = optionRow[0]
    fcodeText = optionRow[1]
    channelOrderNumber = optionRow[2]
    fcode = None
    if fcodeText is not None:
        fcodeText = rex.search(optionRow[1])
        fcode = fcodeText.group()

    updateSql = '''
                update `channel_order` set fcode = %s where id = %s and channel_order_number = %s
                '''
    updateValue = (
        fcode,
        idNo,
        channelOrderNumber
    )
    cursor.execute(updateSql, updateValue)
    print(updateSql, updateValue)

ebayOrderList = f'''
            select `channel_order_number`, `product_name`, `state`, `channel`
            from `channel_order` where `payment_at` >= 2019-11-09 and `channel` in ('gmarket', 'auction', 'g9')
            and state not in ('입금대기', '판매자송금', '구매결정완료', '배송지연/발송예정', '주문확인'); 
            '''

cursor.execute(ebayOrderList)
ebayOrderRows = cursor.fetchall()

wb = Workbook()

ws = wb.active
no = 2

for ebayOrderRow in ebayOrderRows:
    channelOrderNumber = ebayOrderRow[0]
    productName = ebayOrderRow[1]
    state = ebayOrderRow[2]
    channel = ebayOrderRow[3]

    bflowStatus = requestStausChannel(channelOrderNumber, channel)

    if bflowStatus['success'] is True:

        productOrderNumber = bflowStatus['message']['orderItemOptionId']
        orderState = bflowStatus['message']['status']
        productOption = bflowStatus['message']['productOption']
        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

            if claimType is None:
                claim_state = None
            elif claimType is 'cancel':
                claimType = '취소'
                claim_state = claimType + ":" + claimStatus
            elif claimType is 'return':
                claimType = '반품'
                claim_state = claimType + ":" + claimStatus
            elif claimType is 'exchange':
                claimType = '교환'
                claim_state = claimType + ":" + claimStatus
            else:
                claim_state = claimType + ":" + claimStatus
        else:
            claim_state = None

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '외부채널주문번호'
        ws.cell(row=1, column=3).value = '상품명'
        ws.cell(row=1, column=4).value = '상품옵션'
        ws.cell(row=1, column=5).value = '브리치 주문상태'
        ws.cell(row=1, column=6).value = '브리치 클레임상태'
        ws.cell(row=1, column=7).value = '채널 상태'
        ws.cell(row=1, column=8).value = '채널명'

        # 비플로우 결제취소 / 채널상태 취소요청 , 취소중 , 반품완료  => 불필요
        # 비플로우 교환 / 채널상태 배송중, 교환요청 , 구매결정완료 => 불필요
        # 비플로우 반품 / 채널상태 반품보류 , 반품요청 = > 불필요
        # 비플로우 배송준비 / 채널 상태 배송지연 / 발송예정 = > 불필요
        # 비플로우 배송지연 / 채널상태 배송지연 / 발송예정  = > 불필요
        if state == '교환수거완료' \
                or state == '교환수거중' \
                or state == '교환완료' \
                or state == '배송중' \
                or state == '교환요청' \
                or state == '구매결정완료' \
                and orderState == '교환':
            print('skip')
            continue
        elif state == '반품수거완료' \
                or state == '반품수거중' \
                or state == '반품완료' \
                or state == '반품보류' \
                or state == '반품요청' \
                and orderState == '반품':
            print('skip')
            continue
        elif state == '입금확인' \
                or state == '주문확인' \
                or state == '배송지연/발송예정' \
                and orderState == '배송준비' \
                or orderState == '결제확인':
            print('skip')
            continue
        elif state == '취소완료' \
                or state == '환불완료' \
                or state == '취소요청' \
                or state == '취소중' \
                or state == '반품완료' \
                and orderState == '결제취소':
            print('skip')
            continue
        elif state == '배송중' and orderState == '출고완료' or orderState == '배송중':
            print('skip')
            continue
        elif state == '주문확인' or state == '배송지연/발송예정' and orderState == '배송지연':
            print('skip')
            continue
        elif state == '배송완료' or state == '구매결정완료' and orderState == '배송완료':
            print('skip')
            continue
        elif state == '미입금구매취소':
            print('skip')
            continue
        elif state == '입금대기':
            print('skip')
            continue
        elif state == '판매자송금':
            print('skip')
            continue
        elif state == '환불예정' and orderState == '결제취소':
            print('skip')
            continue
        elif state == '반품보류' or state == '미수취신고':
            ws.cell(row=no, column=1).value = productOrderNumber
            ws.cell(row=no, column=2).value = channelOrderNumber
            ws.cell(row=no, column=3).value = productName
            ws.cell(row=no, column=4).value = productOption
            ws.cell(row=no, column=5).value = claim_state
            ws.cell(row=no, column=6).value = orderState
            ws.cell(row=no, column=7).value = state
            ws.cell(row=no, column=8).value = channel
            no += 1
        else:
            ws.cell(row=no, column=1).value = productOrderNumber
            ws.cell(row=no, column=2).value = channelOrderNumber
            ws.cell(row=no, column=3).value = productName
            ws.cell(row=no, column=4).value = productOption
            ws.cell(row=no, column=5).value = claim_state
            ws.cell(row=no, column=6).value = orderState
            ws.cell(row=no, column=7).value = state
            ws.cell(row=no, column=8).value = channel
            no += 1
    else:
        pass
result = config.ST_LOGIN['excelPath'] + 'ebayOrderResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()
print(result)
cursor.close()
db.close()
driver.quit()
display.stop()
