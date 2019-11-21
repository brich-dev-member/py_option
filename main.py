from flask import Flask, render_template, redirect, request, send_file
import openpyxl
from itertools import permutations, combinations_with_replacement
from datetime import datetime

application = Flask(__name__)


@application.route('/')
def hello_world():
    return render_template('main.html')


@application.route('/option')
def option():
    return render_template('option.html')


@application.route('/option/upload', methods=['POST', 'GET'])
def option_slice():
    if request.method == 'POST':

        excel_file = request.files["excel_file"]
        filename = excel_file.filename
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        wa = wb.create_sheet('option')

        no = list()
        name = list()
        price = list()
        color = list()
        size = list()

        iter_row = iter(ws.rows)
        next(iter_row)

        for row in iter_row:
            no.append(row[0].value)
            name.append(row[1].value)
            price.append(row[2].value)
            colorOption = row[3].value
            if colorOption == "ONE COLOR":
                color.append(["ONE COLOR"])
            elif row[5].value == "원플러스":
                colors = colorOption.split(',')
                colorsPermutaions = list(combinations_with_replacement(colors, 2))
                onePluseList = list()
                for colorChain in colorsPermutaions:
                    first = colorChain[0]
                    secound = colorChain[1]
                    sumColor = first + "+" + secound
                    onePluseList.append(sumColor)
                # colorChain = list(chain(*colorsPermutaions))
                color.append(onePluseList)
                print(onePluseList)
            else:
                colors = colorOption.split(',')
                color.append(colors)

            sizeOption = str(row[4].value)
            sizes = sizeOption.split(',')
            size.append(sizes)

        optionLists = list(zip(no, name, price, color, size))
        lastLists = list()

        for optionList in optionLists:
            no = optionList[0]
            product = optionList[1]
            priceLists = optionList[2]
            colorLists = optionList[3]
            sizeLists = optionList[4]
            for colorList in range(len(colorLists)):
                color = colorLists[colorList]
                for sizeList in range(len(sizeLists)):
                    size = sizeLists[sizeList]
                    # total = color + "_" + size
                    lastLists.append(no)
                    lastLists.append(product)
                    lastLists.append(priceLists)
                    lastLists.append(color)
                    lastLists.append(size)

        def chunker(seq, size):
            return (seq[pos:pos + size] for pos in range(0, len(seq), size))

        no = 0

        for group in chunker(lastLists, 5):
            no += 1
            # print(group)
            iNo = group[0]
            iProduct = group[1]
            iPrice = group[2]
            # iTotal = group[3]
            iColor = group[3]
            iSize = group[4]

            wa.cell(row=no, column=1).value = iNo
            wa.cell(row=no, column=2).value = iProduct
            wa.cell(row=no, column=3).value = iPrice
            # wa.cell(row=no, column=4).value = iTotal
            wa.cell(row=no, column=4).value = iColor
            wa.cell(row=no, column=5).value = iSize

        makeToday = datetime.today()
        now = makeToday.strftime("%m%d_%H%M")
        result = './temp/' + filename[0:filename.find('.')] + "_" + now + ".xlsx"
        wb.save(result)
        print(result)
        return send_file(
            result,
            as_attachment=True,
            attachment_filename=result, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        return redirect('/option')


if __name__=='__main__':
    application.run(host='0.0.0.0')
