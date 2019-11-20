import os
import re
import time
from datetime import datetime
from glob import glob

import dateutil.relativedelta
import pymysql
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from slacker import Slacker

import config


def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)


def findSelect(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)


# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# f코드 정규
rex = re.compile('_F[0-9]+')
rexSpace = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

# 셀레니움 셋
options = Options()
# options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/Users/daegukim/py_option/chromedriver', options=options)


driver.get('https://biz.wemakeprice.com/partner/login')
driver.find_element_by_xpath('//*[@id="login_uid"]').send_keys(config.WMP_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="login_upw_biz"]').send_keys(config.WMP_LOGIN['password'])
driver.find_element_by_xpath('//*[@id="loginConfirmBtn_biz"]').click()
countSleep(1, 5)

countPopUp = driver.find_elements_by_xpath('//*[@id="agree"]/div[2]/a')
checkPopup = driver.find_elements_by_xpath('//*[@id="agree"]/div[2]/input')

print(countPopUp)
popupLists = []
for count in countPopUp:
    closeText = count.get_attribute('onclick')
    close = closeText.split("_")
    for check in checkPopup:
        checkText = check.get_attribute('id')
        ch = checkText.split("_")
        if close[1] == ch[1]:
            popupLists.append((closeText, checkText))
        else:
            continue

popupLists.reverse()
print(popupLists)

for popupList in popupLists:
    try:
        driver.find_element_by_xpath(f'//*[@id="{popupList[1]}"]').click()
        driver.find_element_by_xpath(f'//*[@id="agree"]/div[2]/a[@onclick="{popupList[0]}"]').click()
    except Exception as ex:
        print(ex)
countSleep(1, 3)

driver.get('http://biz.wemakeprice.com/dealer/claim_return/find')
countSleep(1, 3)

findSelect('//*[@id="input-perpage"]', '500')
driver.find_element_by_xpath(
    '/html/body/div[11]/div[2]/div[2]/div[1]/form/fieldset/div[1]/dl[1]/dd/div/div[2]/label[3]').click()
# driver.find_element_by_xpath('//*[@id="btn-1month"]').click()

driver.find_element_by_xpath('//*[@id="btn-find"]').click()
countSleep(1, 3)
driver.find_element_by_xpath('//*[@id="btn-excel"]').click()

resultCount = driver.find_element_by_xpath('//*[@id="place-result-count"]/span[1]').text
countSleep(1, 3)

findWmpXlsx = glob(config.ST_LOGIN['excelPath'] + "wmp_refund_list_*.xlsx")
print(findWmpXlsx)
wmpOriExcel = findWmpXlsx[0]
wmpResultExcel = config.ST_LOGIN['excelPath'] + 'wmp_return_' + now + '.xlsx'
print(wmpOriExcel)
os.rename(wmpOriExcel, wmpResultExcel)

wb = load_workbook(wmpResultExcel)
ws = wb.active

channelSql = '''
    INSERT INTO `excel`.`channel_returns` (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at
        ) VALUES (
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s
       )
        ON DUPLICATE KEY UPDATE 
        security_refund = %s,
        security_refund_at = %s,
        return_request_at = %s,
        return_complete_at = %s,
        return_delivery_case = %s,
        return_delivery_fees = %s,
        return_request_case = %s,
        channel_delivery_fees = %s,
        return_respons = %s,
        payment_case = %s,
        return_qty = %s,
        refund_state = %s,
        delivery_company = %s,
        delivery_code = %s,
        return_delivery_arrive_at = %s,
        return_hold_at = %s,
        return_delivery_complete_at = %s,
        fcode = %s,
        claim_state = %s
    '''

for row in ws.iter_rows(min_row=3):
    order_number = replacenone(row[3].value)
    order_number_line = replacenone(row[6].value)
    return_number = replacenone(row[1].value)
    channel = 'wemakeprice'
    security_refund = None
    security_refund_at = None
    return_request_at = row[4].value
    return_complete_at = row[5].value
    return_delivery_case = None
    return_delivery_fees = None
    return_request_case = replacenone(row[11].value)
    channel_delivery_fees = None
    return_respons = replacenone(row[13].value)
    payment_case = None
    return_qty = replaceint(row[9].value)
    refund_state = replacenone(row[2].value)
    product_name = replacenone(row[7].value)
    product_option = replacenone(row[8].value)
    if product_option is not None:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()
    delivery_company = None
    delivery_code = None
    return_delivery_arrive_at = None
    return_hold_at = None
    return_delivery_complete_at = None
    claim_state = '반품'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

driver.get('http://biz.wemakeprice.com/dealer/claim_exchange/find')
countSleep(1, 3)

findSelect('//*[@id="input-perpage"]', '500')
driver.find_element_by_xpath(
    '/html/body/div[11]/div[2]/div[2]/div[1]/form/fieldset/div[1]/dl[1]/dd/div/div[2]/label[3]').click()
driver.find_element_by_xpath('//*[@id="btn-find"]').click()
countSleep(1, 3)
driver.find_element_by_xpath('//*[@id="btn-excel"]').click()

resultCount = driver.find_element_by_xpath('//*[@id="place-result-count"]/span[1]').text
countSleep(1, 3)

findWmpXlsx = glob(config.ST_LOGIN['excelPath'] + "wmp_exchange_list_*.xlsx")
print(findWmpXlsx)
wmpOriExcel = findWmpXlsx[0]
wmpExchangeResultExcel = config.ST_LOGIN['excelPath'] + 'wmp_exchange_' + now + '.xlsx'
print(wmpOriExcel)
os.rename(wmpOriExcel, wmpExchangeResultExcel)

wb = load_workbook(wmpExchangeResultExcel)
ws = wb.active


for row in ws.iter_rows(min_row=3):
    order_number = replacenone(row[3].value)
    order_number_line = replacenone(row[6].value)
    return_number = replacenone(row[1].value)
    channel = 'wemakeprice'
    security_refund = None
    security_refund_at = None
    return_request_at = row[4].value
    return_complete_at = row[5].value
    return_delivery_case = None
    return_delivery_fees = None
    return_request_case = replacenone(row[11].value)
    channel_delivery_fees = None
    return_respons = replacenone(row[13].value)
    payment_case = None
    return_qty = replaceint(row[9].value)
    refund_state = replacenone(row[2].value)
    product_name = replacenone(row[7].value)
    product_option = replacenone(row[8].value)
    if product_option is not None:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()
    delivery_company = None
    delivery_code = None
    return_delivery_arrive_at = None
    return_hold_at = None
    return_delivery_complete_at = None
    claim_state = '교환'

    values = (
        order_number,
        order_number_line,
        return_number,
        channel,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        product_name,
        product_option,
        fcode,
        claim_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        security_refund,
        security_refund_at,
        return_request_at,
        return_complete_at,
        return_delivery_case,
        return_delivery_fees,
        return_request_case,
        channel_delivery_fees,
        return_respons,
        payment_case,
        return_qty,
        refund_state,
        delivery_company,
        delivery_code,
        return_delivery_arrive_at,
        return_hold_at,
        return_delivery_complete_at,
        fcode,
        claim_state
    )
    print(channelSql, values)
    cursor.execute(channelSql, values)

driver.close()