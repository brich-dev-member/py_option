from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import date
from datetime import datetime, timedelta
import pyexcel as p
import pymysql
from openpyxl import load_workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook
from slacker import Slacker

def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)


def findSelect(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)


# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# f코드 정규
rex = re.compile('_F[0-9]+')
rexSpace = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
halfNow = (makeToday - timedelta(weeks=2)).strftime("%Y-%m-%d")
weekNow = (makeToday - timedelta(weeks=1)).strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
print(halfNow)
options = Options()
options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

driver.switch_to.window(driver.window_handles[0])
driver.get('https://partner.brich.co.kr/login')
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/button[2]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[1]').send_keys(
    config.BFLOW_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[2]').send_keys(
    config.BFLOW_LOGIN['password'])
time.sleep(1)
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[2]/button[1]').click()
countSleep(1,4)

# 주문 1개월 업데이트
print('Sell')
driver.get(f'''
    https://partner.brich.co.kr/api/orders-excel-download?type=order&start={halfNow}
    &end={totalNow}&condition=code&content=&period=orders.created_at&orderby=orders.created_at
    &per_page=10&selectedProviderOptimusId=&selectedBrandOptimusId=&selectedCrawlingTarget=
    &productName=&productId=&isTodayDelivery=&isHold=&coupon_optimus_id=&refererDomain=
''')
countSleep(1,10)
bflowOriExcel = config.ST_LOGIN['excelPath'] + "orders_" + halfNow + "_" + totalNow + ".xlsx"
bflowResultExcel = config.ST_LOGIN['excelPath'] + 'bflow_order' + now + '.xlsx'

os.rename(bflowOriExcel, bflowResultExcel)

path = bflowResultExcel

wb = load_workbook(path)

ws = wb.active

sql = '''INSERT INTO `bflow`.`sell` (
        product_order_number,
        order_number,
        payment_at,
        order_state,
        claim,
        provider_name,
        product_name,
        product_option,
        channel,
        product_number,
        product_amount,
        option_amount,
        seller_discount,
        quantity,
        total_amount,
        delivery_at,
        delivery_complete,
        order_complete_at,
        auto_complete_at,
        category_number,
        buyer_email,
        buyer_gender,
        buyer_age,
        crawler,
        provider_number,
        channel_order_number,
        week,
        month,
        fcode
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE payment_at = %s, order_state = %s, claim = %s, delivery_at = %s, delivery_complete = %s,
        order_complete_at =%s, auto_complete_at = %s, channel_order_number = %s, week = %s, month = %s, fcode = %s
         '''

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)

for row in rows:
    product_order_number = replacenone(row[0].value)
    order_number = replacenone(row[1].value)
    payment_at = replacedate(row[4].value)
    order_state = replacenone(row[5].value)
    claim = replacenone(row[6].value)
    provider_name = replacenone(row[7].value)
    product_name = replacenone(row[8].value)
    product_option = replacenone(row[9].value)
    channel = replacenone(row[10].value)
    product_number = replacenone(row[20].value)
    product_amount = replaceint(row[21].value)
    option_amount = replaceint(row[22].value)
    seller_discount = replaceint(row[23].value)
    quantity = replaceint(row[24].value)
    total_amount = replaceint(row[25].value)
    delivery_at = replacedate(row[26].value)
    delivery_complete = replacedate(row[27].value)
    order_complete_at = replacedate(row[28].value)
    auto_complete_at = replacedate(row[29].value)
    category_number = replacenone(row[41].value)
    buyer_email = replacenone(row[42].value)
    buyer_gender = replacenone(row[43].value)
    buyer_age = replacenone(row[44].value)
    crawler = replacenone(row[45].value)
    provider_number = replacenone(row[46].value)
    channel_order_number = replacenone(row[3].value)
    if payment_at is not None:
        monthStr = datetime.strptime(payment_at, '%Y-%m-%d')
        week = monthStr.isocalendar()[1]
        month = monthStr.month
    else:
        week = None
        month = None

    if product_option is None:
        fcode = None
    else:
        makeCode = rex.search(product_option)
        if makeCode is None:
            fcode = None
        else:
            fcode = makeCode.group()

    values = (
        product_order_number, order_number, payment_at, order_state, claim, provider_name, product_name, product_option,
        channel, product_number, product_amount, option_amount, seller_discount, quantity, total_amount, delivery_at,
        delivery_complete, order_complete_at, auto_complete_at, category_number, buyer_email, buyer_gender, buyer_age,
        crawler, provider_number, channel_order_number, week, month, fcode, payment_at, order_state, claim, delivery_at,
        delivery_complete, order_complete_at, auto_complete_at, channel_order_number, week, month, fcode
    )
    cursor.execute(sql, values)
    print(sql, values)
print('distribution')
driver.get(f'''
    https://partner.brich.co.kr/api/distribution-confirm-excel-download?
    start={weekNow}&end={totalNow}&condition=linkage_mall_order_id&content=&status%5B%5D=80&
    period=order_item_options.confirmed_at&orderby=order_item_options.created_at&
    per_page=50&selectedProviderOptimusId=
    ''')
countSleep(1,10)
bflowdistributionOriExcel = config.ST_LOGIN['excelPath'] + "distribution_confirm_" + weekNow + "_" + totalNow +".xlsx"
bflowdistributionExcel = config.ST_LOGIN['excelPath'] + 'bflow_distribution' + now + '.xlsx'

os.rename(bflowdistributionOriExcel, bflowdistributionExcel)

path = bflowdistributionExcel

wb = load_workbook(path)

ws = wb.active

distributionSql = '''INSERT INTO `bflow`.`calculate` (
        product_order_number,
        order_number,
        channel_order_number,
        order_state,
        delivery_at,
        provider_name,
        channel,
        quantity,
        brich_product_price,
        fees,
        brich_calculate,
        channel_calculate,
        complete_at,
        match_at,
        margin,
        profit_rate,
        month
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE order_state = %s, delivery_at = %s, quantity = %s,
         brich_product_price = %s, fees = %s, brich_calculate =%s, channel_calculate = %s,
         complete_at = %s, match_at = %s, margin = %s, profit_rate = %s, month = %s
         '''

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)

for row in rows:
    product_order_number = replacenone(row[0].value)
    order_number = replacenone(row[1].value)
    channel_order_number = replacenone(row[2].value)
    order_state = replacenone(row[3].value)
    delivery_at = replacedate(row[4].value)
    provider_name = replacenone(row[5].value)
    channel = replacenone(row[6].value)
    quantity = replaceint(row[7].value)
    brich_product_price = replaceint(row[8].value)
    fees = replaceint(row[9].value)
    brich_calculate = replaceint(row[10].value)
    channel_calculate = replaceint(row[11].value)
    complete_at = replacedate(row[12].value)
    match_at = replacedate(row[13].value)
    if complete_at is not None:
        monthStr = datetime.strptime(complete_at, '%Y-%m-%d')
        month = monthStr.month
    else:
        month = None
    if brich_calculate is None or channel_calculate is None:
        continue
    else:
        margin = channel_calculate - brich_calculate
        profit_rate = margin / brich_product_price * 100

    values = (
        product_order_number,
        order_number,
        channel_order_number,
        order_state,
        delivery_at,
        provider_name,
        channel,
        quantity,
        brich_product_price,
        fees,
        brich_calculate,
        channel_calculate,
        complete_at,
        match_at,
        margin,
        profit_rate,
        month,
        order_state,
        delivery_at,
        quantity,
        brich_product_price,
        fees,
        brich_calculate,
        channel_calculate,
        complete_at,
        match_at,
        margin,
        profit_rate,
        month
    )

    cursor.execute(distributionSql, values)
    print(distributionSql, values)
print('product_Dump')

baseUrl = 'https://partner.brich.co.kr/api/products/export-dump?params={%22start%22:%22'
subUrl = '%22,%22end%22:%22'
endUrl = '''%22,%22product_optimus_id%22:%22%22,%22name%22:%22%22,%22brand_name%22:%22%22,%22status%22:[],%22period%22:%22created_at%22,%22categories%22:{%22depth1%22:%22%22,%22depth2%22:%22%22,%22depth3%22:%22%22,%22depth4%22:%22%22},%22orderby%22:%22created_at%22,%22per_page%22:50,%22page%22:1,%22approve_status%22:[],%22custom_code%22:%22%22,%22provider_optimus_id%22:%22%22,%22manager_optimus_id%22:%22%22,%22distribution%22:[],%22price%22:{%22type%22:null,%22lte%22:null,%22gte%22:null},%22crawled_product%22:false,%22is_today_delivery%22:false,%22is_today_sale%22:false,%22is_celeb_group_buying%22:false,%22is_group_buying%22:false,%22official_information_status%22:%22%22,%22disApproved%22:false}'''
totlaUrl = baseUrl + weekNow + subUrl + totalNow + endUrl
print(totlaUrl)
driver.get(totlaUrl)

countSleep(1,10)
bflowProductOriExcel = config.ST_LOGIN['excelPath'] + "products.xlsx"
bflowProductExcel = config.ST_LOGIN['excelPath'] + 'bflow_Product_' + now + '.xlsx'

os.rename(bflowProductOriExcel, bflowProductExcel)

path = bflowProductExcel

wb = load_workbook(path)

ws = wb.active

productSql = '''INSERT INTO `bflow`.`product` (
        confirm,
        state,
        product_number,
        provider_name,
        provider_number,
        product_name,
        brand,
        category,
        category_number,
        price,
        start_date,
        end_date,
        create_date,
        update_date,
        week,
        month,
        is_deal,
        ssg_fees,
        gmarket_fees,
        auction_fees,
        11st_fees,
        coupang_fees,
        interpark_fees,
        wemakeprice_fees,
        tmon_fees,
        g9_fees
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE confirm = %s, state = %s, product_name = %s,
         category = %s, category_number = %s, price =%s, start_date = %s, end_date = %s,
         create_date = %s, update_date = %s, week = %s, month = %s, is_deal = %s,
         ssg_fees = %s, gmarket_fees = %s, auction_fees = %s, 11st_fees = %s,
         coupang_fees = %s, interpark_fees = %s, wemakeprice_fees = %s, tmon_fees = %s, g9_fees = %s
         '''

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)

for row in rows:
    confirm = replacenone(row[0].value)
    state = replacenone(row[1].value)
    product_number = replacenone(row[2].value)
    provider_name = replacenone(row[3].value)
    provider_number = replacenone(row[4].value)
    product_name = replacenone(row[5].value)
    brand = replacenone(row[6].value)
    category = replacenone(row[7].value)
    category_number = replacenone(row[8].value)
    price = replaceint(row[9].value)
    deal = replacenone(row[10].value)
    if deal is str('Y'):
        is_deal = 1
    else:
        is_deal = 0
    channel_fees = {}
    channels = row[11].value.replace(" ", "")
    channelSplits = channels.split(',')
    print(channelSplits)
    for channelSplit in channelSplits:
        channel = channelSplit.split(':')
        channel_fees[channel[0]] = channel[1]
    print(channel_fees)
    ssg_fees = float(channel_fees.get('ssg').replace("%", ""))
    gmarket_fees = float(channel_fees.get('gmarket').replace("%", ""))
    auction_fees = float(channel_fees.get('auction').replace("%", ""))
    st_fees = float(channel_fees.get('11st').replace("%", ""))
    coupang_fees = float(channel_fees.get('coupang').replace("%", ""))
    interpark_fees = float(channel_fees.get('interpark').replace("%", ""))
    wemakeprice_fees = float(channel_fees.get('wemakeprice').replace("%", ""))
    tmon_fees = float(channel_fees.get('tmon').replace("%", ""))
    g9_fees = float(channel_fees.get('g9').replace("%", ""))
    saleDate = replacenone(row[12].value).replace(" ", "")
    if saleDate is not None:
        dates = saleDate.split("~")
        start_date = replacedate(dates[0])
        end_date = replacedate(dates[1])
        print(start_date,end_date)
    create_date = replacedate(row[13].value)
    update_date = replacedate(row[14].value)
    if create_date is not None:
        monthStr = datetime.strptime(create_date, '%Y-%m-%d')
        week = monthStr.isocalendar()[1]
        month = monthStr.month
    else:
        week = None
        month = None

    values = (
        confirm,
        state,
        product_number,
        provider_name,
        provider_number,
        product_name,
        brand,
        category,
        category_number,
        price,
        start_date,
        end_date,
        create_date,
        update_date,
        week,
        month,
        is_deal,
        ssg_fees,
        gmarket_fees,
        auction_fees,
        st_fees,
        coupang_fees,
        interpark_fees,
        wemakeprice_fees,
        tmon_fees,
        g9_fees,
        confirm,
        state,
        product_name,
        category,
        category_number,
        price,
        start_date,
        end_date,
        create_date,
        update_date,
        week,
        month,
        is_deal,
        ssg_fees,
        gmarket_fees,
        auction_fees,
        st_fees,
        coupang_fees,
        interpark_fees,
        wemakeprice_fees,
        tmon_fees,
        g9_fees
    )
    print(values)
    cursor.execute(productSql, values)

driver.quit()
db.close()



