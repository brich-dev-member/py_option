import requests
import json
from openpyxl import Workbook
from datetime import date
from datetime import datetime
import dateutil.relativedelta
from bs4 import BeautifulSoup

# 'sort': 'RECENT', 최신순
# 'sort': 'POPULARITY' 인기순
# 'sort': 'REVIEW', 리뷰순

# 날짜 관련
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

def styleWindowDealList(providerNumber, sort, count):
    
    menuUrl = 'https://shopping.naver.com/v1/channels/1000003442/menus'
    menuParams = {
        '_nc_': '',
        'parentId': '0',
        'displayType': 'NO_RULE',
    }
    menuRes = requests.get(url=menuUrl, params=menuParams)
    menuJson = json.loads(menuRes.text)
    menuID = menuJson[0]['id']
    print(menuID)
    swUrl = 'https://shopping.naver.com/v1/products'
    
    swParams = {
        '_nc_': '',
        'subVertical': 'STYLE',
        'page': '1',
        'pageSize': count,
        'sort': sort,
        'filter': 'ALL',
        'displayType': 'CATEGORY_HOME',
        'includeZzim': 'true',
        'includeViewCount': 'true',
        'includeStoreCardInfo': 'true',
        'includeStockQuantity': 'false',
        'includeBrandInfo': 'false',
        'includeBrandLogoImage': 'false',
        'includeRepresentativeReview': 'false',
        'includeListCardAttribute': 'false',
        'includeRanking': 'false',
        'includeRankingByMenus': 'false',
        'includeStoreCategoryName': 'false',
        'menuId': menuID,
        'storeId': providerNumber,
        'standardSizeKeys': '',
        'standardColorKeys':'' ,
        'optionFilters': '',
        'attributeValueIds': '',
        'attributeValueIdsAll': '',
        'certifications': ''
    }
    swRes = requests.get(url=swUrl, params=swParams)
    swJson = json.loads(swRes.text)

    wb = Workbook()
    ws = wb.active
    
    no = 2

    for product in swJson['products']:
        providerName = product['channel']['name']
        productUrl = 'https://shopping.naver.com/style/style/stores/1000003442/products/' + product['_id']
        productName = product['name']
        productSalePrice = product['salePrice']
        productDiscountPrice = product['mobileDiscountPrice']
        productContent = product['contentText']
        isSoldOut = product['soldout']
        productSellCount = product['totalSaleCount']
        productViewCount = product['viewCountFromWindowApi']
        productPopular = product['popularScore']
        updateAt = product['updatedAt']

        optionGetUrl = requests.get(url=productUrl)
        bs = BeautifulSoup(optionGetUrl.text, 'html.parser')
        try:
            color = []
            findColor = bs.find('strong', text='색상을 선택하세요').parent
            for colorOption in findColor.find_all('span'):
                color.append(colorOption.string)
            
            colorList = ','.join(color)
        except Exception as ex:
            colorList = None
            print(ex)

        try:
            size = []
            findSize = bs.find('span', text='사이즈를 선택하세요').parent
            for sizeOption in findSize.find_all('button'):
                size.append(sizeOption.string)

            sizeList = ','.join(size)
        except Exception as ex:
            sizeList = None
            print(ex)
        

        if isSoldOut == 'True':
            continue
        else:
            ws.cell(row=1, column=1).value = '순번'
            ws.cell(row=1, column=2).value = '상품URL'
            ws.cell(row=1, column=3).value = '상품명'
            ws.cell(row=1, column=4).value = '판매가'
            ws.cell(row=1, column=5).value = '할인판매가'
            ws.cell(row=1, column=6).value = '색상'
            ws.cell(row=1, column=7).value = '사이즈'
            ws.cell(row=1, column=8).value = '상품설명'
            ws.cell(row=1, column=9).value = '품절여부'
            ws.cell(row=1, column=10).value = '판매량'
            ws.cell(row=1, column=11).value = '조회수'
            ws.cell(row=1, column=12).value = '인기지수'
            ws.cell(row=1, column=13).value = '수정일'

            ws.cell(row=no, column=1).value = no - 1
            ws.cell(row=no, column=2).value = productUrl
            ws.cell(row=no, column=3).value = productName
            ws.cell(row=no, column=4).value = productSalePrice
            ws.cell(row=no, column=5).value = productDiscountPrice
            ws.cell(row=no, column=6).value = colorList
            ws.cell(row=no, column=7).value = sizeList
            ws.cell(row=no, column=8).value = productContent
            ws.cell(row=no, column=9).value = isSoldOut
            ws.cell(row=no, column=10).value = productSellCount
            ws.cell(row=no, column=11).value = productViewCount
            ws.cell(row=no, column=12).value = productPopular
            ws.cell(row=no, column=13).value = updateAt

            no += 1
    
    result = './temp/' + providerName + "_" + sort + "_" + str(count) + "_" + now + ".xlsx"
    wb.save(result)
    print(result)
    return result


# 'sort': 'RECENT', 최신순
# 'sort': 'POPULARITY' 인기순
# 'sort': 'REVIEW', 리뷰순



