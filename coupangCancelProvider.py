import os
import re
import time
from datetime import datetime
from datetime import datetime, timedelta
import json
import dateutil.relativedelta
from openpyxl import load_workbook, Workbook
from slacker import Slacker
import requests
import config

# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
weekNow = (makeToday - timedelta(weeks=1)).strftime("%Y-%m-%d")

wb = Workbook()

ws = wb.active

no = 2

with requests.Session() as s:

    loginUrl = 'https://wing.coupang.com/login'
    loginData = {
        'username': config.COUPANG_LOGIN['id'],
        'password': config.COUPANG_LOGIN['password']
    }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
    }
    loginRequest = s.post(url=loginUrl, data=loginData, headers=headers)
    print(loginRequest.status_code)
    jsonHeaders = {
        'Content-Type': 'application/json;charset=UTF-8'
    }
    scoreUrl = s.get('https://wing.coupang.com/seller/performance/v2/performance?_=1578542535497', headers=jsonHeaders)
    scoreInfo = json.loads(scoreUrl.text)
    print(scoreInfo)
    score = scoreInfo['cancellation']['last7daysRate']
    print('cancel rate :', score)
    cancelCount = int(scoreInfo['cancellation']['last7DaysDen']) - int(scoreInfo['cancellation']['last7DaysNum'])
    print('cancel count :', cancelCount)

    dataTime = datetime.strptime(scoreInfo['cancellation']['statusChangeDate'], '%Y-%m-%d')
    cancelEndDate = (dataTime - timedelta(days=1)).strftime("%Y-%m-%d")
    cancelstartDate = (datetime.strptime(cancelEndDate, '%Y-%m-%d') - timedelta(weeks=1) + timedelta(days=1)).strftime("%Y-%m-%d")
    print(cancelstartDate, cancelEndDate)

    cancelUrl = 'https://wing.coupang.com/seller/performance/v2/cancellationInfo'
    cancelData = {"includeExceptionCases":'false',"page":0,"pageSize":cancelCount,"orderIds":[],"vendorId":"A00191857","expectCount":cancelCount ,"startDate":cancelstartDate,"endDate":cancelEndDate}
    cancelListPage = s.post(url=cancelUrl, data=json.dumps(cancelData),  headers=jsonHeaders)
    cancelJsonLists = json.loads(cancelListPage.text)

    for cancelList in cancelJsonLists['content']:
        productName = cancelList['productName'].split('/')
        print(productName)
        cancelUnitCount = cancelList['cancelUnitCount']
        cancelReason = cancelList['cancelReason']
        cancellationDate = datetime.fromtimestamp(cancelList['cancellationDate']/1000)

        ws.cell(row=1, column=1).value = '순번'
        ws.cell(row=1, column=2).value = '업체명'
        ws.cell(row=1, column=3).value = '상품명'
        ws.cell(row=1, column=4).value = '취소수량'
        ws.cell(row=1, column=5).value = '취소사유'
        ws.cell(row=1, column=6).value = '취소일'
        
        ws.cell(row=no, column=1).value = no - 1
        if len(productName) > 1:
            ws.cell(row=no, column=2).value = productName[0]
            ws.cell(row=no, column=3).value = productName[1]
        else:
            ws.cell(row=no, column=2).value = None
            ws.cell(row=no, column=3).value = productName[0] 
        ws.cell(row=no, column=4).value = cancelUnitCount
        ws.cell(row=no, column=5).value = cancelReason
        ws.cell(row=no, column=6).value = cancellationDate

        no += 1


result = config.ST_LOGIN['excelPath'] + 'coupangCancelList_' + cancelstartDate + "_" + cancelEndDate +  "_" + now + '.xlsx'
print(result)
wb.save(result)
wb.close()

sendResult = open(result, 'rb')
slack.files.upload(
        file_=sendResult,
        channels=config.SLACK_API['channels2'],
        title=f'주문이행점수 : {score} 점 / 취소건수 : {cancelCount} 건',
    )
print('Remove File :', result)
os.remove(result)


# URL
# https://wing.coupang.com/seller/performance?showPanel=cancellation
# 디데일 팝업 get
# https://wing.coupang.com/seller/performance/v2/orderSearch/4000060275683/simpleview?_=1578539119942


