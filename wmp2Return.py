import os
import re
import time
from datetime import datetime
from datetime import datetime, timedelta
from glob import glob
import json
import dateutil.relativedelta
import pymysql
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from slacker import Slacker
from reqStatus import requestStaus, requestStausChannel

import config
# from pyvirtualdisplay import Display

# display = Display(visible=0, size=(1200, 900))
# display.start()


def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)


def findSelect(xpath, value):
    el = Select(driver.find_element_by_xpath(xpath))
    el.select_by_value(value)


# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

# f코드 정규
rex = re.compile('_F[0-9]+')
rexSpace = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
weekNow = (makeToday - timedelta(weeks=1)).strftime("%Y-%m-%d")

# 셀레니움 셋
options = Options()
options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

driver.get('https://wpartner.wemakeprice.com/login')
countSleep(1, 3)
driver.find_element_by_xpath('/html/body/div[3]/div[1]/div/div[1]/span[1]/input').send_keys(config.WMP_LOGIN['id'])
driver.find_element_by_xpath('/html/body/div[3]/div[1]/div/div[1]/span[2]/input').send_keys(config.WMP_LOGIN['password'])
driver.find_element_by_xpath('//*[@id="login"]').click()
countSleep(1, 3)

driver.get('https://wpartner.wemakeprice.com/claim/cancelMain?dhxr1576662488197=1')

#취소신청 리스트
                      
wpCanceRequestlUrl = f'https://wpartner.wemakeprice.com/claim/getClaimCancelList.json?schTotalFlag=&schDateType=requestDate&schStartDate={endNow}&schEndDate={totalNow}&schClaimStatus=&schType=&schValue=&schApproveId=&schLimitCnt=300&schPageNo=1&_=1577064494779'
driver.get(wpCanceRequestlUrl)
print(wpCanceRequestlUrl)
print('cancel Request')
countSleep(1, 3)
try:
    wpCancelRequestText = driver.find_element_by_tag_name('pre').text
    wpCancelRequestLists = json.loads(wpCancelRequestText)
except Exception as ex:
    wpCancelRequestLists = []
    print(ex)
#취소완료 리스트
wpCanceCompletelUrl = f'https://wpartner.wemakeprice.com/claim/getClaimCancelList.json?schTotalFlag=cancelComplete&schDateType=requestDate&schStartDate={endNow}&schEndDate={totalNow}&schClaimStatus=C1&schType=&schValue=&schApproveId=&schLimitCnt=300&schPageNo=1&_=1576663651065'
driver.get(wpCanceCompletelUrl)
print(wpCanceCompletelUrl)
print('cancel Complete')
countSleep(1, 3)
try:
    wpCancelCompleteText = driver.find_element_by_tag_name('pre').text
    wpCancelCompleteLists = json.loads(wpCancelCompleteText)
except Exception as ex:
    wpCancelCompleteLists = []
    print(ex)
#반품관리
wpReturnRequestUrl = f'https://wpartner.wemakeprice.com/claim/getClaimReturnList.json?schTotalFlag=&schDateType=requestDate&schStartDate={endNow}&schEndDate={totalNow}&schClaimStatus=&schPickupStatus=&schType=&schValue=&schApproveId=&schLimitCnt=300&schPageNo=1&_=1577064517017'
driver.get(wpReturnRequestUrl)
print(wpReturnRequestUrl)
print('return Request')
countSleep(1, 3)
try:
    wpReturnRequestText = driver.find_element_by_tag_name('pre').text
    wpReturnRequestLists = json.loads(wpReturnRequestText)
except Exception as ex:
    wpReturnRequestLists = []
    print(ex)

#교환관리 
wpExchangeRequestUrl = f'https://wpartner.wemakeprice.com/claim/getClaimExchangeList.json?schTotalFlag=&schPageNo=1&schDateType=requestDate&schStartDate={endNow}&schEndDate={totalNow}&schClaimStatus=&schPickupStatus=&schReceiveStatus=&schType=&schValue=&schLimitCnt=300&schPageNo=1&schIsNoLimit=N&_=1577064531915'
driver.get(wpExchangeRequestUrl)
print(wpExchangeRequestUrl)
print('Exchange Request')
countSleep(1, 3)
try:
    wpExchangeRequestText = driver.find_element_by_tag_name('pre').text
    wpExchangeRequestLists = json.loads(wpExchangeRequestText)
except Exception as ex:
    wpExchangeRequestLists = []
    print(ex)
 
wpResultLists = wpCancelRequestLists + wpCancelCompleteLists + wpReturnRequestLists + wpExchangeRequestLists
wb = Workbook()

ws = wb.active

no = 2

for wpResult in wpResultLists:
    print(wpResult)
    ChannelOrderNumber = wpResult['orderNo']
    productOptionRex = wpResult['optNm']
    wpClaimNo = wpResult['claimBundleNo']
    wpStatus = wpResult['claimStatusNm']
    wpRequestDate = wpResult['requestDate']
    fcode = rex.search(productOptionRex).group()


    bflowStatus = requestStaus(ChannelOrderNumber, fcode)
    print(bflowStatus['message'])
    if bflowStatus['success'] is True:
        productOrderNo = bflowStatus['message']['orderItemOptionId']
        orderNumber = bflowStatus['message']['orderCode']
        channel = bflowStatus['message']['channel'] 
        status = bflowStatus['message']['status']
        paymentAt = bflowStatus['message']['payCompletedAt']
        if len(bflowStatus['message']['claims']) > 0:
                claimType = bflowStatus['message']['claims'][0]['claimType']
                claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

                if claimType is None:
                    claim_state = None
                elif claimType == 'cancel':
                    claimType = '취소'
                    claim_state = claimType + ":" + claimStatus
                elif claimType == 'return':
                    claimType = '반품'
                    claim_state = claimType + ":" + claimStatus
                elif claimType == 'exchange':
                    claimType = '교환'
                    claim_state = claimType + ":" + claimStatus
                else:
                    claim_state = claimType + ":" + claimStatus
                
                if claimType == '반품' and wpStatus == '반품신청' or wpStatus == '반품지연' or wpStatus == '반품보류' or wpStatus == '반품완료':
                    claimPaymentCase = wpResult['claimShipFeeEncloseNm']
                    claimRespons = wpResult['whoReason']
                    claimReturnShipPrice = wpResult['claimReturnShipPrice']
                    shipCompleteDt = wpResult['pickupP3Date']
                    companyNm = wpResult['pickupCompanyNm']
                    invoiceNo = wpResult['pickupInvoice']
                elif claimType == '교환' and wpStatus == '교환신청' or wpStatus == '교환지연' or wpStatus == '교환보류' or wpStatus == '교환완료':
                    claimPaymentCase = wpResult['claimShipFeeEncloseNm']
                    claimRespons = wpResult['whoReason']
                    claimReturnShipPrice = wpResult['claimExchangeShipPrice']
                    shipCompleteDt = wpResult['pickupP3Date']
                    companyNm = wpResult['pickupCompanyNm']
                    invoiceNo = wpResult['pickupInvoice']
        else:
                claim_state = None
    
        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '주문번호'
        ws.cell(row=1, column=3).value = '채널 주문번호'
        ws.cell(row=1, column=4).value = '결제일'
        ws.cell(row=1, column=5).value = '채널'
        ws.cell(row=1, column=6).value = '브리치 주문상태'
        ws.cell(row=1, column=7).value = '브리치 클레임'
        ws.cell(row=1, column=8).value = '채널 상태'
        ws.cell(row=1, column=9).value = '채널 클레임 요청일'
        ws.cell(row=1, column=10).value = '반품 배송비'
        ws.cell(row=1, column=11).value = '배송비 지불방법'
        ws.cell(row=1, column=12).value = '귀책여부'
        ws.cell(row=1, column=13).value = '회수 택배사'
        ws.cell(row=1, column=14).value = '회수 송장번호'
        ws.cell(row=1, column=15).value = '회수 완료일'
        
        # 비플로우 / 외부채널(위메프 2.0)
        # 취소:요청 / 취소신청
        # 취소:처리완료 / 취소신청
        # 취소:처리완료 / 취소승인
        # 반품:수거중 / 반품신청
        # 교환:수거중 / 교환신청

        if claim_state == '취소:처리완료' and wpStatus == '취소완료':
            continue
        elif claim_state == '취소:요청' and wpStatus == '취소신청':
            continue
        elif claim_state == '취소:처리완료' and wpStatus == '취소신청':
            continue
        elif claim_state == '취소:처리완료' and wpStatus == '취소승인':
            continue
        elif claim_state == '반품:수거중' and wpStatus == '반품신청':
            continue
        elif claim_state == '교환:수거중' and wpStatus == '교환신청':
            continue
        elif wpStatus == '취소신청':
            ws.cell(row=no, column=1).value = productOrderNo
            ws.cell(row=no, column=2).value = orderNumber
            ws.cell(row=no, column=3).value = ChannelOrderNumber
            ws.cell(row=no, column=4).value = paymentAt
            ws.cell(row=no, column=5).value = channel
            ws.cell(row=no, column=6).value = status
            ws.cell(row=no, column=7).value = claim_state
            ws.cell(row=no, column=8).value = wpStatus
            ws.cell(row=no, column=9).value = wpRequestDate

            no += 1
        elif status == '반품' or wpStatus == '반품신청':
            try:
                ws.cell(row=no, column=1).value = productOrderNo
                ws.cell(row=no, column=2).value = orderNumber
                ws.cell(row=no, column=3).value = ChannelOrderNumber
                ws.cell(row=no, column=4).value = paymentAt
                ws.cell(row=no, column=5).value = channel
                ws.cell(row=no, column=6).value = status
                ws.cell(row=no, column=7).value = claim_state
                ws.cell(row=no, column=8).value = wpStatus
                ws.cell(row=no, column=9).value = wpRequestDate
                ws.cell(row=no, column=10).value = claimReturnShipPrice
                ws.cell(row=no, column=11).value = claimPaymentCase
                ws.cell(row=no, column=12).value = claimRespons
                ws.cell(row=no, column=13).value = companyNm
                ws.cell(row=no, column=14).value = invoiceNo
                ws.cell(row=no, column=15).value = shipCompleteDt

                no += 1
            except Exception as ex:
                print(ex)
        elif status == '교환' or wpStatus == '교환신청':
            try:
                ws.cell(row=no, column=1).value = productOrderNo
                ws.cell(row=no, column=2).value = orderNumber
                ws.cell(row=no, column=3).value = ChannelOrderNumber
                ws.cell(row=no, column=4).value = paymentAt
                ws.cell(row=no, column=5).value = channel
                ws.cell(row=no, column=6).value = status
                ws.cell(row=no, column=7).value = claim_state
                ws.cell(row=no, column=8).value = wpStatus
                ws.cell(row=no, column=9).value = wpRequestDate
                ws.cell(row=no, column=10).value = claimReturnShipPrice
                ws.cell(row=no, column=11).value = claimPaymentCase
                ws.cell(row=no, column=12).value = claimRespons
                ws.cell(row=no, column=13).value = companyNm
                ws.cell(row=no, column=14).value = invoiceNo
                ws.cell(row=no, column=15).value = shipCompleteDt

                no += 1
            except Exception as ex:
                print(ex)
        else :
            ws.cell(row=no, column=1).value = productOrderNo
            ws.cell(row=no, column=2).value = orderNumber
            ws.cell(row=no, column=3).value = ChannelOrderNumber
            ws.cell(row=no, column=4).value = paymentAt
            ws.cell(row=no, column=5).value = channel
            ws.cell(row=no, column=6).value = status
            ws.cell(row=no, column=7).value = claim_state
            ws.cell(row=no, column=8).value = wpStatus
            ws.cell(row=no, column=9).value = wpRequestDate

            no += 1
    
    elif bflowStatus['success'] is False:
        continue

result = config.ST_LOGIN['excelPath'] + 'wpNewResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
driver.close()
driver.quit()
print(result)
#https://wpartner.wemakeprice.com/login