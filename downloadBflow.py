import requests
import config
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import time
import json
import config
import os
from datetime import date
from datetime import datetime, timedelta
import pymysql
from openpyxl import load_workbook
import dateutil.relativedelta
import re
from openpyxl import Workbook
from slacker import Slacker
from tqdm import tqdm

def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text

def countSleep(sleep, total):
    for count in range(1, total):
        print(count)
        time.sleep(sleep)

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# f코드 정규
rex = re.compile('_F[0-9]+')
# 날짜 모듈
makeToday = datetime.now()
makeWeek = datetime.weekday(makeToday)
makeTime = datetime.time(makeToday).strftime('%H:%M')
now = makeToday.strftime("%m-%d_%H-%M-%S")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")
findNow = makeToday.strftime("%Y%m%d%H")

options = Options()
options.add_argument('--headless')
options.add_argument("disable-gpu")
prefs = {
    "download.default_directory": config.ST_LOGIN['excelPath'],
    "directory_upgrade": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(executable_path='/usr/bin/chromedriver', options=options)

print(driver.window_handles)
driver.get('https://partner.brich.co.kr/login')
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/button[2]').click()
time.sleep(2)
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[1]').send_keys(
    config.BFLOW_LOGIN['id'])
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[2]').send_keys(
    config.BFLOW_LOGIN['password'])
time.sleep(1)
driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[2]/button[1]').click()
time.sleep(4)
print('login')



downLoadList = []

def findBflowExcel(fileName):
    excelDownUrl = 'https://partner.brich.co.kr/api/excel-downloads?page=1&search={"start":"","end":"","orderBy":"created_at","period":"created_at","perPage":100,"content":null,"page":1}'
    driver.get(excelDownUrl)
    excelText = driver.find_element_by_tag_name('pre').text
    excelData = json.loads(excelText)

    dataRows = excelData['data']['data']
    print(dataRows)
    for row in dataRows:
        print(row['filename'], row['status'])
        if fileName not in row['filename']:
            continue
        elif fileName in row['filename']:
            if row['status'] == 'complete':
                fileList = {
                    'download_url':row['download_url'],
                    'filename':row['filename'],
                    'division_file_count':row['division_file_count'],
                    'total_file_count':row['total_file_count']
                }
                downLoadList.append(fileList)
            else:
                continue
    
    print(downLoadList)

def insertDB():
    for downList in downLoadList:
        driver.get(downList['download_url'])
        countSleep(1,3)
        insertFile = config.ST_LOGIN['excelPath'] + downList['filename']
        print(insertFile)
        if 'order_list_' in insertFile:
            wb = load_workbook(insertFile)

            ws = wb.active

            sql = '''INSERT INTO `bflow`.`sell` (
                    product_order_number,
                    order_number,
                    payment_at,
                    order_state,
                    claim,
                    provider_name,
                    product_name,
                    product_option,
                    channel,
                    product_number,
                    product_amount,
                    option_amount,
                    seller_discount,
                    quantity,
                    total_amount,
                    delivery_at,
                    delivery_complete,
                    order_complete_at,
                    auto_complete_at,
                    category_number,
                    buyer_gender,
                    buyer_age,
                    crawler,
                    provider_number,
                    channel_order_number,
                    week,
                    month,
                    fcode
                    ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE payment_at = %s, order_state = %s, claim = %s, delivery_at = %s, delivery_complete = %s,
                    order_complete_at =%s, auto_complete_at = %s, channel_order_number = %s, week = %s, month = %s, fcode = %s
                    '''

            iter_row = iter(ws.rows)
            next(iter_row)

            rows = tqdm(iter_row)

            for row in rows:
                product_order_number = replacenone(row[0].value)
                order_number = replacenone(row[1].value)
                payment_at = replacedate(row[5].value)
                order_state = replacenone(row[6].value)
                claim = replacenone(row[7].value)
                provider_name = replacenone(row[8].value)
                product_name = replacenone(row[9].value)
                product_option = replacenone(row[10].value)
                korChannel = replacenone(row[2].value)
                if korChannel == '브리치':
                    channel = 'brich'
                elif korChannel == '지마켓':
                    channel = 'gmarket'
                elif korChannel == '옥션':
                    channel = 'auction'
                elif korChannel == '11번가':
                    channel = '11st'
                elif korChannel == '인터파크':
                    channel = 'interpark'
                elif korChannel == '위메프':
                    channel = 'wemakeprice'
                elif korChannel == '티몬':
                    channel = 'tmon'
                elif korChannel == 'g9':
                    channel = 'g9'
                elif korChannel == 'ssg':
                    channel = 'ssg'
                product_number = replacenone(row[20].value)
                product_amount = replaceint(row[21].value)
                option_amount = replaceint(row[22].value)
                seller_discount = replaceint(row[23].value)
                quantity = replaceint(row[24].value)
                total_amount = replaceint(row[25].value)
                delivery_at = replacedate(row[26].value)
                delivery_complete = replacedate(row[27].value)
                order_complete_at = replacedate(row[28].value)
                auto_complete_at = replacedate(row[29].value)
                category_number = replacenone(row[41].value)
                buyer_gender = replacenone(row[42].value)
                buyer_age = replacenone(row[43].value)
                crawler = replacenone(row[44].value)
                provider_number = replacenone(row[45].value)
                channel_order_number = replacenone(row[3].value)
                if payment_at is not None:
                    monthStr = datetime.strptime(payment_at, '%Y-%m-%d')
                    week = monthStr.isocalendar()[1]
                    month = monthStr.month
                else:
                    week = None
                    month = None

                if product_option is None:
                    fcode = None
                else:
                    makeCode = rex.search(product_option)
                    if makeCode is None:
                        fcode = None
                    else:
                        fcode = makeCode.group()

                values = (
                    product_order_number, order_number, payment_at, order_state, claim, provider_name, product_name, product_option,
                    channel, product_number, product_amount, option_amount, seller_discount, quantity, total_amount, delivery_at,
                    delivery_complete, order_complete_at, auto_complete_at, category_number, buyer_gender, buyer_age,
                    crawler, provider_number, channel_order_number, week, month, fcode, payment_at, order_state, claim, delivery_at,
                    delivery_complete, order_complete_at, auto_complete_at, channel_order_number, week, month, fcode
                )
                cursor.execute(sql, values)
                print(sql, values)
            os.remove(insertFile)
            print('delete file : ', insertFile )
        elif 'product_dump_list_' in insertFile:
            wb = load_workbook(insertFile)

            ws = wb.active

            productSql = '''INSERT INTO `bflow`.`product` (
                    confirm,
                    state,
                    product_number,
                    provider_name,
                    provider_number,
                    product_name,
                    brand,
                    category,
                    category_number,
                    price,
                    start_date,
                    end_date,
                    create_date,
                    update_date,
                    week,
                    month,
                    is_deal,
                    ssg_fees,
                    gmarket_fees,
                    auction_fees,
                    11st_fees,
                    coupang_fees,
                    interpark_fees,
                    wemakeprice_fees,
                    tmon_fees,
                    g9_fees
                    ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE confirm = %s, state = %s, product_name = %s,
                    category = %s, category_number = %s, price =%s, start_date = %s, end_date = %s,
                    create_date = %s, update_date = %s, week = %s, month = %s, is_deal = %s,
                    ssg_fees = %s, gmarket_fees = %s, auction_fees = %s, 11st_fees = %s,
                    coupang_fees = %s, interpark_fees = %s, wemakeprice_fees = %s, tmon_fees = %s, g9_fees = %s
                    '''

            iter_row = iter(ws.rows)
            next(iter_row)

            rows = tqdm(iter_row)

            for row in rows:
                confirm = replacenone(row[0].value)
                state = replacenone(row[1].value)
                product_number = replacenone(row[2].value)
                provider_name = replacenone(row[3].value)
                provider_number = replacenone(row[4].value)
                product_name = replacenone(row[5].value)
                brand = replacenone(row[6].value)
                category = replacenone(row[7].value)
                category_number = replacenone(row[8].value)
                price = replaceint(row[9].value)
                deal = replacenone(row[10].value)
                if deal is str('Y'):
                    is_deal = 1
                else:
                    is_deal = 0
                channel_fees = {}
                channels = row[11].value.replace(" ", "")
                channelSplits = channels.split(',')
                print(channelSplits)
                for channelSplit in channelSplits:
                    channel = channelSplit.split(':')
                    channel_fees[channel[0]] = channel[1]
                print(channel_fees)
                ssg_fees = float(channel_fees.get('ssg').replace("%", ""))
                gmarket_fees = float(channel_fees.get('gmarket').replace("%", ""))
                auction_fees = float(channel_fees.get('auction').replace("%", ""))
                st_fees = float(channel_fees.get('11st').replace("%", ""))
                coupang_fees = float(channel_fees.get('coupang').replace("%", ""))
                interpark_fees = float(channel_fees.get('interpark').replace("%", ""))
                wemakeprice_fees = float(channel_fees.get('wemakeprice').replace("%", ""))
                tmon_fees = float(channel_fees.get('tmon').replace("%", ""))
                g9_fees = float(channel_fees.get('g9').replace("%", ""))
                saleDate = replacenone(row[12].value).replace(" ", "")
                if saleDate is not None:
                    dates = saleDate.split("~")
                    start_date = replacedate(dates[0])
                    end_date = replacedate(dates[1])
                    print(start_date,end_date)
                create_date = replacedate(row[13].value)
                update_date = replacedate(row[14].value)
                if create_date is not None:
                    monthStr = datetime.strptime(create_date, '%Y-%m-%d')
                    week = monthStr.isocalendar()[1]
                    month = monthStr.month
                else:
                    week = None
                    month = None

                values = (
                    confirm,
                    state,
                    product_number,
                    provider_name,
                    provider_number,
                    product_name,
                    brand,
                    category,
                    category_number,
                    price,
                    start_date,
                    end_date,
                    create_date,
                    update_date,
                    week,
                    month,
                    is_deal,
                    ssg_fees,
                    gmarket_fees,
                    auction_fees,
                    st_fees,
                    coupang_fees,
                    interpark_fees,
                    wemakeprice_fees,
                    tmon_fees,
                    g9_fees,
                    confirm,
                    state,
                    product_name,
                    category,
                    category_number,
                    price,
                    start_date,
                    end_date,
                    create_date,
                    update_date,
                    week,
                    month,
                    is_deal,
                    ssg_fees,
                    gmarket_fees,
                    auction_fees,
                    st_fees,
                    coupang_fees,
                    interpark_fees,
                    wemakeprice_fees,
                    tmon_fees,
                    g9_fees
                )
                print(values)
                cursor.execute(productSql, values)
            os.remove(insertFile)
            print('delete file : ', insertFile )
        elif 'order_distribution_confirm_' in insertFile:
            wb = load_workbook(insertFile)

            ws = wb.active

            distributionSql = '''INSERT INTO `bflow`.`calculate` (
                    product_order_number,
                    order_number,
                    channel_order_number,
                    order_state,
                    delivery_at,
                    provider_name,
                    channel,
                    quantity,
                    brich_product_price,
                    fees,
                    brich_calculate,
                    channel_calculate,
                    complete_at,
                    match_at,
                    margin,
                    profit_rate,
                    month
                    ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE order_state = %s, delivery_at = %s, quantity = %s,
                    brich_product_price = %s, fees = %s, brich_calculate =%s, channel_calculate = %s,
                    complete_at = %s, match_at = %s, margin = %s, profit_rate = %s, month = %s
                    '''

            iter_row = iter(ws.rows)
            next(iter_row)

            rows = tqdm(iter_row)

            for row in rows:
                product_order_number = replacenone(row[0].value)
                order_number = replacenone(row[1].value)
                channel_order_number = replacenone(row[2].value)
                order_state = replacenone(row[3].value)
                delivery_at = replacedate(row[4].value)
                provider_name = replacenone(row[5].value)
                channel = replacenone(row[6].value)
                quantity = replaceint(row[7].value)
                brich_product_price = replaceint(row[8].value)
                fees = replaceint(row[9].value)
                brich_calculate = replaceint(row[10].value)
                channel_calculate = replaceint(row[11].value)
                complete_at = replacedate(row[12].value)
                match_at = replacedate(row[13].value)
                if complete_at is not None:
                    monthStr = datetime.strptime(complete_at, '%Y-%m-%d')
                    month = monthStr.month
                else:
                    month = None
                if brich_calculate is None or channel_calculate is None:
                    continue
                else:
                    margin = channel_calculate - brich_calculate
                    profit_rate = margin / brich_product_price * 100

                values = (
                    product_order_number,
                    order_number,
                    channel_order_number,
                    order_state,
                    delivery_at,
                    provider_name,
                    channel,
                    quantity,
                    brich_product_price,
                    fees,
                    brich_calculate,
                    channel_calculate,
                    complete_at,
                    match_at,
                    margin,
                    profit_rate,
                    month,
                    order_state,
                    delivery_at,
                    quantity,
                    brich_product_price,
                    fees,
                    brich_calculate,
                    channel_calculate,
                    complete_at,
                    match_at,
                    margin,
                    profit_rate,
                    month
                )

                cursor.execute(distributionSql, values)
                print(distributionSql, values)
            os.remove(insertFile)
            print('delete file : ', insertFile )
        elif 'order_return_list_' in insertFile:

            wb = load_workbook(insertFile)

            ws = wb.active

            sql = '''
                INSERT INTO `bflow`.`Bflow_returns` (
                    product_order_number,
                    order_number,
                    return_number,
                    channel,
                    make,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    provider_name,
                    product_name,
                    product_option,
                    fcode
                    ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE 
                    payment_at = %s,
                    return_request_at = %s,
                    return_complete_at = %s,
                    return_delivery_case = %s,
                    return_delivery_fees = %s,
                    return_request_case = %s,
                    return_qty = %s,
                    refund_state = %s,
                    claim_state = %s,
                    fast_refund = %s,
                    fcode = %s

                    '''

            # Excel read

            maxRow = ws.max_row

            for row in ws.iter_rows(min_row=2, max_row=maxRow):
                product_order_number = replacenone(row[0].value)
                order_number = replacenone(row[1].value)
                return_number = replacenone(row[2].value)
                channel = replacenone(row[3].value)
                make = replacenone(row[4].value)
                payment_at = row[5].value
                return_request_at = row[6].value
                return_complete_at = row[7].value
                return_delivery_case = replacenone(row[8].value)
                return_delivery_fees = replaceint(row[9].value)
                return_request_case = replacenone(row[11].value)
                return_qty = replaceint(row[12].value)
                refund_state = replacenone(row[13].value)
                claim_state = replacenone(row[15].value)
                fast_refund = replacenone(row[17].value)
                provider_name = replacenone(row[18].value)
                product_name = replacenone(row[19].value)
                product_option = replacenone(row[20].value)
                if product_option is not None:
                    makeCode = rex.search(product_option)
                    if makeCode is None:
                        fcode = None
                    else:
                        fcode = makeCode.group()

                values = (
                    product_order_number,
                    order_number,
                    return_number,
                    channel,
                    make,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    provider_name,
                    product_name,
                    product_option,
                    fcode,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    fcode,
                )

                cursor.execute(sql, values)
                print(sql, values)
            os.remove(insertFile)
            print('delete file : ', insertFile )
        elif 'order_exchange_list' in insertFile:

            wb = load_workbook(insertFile)

            ws = wb.active

            maxRow = ws.max_row

            sql = '''
                INSERT INTO `bflow`.`Bflow_returns` (
                    product_order_number,
                    order_number,
                    return_number,
                    channel,
                    make,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    provider_name,
                    product_name,
                    product_option,
                    fcode
                    ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE 
                    payment_at = %s,
                    return_request_at = %s,
                    return_complete_at = %s,
                    return_delivery_case = %s,
                    return_delivery_fees = %s,
                    return_request_case = %s,
                    return_qty = %s,
                    refund_state = %s,
                    claim_state = %s,
                    fast_refund = %s,
                    fcode = %s

                    '''

            for row in ws.iter_rows(min_row=2, max_row=maxRow):
                product_order_number = replacenone(row[0].value)
                order_number = replacenone(row[1].value)
                return_number = replacenone(row[2].value)
                channel = replacenone(row[3].value)
                make = replacenone(row[4].value)
                payment_at = row[5].value
                return_request_at = row[6].value
                return_complete_at = None
                return_delivery_case = replacenone(row[8].value)
                return_delivery_fees = replaceint(row[9].value)
                return_request_case = replacenone(row[7].value)
                return_qty = replaceint(row[11].value)
                refund_state = replacenone(row[12].value)
                claim_state = replacenone(row[14].value)
                fast_refund = None
                provider_name = replacenone(row[15].value)
                product_name = replacenone(row[16].value)
                product_option = replacenone(row[17].value)
                if product_option is not None:
                    makeCode = rex.search(product_option)
                    if makeCode is None:
                        fcode = None
                    else:
                        fcode = makeCode.group()

                values = (
                    product_order_number,
                    order_number,
                    return_number,
                    channel,
                    make,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    provider_name,
                    product_name,
                    product_option,
                    fcode,
                    payment_at,
                    return_request_at,
                    return_complete_at,
                    return_delivery_case,
                    return_delivery_fees,
                    return_request_case,
                    return_qty,
                    refund_state,
                    claim_state,
                    fast_refund,
                    fcode,
                )

                cursor.execute(sql, values)
                print(sql, values)
            os.remove(insertFile)
            print('delete file : ', insertFile )
        countSleep(1,3)
    del downLoadList[:]

orderName = 'order_list_' + findNow
productName = 'product_dump_list_' + findNow
distributionName = 'order_distribution_confirm_pending_confirm_' + findNow
returnName = 'order_return_list_' + findNow
exchangeName = 'order_exchange_list_' + findNow
findBflowExcel(orderName)
insertDB()
findBflowExcel(exchangeName)
insertDB()

findBflowExcel(returnName)
insertDB()

findBflowExcel(distributionName)
insertDB()
findBflowExcel(productName)
insertDB()


db.close()
driver.quit()
