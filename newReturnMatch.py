import config
from openpyxl import Workbook
import pymysql
from datetime import date
from datetime import datetime
import dateutil.relativedelta
from reqStatus import requestStaus, requestStausChannel

# 날짜 모듈
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")


# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

sql = '''
    select c.`order_number`, c.`refund_state`,c.`return_request_at`,
    c.`return_complete_at`, c.`return_respons`, c.`payment_case`, c.`delivery_company`,
    c.`delivery_code`,c.`return_delivery_arrive_at`, c.`fcode`, c.`channel`
    from `channel_returns` as c 
    where c.`refund_state` is not NULL
    and not c.`refund_state` in ('반품완료'); 
    '''

cursor.execute(sql)
newReturn = cursor.fetchall()

wb = Workbook()

ws = wb.active

no = 2

for returnList in newReturn:
    channel_order_number = returnList[0]
    refund_state = returnList[1]
    return_request_at = returnList[2]
    return_complete_at = returnList[3]
    return_respons = returnList[4]
    payment_case = returnList[5]
    delivery_company = returnList[6]
    delivery_code = returnList[7]
    return_delivery_arrive_at = returnList[8]
    fcode = returnList[9]
    channel = returnList[10]

    if channel == 'gmarket':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    elif channel == 'auction':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    elif channel == 'g9':
        bflowStatus = requestStausChannel(channel_order_number, channel)
        print(bflowStatus['message'])
    else:
        bflowStatus = requestStaus(channel_order_number, fcode)
        print(bflowStatus['message'])
    
    if bflowStatus['success'] is True:
        product_order_number = bflowStatus['message']['orderItemOptionId']
        order_number = bflowStatus['message']['orderCode']
        channel = bflowStatus['message']['channel']
        payment_at = bflowStatus['message']['payCompletedAt']
        order_state = bflowStatus['message']['status']
        if len(bflowStatus['message']['claims']) > 0:
            claimType = bflowStatus['message']['claims'][0]['claimType']
            claimStatus = bflowStatus['message']['claims'][0]['claimStatus']

            if claimType is None:
                claim_state = None
            elif claimType == 'cancel':
                claimType = '취소'
                claim_state = claimType + ":" + claimStatus
            elif claimType == 'return':
                claimType = '반품'
                claim_state = claimType + ":" + claimStatus
            elif claimType == 'exchange':
                claimType = '교환'
                claim_state = claimType + ":" + claimStatus
            else:
                claim_state = claimType + ":" + claimStatus
        else:
            claim_state = None
    
        print(claim_state, refund_state)

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '주문번호'
        ws.cell(row=1, column=3).value = '채널 주문번호'
        ws.cell(row=1, column=4).value = '결제일'
        ws.cell(row=1, column=5).value = '채널'
        ws.cell(row=1, column=6).value = '브리치 주문상태'
        ws.cell(row=1, column=7).value = '브리치 클레임'
        ws.cell(row=1, column=8).value = '채널 상태'
        ws.cell(row=1, column=9).value = '채널 클레임 요청일'
        ws.cell(row=1, column=10).value = '채널 클레임 완료일'
        ws.cell(row=1, column=11).value = '귀책여부'
        ws.cell(row=1, column=12).value = '비용처리'
        ws.cell(row=1, column=13).value = '택배사'
        ws.cell(row=1, column=14).value = '송장번호'
        ws.cell(row=1, column=15).value = '수거 완료일'
        if refund_state == '반품완료':
            continue
        elif order_state == '결제취소' or order_state == '반품' or order_state == '교환':
            continue
        else:
            ws.cell(row=no, column=1).value =  product_order_number
            ws.cell(row=no, column=2).value =  order_number
            ws.cell(row=no, column=3).value =  channel_order_number
            ws.cell(row=no, column=4).value =  payment_at
            ws.cell(row=no, column=5).value =  channel
            ws.cell(row=no, column=6).value =  order_state
            ws.cell(row=no, column=7).value =  claim_state
            ws.cell(row=no, column=8).value =  refund_state
            ws.cell(row=no, column=9).value =  return_request_at
            ws.cell(row=no, column=10).value = return_complete_at
            ws.cell(row=no, column=11).value = return_respons
            ws.cell(row=no, column=12).value = payment_case
            ws.cell(row=no, column=13).value = delivery_company
            ws.cell(row=no, column=14).value = delivery_code
            ws.cell(row=no, column=15).value = return_delivery_arrive_at

            no += 1
    elif bflowStatus['success'] is False:
        pass
result = config.ST_LOGIN['excelPath'] + 'channelReturnMissMatch_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()
db.close()
print(result)


