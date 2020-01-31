from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import datetime
from datetime import datetime, timedelta
import pyexcel as p
import pymysql
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook
from reqStatus import requestStaus, requestStausChannel
import json
from slacker import Slacker
import requests

# 날짜 관련
makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()
# 슬랙 인증
slack = Slacker(config.SLACK_API['token'])

orderSql = f'''
            select `channel_order_number`, `fcode`, `channel_amount`, `channel_calculate`
            from `channel_order`
            where `channel` = '11st' and date(`payment_at`) >= date(subdate(now(), INTERVAL 2 DAY)) and date(`payment_at`) <= date(now());
            '''
cursor.execute(orderSql)
stOrderRequests = cursor.fetchall()

wb = Workbook()
ws = wb.active
no = 2

for orderRequest in stOrderRequests:
    channe_order_number = orderRequest[0]
    fcode = orderRequest[1]
    channel_amount = int(orderRequest[2].replace(',',''))
    channel_calculate = int(orderRequest[3].replace(',',''))

    bflowStatus = requestStaus(channe_order_number, fcode)
    print(bflowStatus['message'])

    if bflowStatus['success'] is True:
        bflowProductOrderNumber = bflowStatus['message']['orderItemOptionId']
        bflowProductName = bflowStatus['message']['productName']
        bflowProductOption = bflowStatus['message']['productOption']
        bflowTotalAmount = bflowStatus['message']['totalBuyPrice']
        bflowPaymentAt = bflowStatus['message']['payCompletedAt']
        profit = round(channel_calculate / bflowTotalAmount * 100, 2)
        print('수익율 :', profit)

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '채널명'
        ws.cell(row=1, column=3).value = '상품명'
        ws.cell(row=1, column=4).value = '옵션명'
        ws.cell(row=1, column=5).value = '주문일'
        ws.cell(row=1, column=6).value = '비플로우 판매금액'
        ws.cell(row=1, column=7).value = '채널    판매금액'
        ws.cell(row=1, column=8).value = '채널정산 예정금액'
        ws.cell(row=1, column=9).value = '수익율'
    
        if profit < 85:
            ws.cell(row=no, column=1).value = bflowProductOrderNumber
            ws.cell(row=no, column=2).value = '11st'
            ws.cell(row=no, column=3).value = bflowProductName
            ws.cell(row=no, column=4).value = bflowProductOption
            ws.cell(row=no, column=5).value = bflowPaymentAt
            ws.cell(row=no, column=6).value = bflowTotalAmount
            ws.cell(row=no, column=7).value = channel_amount
            ws.cell(row=no, column=8).value = channel_calculate
            ws.cell(row=no, column=9).value = profit

            no += 1

    elif bflowStatus['success'] is False:
        continue
result = '11stProfit_' + now + '.xlsx'
wb.save(result)