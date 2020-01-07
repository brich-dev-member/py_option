import pymysql
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils.cell import get_column_letter
import config

wb = Workbook()

ws = wb.active
wa = wb.create_sheet('통합지표')

# DB
db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)
border_right = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='medium', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
)

toImportant = NamedStyle(name="toImportant")

toImportant.border = border
toImportant.fill = PatternFill('solid', fgColor='EEEEEE')
toImportant.font = Font(size=11, bold=True)
toImportant.alignment = Alignment(horizontal='center', vertical='center')


def noZerodiv(a, b):
    if a is not None and b is not None:
        return round(b / a, 2)
    else:
        return None


def noneCheck(a, b):
    if a is not None and b is not None:
        return a - b
    else:
        return None


startMonth = 1
endMonth = startMonth + 1
year = 2020

for i in range(startMonth, endMonth):
    sql = f'''
    SELECT 
    date,
    min(date),
    max(date),
    sum(brich_total_amount),
    sum(brich_qty),
    sum(brich_sales),
    sum(brich_cogs),
    sum(gmarket_total_amount),
    sum(gmarket_qty),
    sum(gmarket_sales),
    sum(gmarket_cogs),
    sum(auction_total_amount),
    sum(auction_qty),
    sum(auction_sales),
    sum(auction_cogs),
    sum(11st_total_amount),
    sum(11st_qty),
    sum(11st_sales),
    sum(11st_cogs),
    sum(wemakeprice_total_amount),
    sum(wemakeprice_qty),
    sum(wemakeprice_sales),
    sum(wemakeprice_cogs),
    sum(interpark_total_amount),
    sum(interpark_qty),
    sum(interpark_sales),
    sum(interpark_cogs),
    sum(coupang_total_amount),
    sum(coupang_qty),
    sum(coupang_sales),
    sum(coupang_cogs),
    sum(ssg_total_amount),
    sum(ssg_qty),
    sum(ssg_sales),
    sum(ssg_cogs),
    sum(g9_total_amount),
    sum(g9_qty),
    sum(g9_sales),
    sum(g9_cogs),
    sum(tmon_total_amount),
    sum(tmon_qty),
    sum(tmon_sales),
    sum(tmon_cogs)
    FROM sell_to_channel WHERE month = {i} and year = {year} GROUP BY date
    '''

    cursor.execute(sql)
    rows = cursor.fetchall()

    refundSql = f'''
    SELECT 
    month,
    min(date),
    max(date),
    sum(brich_refund_amount),
    sum(brich_refund_qty),
    sum(gmarket_refund_amount),
    sum(gmarket_refund_qty),
    sum(auction_refund_amount),
    sum(auction_refund_qty),
    sum(11st_refund_amount),
    sum(11st_refund_qty),
    sum(wemakeprice_refund_amount),
    sum(wemakeprice_refund_qty),
    sum(interpark_refund_amount),
    sum(interpark_refund_qty),
    sum(coupang_refund_amount),
    sum(coupang_refund_qty),
    sum(ssg_refund_amount),
    sum(ssg_refund_qty),
    sum(g9_refund_amount),
    sum(g9_refund_qty),
    sum(tmon_refund_amount),
    sum(tmon_refund_qty)
    FROM sell_to_channel WHERE month = {i} and year = {year} GROUP BY month
    '''

    cursor.execute(refundSql)
    refundRows = cursor.fetchall()

    endRow = ws.max_row + 1
    no = 0 + endRow

    for row in rows:
        week = row[0]
        weekstr = datetime.strftime(row[1], '%Y-%m-%d') + "~" + datetime.strftime(row[2], '%Y-%m-%d')
        brich_total_amount = row[3]
        brich_qty = row[4]
        brich_CT = noZerodiv(brich_qty, brich_total_amount)
        brich_sales = row[5]
        brich_cogs = row[6]
        brich_margin = noneCheck(brich_sales, brich_cogs)
        gmarket_total_amount = row[7]
        gmarket_qty = row[8]
        gmarket_CT = noZerodiv(gmarket_qty, gmarket_total_amount)
        gmarket_sales = row[9]
        gmarket_cogs = row[10]
        gmarket_margin = noneCheck(gmarket_sales, gmarket_cogs)
        auction_total_amount = row[11]
        auction_qty = row[12]
        auction_CT = noZerodiv(auction_qty, auction_total_amount)
        auction_sales = row[13]
        auction_cogs = row[14]
        auction_margin = noneCheck(auction_sales, auction_cogs)
        st_total_amount = row[15]
        st_qty = row[16]
        st_CT = noZerodiv(st_qty, st_total_amount)
        st_sales = row[17]
        st_cogs = row[18]
        st_margin = noneCheck(st_sales, st_cogs)
        wemakeprice_total_amount = row[19]
        wemakeprice_qty = row[20]
        wemakeprice_CT = noZerodiv(wemakeprice_qty, wemakeprice_total_amount)
        wemakeprice_sales = row[21]
        wemakeprice_cogs = row[22]
        wemakeprice_margin = noneCheck(wemakeprice_sales, wemakeprice_cogs)
        interpark_total_amount = row[23]
        interpark_qty = row[24]
        interpark_CT = noZerodiv(interpark_qty, interpark_total_amount)
        interpark_sales = row[25]
        interpark_cogs = row[26]
        interpark_margin = noneCheck(interpark_sales, interpark_cogs)
        coupnag_total_amount = row[27]
        coupnag_qty = row[28]
        coupnag_CT = noZerodiv(coupnag_qty, coupnag_total_amount)
        coupnag_sales = row[29]
        coupnag_cogs = row[30]
        coupnag_margin = noneCheck(coupnag_sales, coupnag_cogs)
        ssg_total_amount = row[31]
        ssg_qty = row[32]
        ssg_CT = noZerodiv(ssg_qty, ssg_total_amount)
        ssg_sales = row[33]
        ssg_cogs = row[34]
        ssg_margin = noneCheck(ssg_sales, ssg_cogs)
        g9_total_amount = row[35]
        g9_qty = row[36]
        g9_CT = noZerodiv(g9_qty, g9_total_amount)
        g9_sales = row[37]
        g9_cogs = row[38]
        g9_margin = noneCheck(g9_sales, g9_cogs)
        tmon_total_amount = row[39]
        tmon_qty = row[40]
        tmon_CT = noZerodiv(tmon_qty, tmon_total_amount)
        tmon_sales = row[41]
        tmon_cogs = row[42]
        tmon_margin = noneCheck(tmon_sales, tmon_cogs)

        ws.cell(row=1, column=1).value = '일자'
        ws.cell(row=1, column=1).style = toImportant
        ws.cell(row=1, column=1).border = border
        ws.cell(row=1, column=2).value = '전체 거래액'
        ws.cell(row=1, column=2).style = toImportant
        ws.cell(row=1, column=2).border = border
        ws.cell(row=1, column=3).value = '전체 판매수'
        ws.cell(row=1, column=3).style = toImportant
        ws.cell(row=1, column=3).border = border
        ws.cell(row=1, column=4).value = '전체 객단가'
        ws.cell(row=1, column=4).style = toImportant
        ws.cell(row=1, column=4).border = border
        ws.cell(row=1, column=5).value = '전체 매출'
        ws.cell(row=1, column=5).style = toImportant
        ws.cell(row=1, column=5).border = border
        ws.cell(row=1, column=6).value = '전체 매출원가'
        ws.cell(row=1, column=6).style = toImportant
        ws.cell(row=1, column=6).border = border
        ws.cell(row=1, column=7).value = '전체 마진'
        ws.cell(row=1, column=7).style = toImportant
        ws.cell(row=1, column=7).border = border

        ws.cell(row=1, column=8).value = '브리치 거래액'
        ws.cell(row=1, column=8).style = toImportant
        ws.cell(row=1, column=8).border = border
        ws.cell(row=1, column=9).value = '브리치 판매수'
        ws.cell(row=1, column=9).style = toImportant
        ws.cell(row=1, column=9).border = border
        ws.cell(row=1, column=10).value = '브리치 객단가'
        ws.cell(row=1, column=10).style = toImportant
        ws.cell(row=1, column=10).border = border
        ws.cell(row=1, column=11).value = '브리치 매출'
        ws.cell(row=1, column=11).style = toImportant
        ws.cell(row=1, column=11).border = border
        ws.cell(row=1, column=12).value = '브리치 매출원가'
        ws.cell(row=1, column=12).style = toImportant
        ws.cell(row=1, column=12).border = border
        ws.cell(row=1, column=13).value = '브리치 마진'
        ws.cell(row=1, column=13).style = toImportant
        ws.cell(row=1, column=13).border = border

        ws.cell(row=1, column=14).value = '지마켓 거래액'
        ws.cell(row=1, column=14).style = toImportant
        ws.cell(row=1, column=14).border = border
        ws.cell(row=1, column=15).value = '지마켓 판매수'
        ws.cell(row=1, column=15).style = toImportant
        ws.cell(row=1, column=15).border = border
        ws.cell(row=1, column=16).value = '지마켓 객단가'
        ws.cell(row=1, column=16).style = toImportant
        ws.cell(row=1, column=16).border = border
        ws.cell(row=1, column=17).value = '지마켓 매출'
        ws.cell(row=1, column=17).style = toImportant
        ws.cell(row=1, column=17).border = border
        ws.cell(row=1, column=18).value = '지마켓 매출원가'
        ws.cell(row=1, column=18).style = toImportant
        ws.cell(row=1, column=18).border = border
        ws.cell(row=1, column=19).value = '지마켓 마진'
        ws.cell(row=1, column=19).style = toImportant
        ws.cell(row=1, column=19).border = border

        ws.cell(row=1, column=20).value = '옥션 거래액'
        ws.cell(row=1, column=20).style = toImportant
        ws.cell(row=1, column=20).border = border
        ws.cell(row=1, column=21).value = '옥션 판매수'
        ws.cell(row=1, column=21).style = toImportant
        ws.cell(row=1, column=21).border = border
        ws.cell(row=1, column=22).value = '옥션 객단가'
        ws.cell(row=1, column=22).style = toImportant
        ws.cell(row=1, column=22).border = border
        ws.cell(row=1, column=23).value = '옥션 매출'
        ws.cell(row=1, column=23).style = toImportant
        ws.cell(row=1, column=23).border = border
        ws.cell(row=1, column=24).value = '옥션 매출원가'
        ws.cell(row=1, column=24).style = toImportant
        ws.cell(row=1, column=24).border = border
        ws.cell(row=1, column=25).value = '옥션 마진'
        ws.cell(row=1, column=25).style = toImportant
        ws.cell(row=1, column=25).border = border

        ws.cell(row=1, column=26).value = '11번가 거래액'
        ws.cell(row=1, column=26).style = toImportant
        ws.cell(row=1, column=26).border = border
        ws.cell(row=1, column=27).value = '11번가 판매수'
        ws.cell(row=1, column=27).style = toImportant
        ws.cell(row=1, column=27).border = border
        ws.cell(row=1, column=28).value = '11번가 객단가'
        ws.cell(row=1, column=28).style = toImportant
        ws.cell(row=1, column=28).border = border
        ws.cell(row=1, column=29).value = '11번가 매출'
        ws.cell(row=1, column=29).style = toImportant
        ws.cell(row=1, column=29).border = border
        ws.cell(row=1, column=30).value = '11번가 매출원가'
        ws.cell(row=1, column=30).style = toImportant
        ws.cell(row=1, column=30).border = border
        ws.cell(row=1, column=31).value = '11번가 마진'
        ws.cell(row=1, column=31).style = toImportant
        ws.cell(row=1, column=31).border = border

        ws.cell(row=1, column=32).value = '위메프 거래액'
        ws.cell(row=1, column=32).style = toImportant
        ws.cell(row=1, column=32).border = border
        ws.cell(row=1, column=33).value = '위메프 판매수'
        ws.cell(row=1, column=33).style = toImportant
        ws.cell(row=1, column=33).border = border
        ws.cell(row=1, column=34).value = '위메프 객단가'
        ws.cell(row=1, column=34).style = toImportant
        ws.cell(row=1, column=34).border = border
        ws.cell(row=1, column=35).value = '위메프 매출'
        ws.cell(row=1, column=35).style = toImportant
        ws.cell(row=1, column=35).border = border
        ws.cell(row=1, column=36).value = '위메프 매출원가'
        ws.cell(row=1, column=36).style = toImportant
        ws.cell(row=1, column=36).border = border
        ws.cell(row=1, column=37).value = '위메프 마진'
        ws.cell(row=1, column=37).style = toImportant
        ws.cell(row=1, column=37).border = border

        ws.cell(row=1, column=38).value = '인터파크 거래액'
        ws.cell(row=1, column=38).style = toImportant
        ws.cell(row=1, column=38).border = border
        ws.cell(row=1, column=39).value = '인터파크 판매수'
        ws.cell(row=1, column=39).style = toImportant
        ws.cell(row=1, column=39).border = border
        ws.cell(row=1, column=40).value = '인터파크 객단가'
        ws.cell(row=1, column=40).style = toImportant
        ws.cell(row=1, column=40).border = border
        ws.cell(row=1, column=41).value = '인터파크 매출'
        ws.cell(row=1, column=41).style = toImportant
        ws.cell(row=1, column=41).border = border
        ws.cell(row=1, column=42).value = '인터파크 매출원가'
        ws.cell(row=1, column=42).style = toImportant
        ws.cell(row=1, column=42).border = border
        ws.cell(row=1, column=43).value = '인터파크 마진'
        ws.cell(row=1, column=43).style = toImportant
        ws.cell(row=1, column=43).border = border

        ws.cell(row=1, column=44).value = '쿠팡 거래액'
        ws.cell(row=1, column=44).style = toImportant
        ws.cell(row=1, column=44).border = border
        ws.cell(row=1, column=45).value = '쿠팡 판매수'
        ws.cell(row=1, column=45).style = toImportant
        ws.cell(row=1, column=45).border = border
        ws.cell(row=1, column=46).value = '쿠팡 객단가'
        ws.cell(row=1, column=46).style = toImportant
        ws.cell(row=1, column=46).border = border
        ws.cell(row=1, column=47).value = '쿠팡 매출'
        ws.cell(row=1, column=47).style = toImportant
        ws.cell(row=1, column=47).border = border
        ws.cell(row=1, column=48).value = '쿠팡 매출원가'
        ws.cell(row=1, column=48).style = toImportant
        ws.cell(row=1, column=48).border = border
        ws.cell(row=1, column=49).value = '쿠팡 마진'
        ws.cell(row=1, column=49).style = toImportant
        ws.cell(row=1, column=49).border = border

        ws.cell(row=1, column=50).value = 'SSG 거래액'
        ws.cell(row=1, column=50).style = toImportant
        ws.cell(row=1, column=50).border = border
        ws.cell(row=1, column=51).value = 'SSG 판매수'
        ws.cell(row=1, column=51).style = toImportant
        ws.cell(row=1, column=51).border = border
        ws.cell(row=1, column=52).value = 'SSG 객단가'
        ws.cell(row=1, column=52).style = toImportant
        ws.cell(row=1, column=52).border = border
        ws.cell(row=1, column=53).value = 'SSG 매출'
        ws.cell(row=1, column=53).style = toImportant
        ws.cell(row=1, column=53).border = border
        ws.cell(row=1, column=54).value = 'SSG 매출원가'
        ws.cell(row=1, column=54).style = toImportant
        ws.cell(row=1, column=54).border = border
        ws.cell(row=1, column=55).value = 'SSG 마진'
        ws.cell(row=1, column=55).style = toImportant
        ws.cell(row=1, column=55).border = border

        ws.cell(row=1, column=56).value = 'G9 거래액'
        ws.cell(row=1, column=56).style = toImportant
        ws.cell(row=1, column=56).border = border
        ws.cell(row=1, column=57).value = 'G9 판매수'
        ws.cell(row=1, column=57).style = toImportant
        ws.cell(row=1, column=57).border = border
        ws.cell(row=1, column=58).value = 'G9 객단가'
        ws.cell(row=1, column=58).style = toImportant
        ws.cell(row=1, column=58).border = border
        ws.cell(row=1, column=59).value = 'G9 매출'
        ws.cell(row=1, column=59).style = toImportant
        ws.cell(row=1, column=59).border = border
        ws.cell(row=1, column=60).value = 'G9 매출원가'
        ws.cell(row=1, column=60).style = toImportant
        ws.cell(row=1, column=60).border = border
        ws.cell(row=1, column=61).value = 'G9 마진'
        ws.cell(row=1, column=61).style = toImportant
        ws.cell(row=1, column=61).border = border

        ws.cell(row=1, column=62).value = '티몬 거래액'
        ws.cell(row=1, column=62).style = toImportant
        ws.cell(row=1, column=62).border = border
        ws.cell(row=1, column=63).value = '티몬 판매수'
        ws.cell(row=1, column=63).style = toImportant
        ws.cell(row=1, column=63).border = border
        ws.cell(row=1, column=64).value = '티몬 객단가'
        ws.cell(row=1, column=64).style = toImportant
        ws.cell(row=1, column=64).border = border
        ws.cell(row=1, column=65).value = '티몬 매출'
        ws.cell(row=1, column=65).style = toImportant
        ws.cell(row=1, column=65).border = border
        ws.cell(row=1, column=66).value = '티몬 매출원가'
        ws.cell(row=1, column=66).style = toImportant
        ws.cell(row=1, column=66).border = border
        ws.cell(row=1, column=67).value = '티몬 마진'
        ws.cell(row=1, column=67).style = toImportant
        ws.cell(row=1, column=67).border = border

        ws.cell(row=no, column=1).value = week
        ws.cell(row=no, column=1).border = border
        ws.cell(row=no, column=2).value = f'=sum(h{no}+n{no}+t{no}+z{no}+af{no}+al{no}+ar{no}+ax{no}+bd{no}+bj{no})'
        ws.cell(row=no, column=2).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=2).border = border
        ws.cell(row=no, column=3).value = f'=sum(i{no}+o{no}+u{no}+aa{no}+ag{no}+am{no}+as{no}+ay{no}+be{no}+bk{no})'
        ws.cell(row=no, column=3).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=3).border = border
        ws.cell(row=no, column=4).value = f'=(b{no}/c{no})'
        ws.cell(row=no, column=4).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=4).border = border
        ws.cell(row=no, column=5).value = f'=sum(k{no}+q{no}+w{no}+ac{no}+ai{no}+ao{no}+au{no}+ba{no}+bg{no}+bm{no})'
        ws.cell(row=no, column=5).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=5).border = border
        ws.cell(row=no, column=6).value = f'=sum(l{no}+r{no}+x{no}+ad{no}+aj{no}+ap{no}+av{no}+bb{no}+bh{no}+bn{no})'
        ws.cell(row=no, column=6).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=6).border = border
        ws.cell(row=no, column=7).value = f'=sum(m{no}+s{no}+y{no}+ae{no}+ak{no}+aq{no}+aw{no}+bc{no}+bi{no}+bo{no})'
        ws.cell(row=no, column=7).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=7).border = border

        ws.cell(row=no, column=8).value = brich_total_amount
        ws.cell(row=no, column=8).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=8).border = border
        ws.cell(row=no, column=9).value = brich_qty
        ws.cell(row=no, column=9).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=9).border = border
        ws.cell(row=no, column=10).value = brich_CT
        ws.cell(row=no, column=10).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=10).border = border
        ws.cell(row=no, column=11).value = brich_sales
        ws.cell(row=no, column=11).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=11).border = border
        ws.cell(row=no, column=12).value = brich_cogs
        ws.cell(row=no, column=12).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=12).border = border
        ws.cell(row=no, column=13).value = brich_margin
        ws.cell(row=no, column=13).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=13).border = border

        ws.cell(row=no, column=14).value = gmarket_total_amount
        ws.cell(row=no, column=14).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=14).border = border
        ws.cell(row=no, column=15).value = gmarket_qty
        ws.cell(row=no, column=15).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=15).border = border
        ws.cell(row=no, column=16).value = gmarket_CT
        ws.cell(row=no, column=16).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=16).border = border
        ws.cell(row=no, column=17).value = gmarket_sales
        ws.cell(row=no, column=17).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=17).border = border
        ws.cell(row=no, column=18).value = gmarket_cogs
        ws.cell(row=no, column=18).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=18).border = border
        ws.cell(row=no, column=19).value = gmarket_margin
        ws.cell(row=no, column=19).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=19).border = border

        ws.cell(row=no, column=20).value = auction_total_amount
        ws.cell(row=no, column=20).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=20).border = border
        ws.cell(row=no, column=21).value = auction_qty
        ws.cell(row=no, column=21).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=21).border = border
        ws.cell(row=no, column=22).value = auction_CT
        ws.cell(row=no, column=22).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=22).border = border
        ws.cell(row=no, column=23).value = auction_sales
        ws.cell(row=no, column=23).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=23).border = border
        ws.cell(row=no, column=24).value = auction_cogs
        ws.cell(row=no, column=24).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=24).border = border
        ws.cell(row=no, column=25).value = auction_margin
        ws.cell(row=no, column=25).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=25).border = border

        ws.cell(row=no, column=26).value = st_total_amount
        ws.cell(row=no, column=26).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=26).border = border
        ws.cell(row=no, column=27).value = st_qty
        ws.cell(row=no, column=27).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=27).border = border
        ws.cell(row=no, column=28).value = st_CT
        ws.cell(row=no, column=28).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=28).border = border
        ws.cell(row=no, column=29).value = st_sales
        ws.cell(row=no, column=29).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=29).border = border
        ws.cell(row=no, column=30).value = st_cogs
        ws.cell(row=no, column=30).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=30).border = border
        ws.cell(row=no, column=31).value = st_margin
        ws.cell(row=no, column=31).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=31).border = border

        ws.cell(row=no, column=32).value = wemakeprice_total_amount
        ws.cell(row=no, column=32).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=32).border = border
        ws.cell(row=no, column=33).value = wemakeprice_qty
        ws.cell(row=no, column=33).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=33).border = border
        ws.cell(row=no, column=34).value = wemakeprice_CT
        ws.cell(row=no, column=34).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=34).border = border
        ws.cell(row=no, column=35).value = wemakeprice_sales
        ws.cell(row=no, column=35).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=35).border = border
        ws.cell(row=no, column=36).value = wemakeprice_cogs
        ws.cell(row=no, column=36).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=36).border = border
        ws.cell(row=no, column=37).value = wemakeprice_margin
        ws.cell(row=no, column=37).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=37).border = border

        ws.cell(row=no, column=38).value = interpark_total_amount
        ws.cell(row=no, column=38).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=38).border = border
        ws.cell(row=no, column=39).value = interpark_qty
        ws.cell(row=no, column=39).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=39).border = border
        ws.cell(row=no, column=40).value = interpark_CT
        ws.cell(row=no, column=40).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=40).border = border
        ws.cell(row=no, column=41).value = interpark_sales
        ws.cell(row=no, column=41).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=41).border = border
        ws.cell(row=no, column=42).value = interpark_cogs
        ws.cell(row=no, column=42).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=42).border = border
        ws.cell(row=no, column=43).value = interpark_margin
        ws.cell(row=no, column=43).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=43).border = border

        ws.cell(row=no, column=44).value = coupnag_total_amount
        ws.cell(row=no, column=44).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=44).border = border
        ws.cell(row=no, column=45).value = coupnag_qty
        ws.cell(row=no, column=45).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=45).border = border
        ws.cell(row=no, column=46).value = coupnag_CT
        ws.cell(row=no, column=46).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=46).border = border
        ws.cell(row=no, column=47).value = coupnag_sales
        ws.cell(row=no, column=47).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=47).border = border
        ws.cell(row=no, column=48).value = coupnag_cogs
        ws.cell(row=no, column=48).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=48).border = border
        ws.cell(row=no, column=49).value = coupnag_margin
        ws.cell(row=no, column=49).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=49).border = border

        ws.cell(row=no, column=50).value = ssg_total_amount
        ws.cell(row=no, column=50).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=50).border = border
        ws.cell(row=no, column=51).value = ssg_qty
        ws.cell(row=no, column=51).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=51).border = border
        ws.cell(row=no, column=52).value = ssg_CT
        ws.cell(row=no, column=52).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=52).border = border
        ws.cell(row=no, column=53).value = ssg_sales
        ws.cell(row=no, column=53).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=53).border = border
        ws.cell(row=no, column=54).value = ssg_cogs
        ws.cell(row=no, column=54).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=54).border = border
        ws.cell(row=no, column=55).value = ssg_margin
        ws.cell(row=no, column=55).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=55).border = border

        ws.cell(row=no, column=56).value = g9_total_amount
        ws.cell(row=no, column=56).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=56).border = border
        ws.cell(row=no, column=57).value = g9_qty
        ws.cell(row=no, column=57).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=57).border = border
        ws.cell(row=no, column=58).value = g9_CT
        ws.cell(row=no, column=58).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=58).border = border
        ws.cell(row=no, column=59).value = g9_sales
        ws.cell(row=no, column=59).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=59).border = border
        ws.cell(row=no, column=60).value = g9_cogs
        ws.cell(row=no, column=60).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=60).border = border
        ws.cell(row=no, column=61).value = g9_margin
        ws.cell(row=no, column=61).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=61).border = border

        ws.cell(row=no, column=62).value = tmon_total_amount
        ws.cell(row=no, column=62).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=62).border = border
        ws.cell(row=no, column=63).value = tmon_qty
        ws.cell(row=no, column=63).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=63).border = border
        ws.cell(row=no, column=64).value = tmon_CT
        ws.cell(row=no, column=64).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=64).border = border
        ws.cell(row=no, column=65).value = tmon_sales
        ws.cell(row=no, column=65).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=65).border = border
        ws.cell(row=no, column=66).value = tmon_cogs
        ws.cell(row=no, column=66).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=66).border = border
        ws.cell(row=no, column=67).value = tmon_margin
        ws.cell(row=no, column=67).number_format = '#,##0;[red]-#,##0'
        ws.cell(row=no, column=67).border = border

        no += 1

    lastRow = ws.max_row
    firstRow = 1 + lastRow - len(rows)
    nowRow = lastRow + 1
    # total
    ws.cell(row=lastRow + 1, column=2).value = f'=sum(b{firstRow}:b{lastRow})'
    ws.cell(row=lastRow + 1, column=2).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=2).border = border
    ws.cell(row=lastRow + 1, column=3).value = f'=sum(c{firstRow}:c{lastRow})'
    ws.cell(row=lastRow + 1, column=3).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=3).border = border
    ws.cell(row=lastRow + 1, column=4).value = f'=average(d{firstRow}:d{lastRow})'
    ws.cell(row=lastRow + 1, column=4).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=4).border = border
    ws.cell(row=lastRow + 1, column=5).value = f'=sum(e{firstRow}:e{lastRow})'
    ws.cell(row=lastRow + 1, column=5).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=5).border = border
    ws.cell(row=lastRow + 1, column=6).value = f'=sum(f{firstRow}:f{lastRow})'
    ws.cell(row=lastRow + 1, column=6).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=6).border = border
    ws.cell(row=lastRow + 1, column=7).value = f'=sum(g{firstRow}:g{lastRow})'
    ws.cell(row=lastRow + 1, column=7).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=7).border = border
    # brich
    ws.cell(row=lastRow + 1, column=8).value = f'=sum(h{firstRow}:h{lastRow})'
    ws.cell(row=lastRow + 1, column=8).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=8).border = border
    ws.cell(row=lastRow + 1, column=9).value = f'=sum(i{firstRow}:i{lastRow})'
    ws.cell(row=lastRow + 1, column=9).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=9).border = border
    ws.cell(row=lastRow + 1, column=10).value = f'=average(j{firstRow}:j{lastRow})'
    ws.cell(row=lastRow + 1, column=10).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=10).border = border
    ws.cell(row=lastRow + 1, column=11).value = f'=sum(k{firstRow}:k{lastRow})'
    ws.cell(row=lastRow + 1, column=11).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=11).border = border
    ws.cell(row=lastRow + 1, column=12).value = f'=sum(l{firstRow}:l{lastRow})'
    ws.cell(row=lastRow + 1, column=12).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=12).border = border
    ws.cell(row=lastRow + 1, column=13).value = f'=sum(m{firstRow}:m{lastRow})'
    ws.cell(row=lastRow + 1, column=13).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=13).border = border
    # gmarket
    ws.cell(row=lastRow + 1, column=14).value = f'=sum(n{firstRow}:n{lastRow})'
    ws.cell(row=lastRow + 1, column=14).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=14).border = border
    ws.cell(row=lastRow + 1, column=15).value = f'=sum(o{firstRow}:o{lastRow})'
    ws.cell(row=lastRow + 1, column=15).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=15).border = border
    ws.cell(row=lastRow + 1, column=16).value = f'=average(p{firstRow}:p{lastRow})'
    ws.cell(row=lastRow + 1, column=16).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=16).border = border
    ws.cell(row=lastRow + 1, column=17).value = f'=sum(q{firstRow}:q{lastRow})'
    ws.cell(row=lastRow + 1, column=17).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=17).border = border
    ws.cell(row=lastRow + 1, column=18).value = f'=sum(r{firstRow}:r{lastRow})'
    ws.cell(row=lastRow + 1, column=18).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=18).border = border
    ws.cell(row=lastRow + 1, column=19).value = f'=sum(s{firstRow}:s{lastRow})'
    ws.cell(row=lastRow + 1, column=19).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=19).border = border
    # auction
    ws.cell(row=lastRow + 1, column=20).value = f'=sum(t{firstRow}:t{lastRow})'
    ws.cell(row=lastRow + 1, column=20).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=20).border = border
    ws.cell(row=lastRow + 1, column=21).value = f'=sum(u{firstRow}:u{lastRow})'
    ws.cell(row=lastRow + 1, column=21).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=21).border = border
    ws.cell(row=lastRow + 1, column=22).value = f'=average(v{firstRow}:v{lastRow})'
    ws.cell(row=lastRow + 1, column=22).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=22).border = border
    ws.cell(row=lastRow + 1, column=23).value = f'=sum(w{firstRow}:w{lastRow})'
    ws.cell(row=lastRow + 1, column=23).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=23).border = border
    ws.cell(row=lastRow + 1, column=24).value = f'=sum(x{firstRow}:x{lastRow})'
    ws.cell(row=lastRow + 1, column=24).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=24).border = border
    ws.cell(row=lastRow + 1, column=25).value = f'=sum(y{firstRow}:y{lastRow})'
    ws.cell(row=lastRow + 1, column=25).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=25).border = border

    # 11st
    ws.cell(row=lastRow + 1, column=26).value = f'=sum(z{firstRow}:z{lastRow})'
    ws.cell(row=lastRow + 1, column=26).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=26).border = border
    ws.cell(row=lastRow + 1, column=27).value = f'=sum(aa{firstRow}:aa{lastRow})'
    ws.cell(row=lastRow + 1, column=27).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=27).border = border
    ws.cell(row=lastRow + 1, column=28).value = f'=average(ab{firstRow}:ab{lastRow})'
    ws.cell(row=lastRow + 1, column=28).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=28).border = border
    ws.cell(row=lastRow + 1, column=29).value = f'=sum(ac{firstRow}:ac{lastRow})'
    ws.cell(row=lastRow + 1, column=29).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=29).border = border
    ws.cell(row=lastRow + 1, column=30).value = f'=sum(ad{firstRow}:ad{lastRow})'
    ws.cell(row=lastRow + 1, column=30).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=30).border = border
    ws.cell(row=lastRow + 1, column=31).value = f'=sum(ae{firstRow}:ae{lastRow})'
    ws.cell(row=lastRow + 1, column=31).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=31).border = border

    # wemakeprice
    ws.cell(row=lastRow + 1, column=32).value = f'=sum(af{firstRow}:af{lastRow})'
    ws.cell(row=lastRow + 1, column=32).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=32).border = border
    ws.cell(row=lastRow + 1, column=33).value = f'=sum(ag{firstRow}:ag{lastRow})'
    ws.cell(row=lastRow + 1, column=33).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=33).border = border
    ws.cell(row=lastRow + 1, column=34).value = f'=average(ah{firstRow}:ah{lastRow})'
    ws.cell(row=lastRow + 1, column=34).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=34).border = border
    ws.cell(row=lastRow + 1, column=35).value = f'=sum(ai{firstRow}:ai{lastRow})'
    ws.cell(row=lastRow + 1, column=35).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=35).border = border
    ws.cell(row=lastRow + 1, column=36).value = f'=sum(aj{firstRow}:aj{lastRow})'
    ws.cell(row=lastRow + 1, column=36).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=36).border = border
    ws.cell(row=lastRow + 1, column=37).value = f'=sum(ak{firstRow}:ak{lastRow})'
    ws.cell(row=lastRow + 1, column=37).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=37).border = border
    # interpark
    ws.cell(row=lastRow + 1, column=38).value = f'=sum(al{firstRow}:al{lastRow})'
    ws.cell(row=lastRow + 1, column=38).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=38).border = border
    ws.cell(row=lastRow + 1, column=39).value = f'=sum(am{firstRow}:am{lastRow})'
    ws.cell(row=lastRow + 1, column=39).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=39).border = border
    ws.cell(row=lastRow + 1, column=40).value = f'=average(an{firstRow}:an{lastRow})'
    ws.cell(row=lastRow + 1, column=40).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=40).border = border
    ws.cell(row=lastRow + 1, column=41).value = f'=sum(ao{firstRow}:ao{lastRow})'
    ws.cell(row=lastRow + 1, column=41).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=41).border = border
    ws.cell(row=lastRow + 1, column=42).value = f'=sum(ap{firstRow}:ap{lastRow})'
    ws.cell(row=lastRow + 1, column=42).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=42).border = border
    ws.cell(row=lastRow + 1, column=43).value = f'=sum(aq{firstRow}:aq{lastRow})'
    ws.cell(row=lastRow + 1, column=43).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=43).border = border
    # coupang
    ws.cell(row=lastRow + 1, column=44).value = f'=sum(ar{firstRow}:ar{lastRow})'
    ws.cell(row=lastRow + 1, column=44).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=44).border = border
    ws.cell(row=lastRow + 1, column=45).value = f'=sum(as{firstRow}:as{lastRow})'
    ws.cell(row=lastRow + 1, column=45).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=45).border = border
    ws.cell(row=lastRow + 1, column=46).value = f'=average(at{firstRow}:at{lastRow})'
    ws.cell(row=lastRow + 1, column=46).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=46).border = border
    ws.cell(row=lastRow + 1, column=47).value = f'=sum(au{firstRow}:au{lastRow})'
    ws.cell(row=lastRow + 1, column=47).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=47).border = border
    ws.cell(row=lastRow + 1, column=48).value = f'=sum(av{firstRow}:av{lastRow})'
    ws.cell(row=lastRow + 1, column=48).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=48).border = border
    ws.cell(row=lastRow + 1, column=49).value = f'=sum(aw{firstRow}:aw{lastRow})'
    ws.cell(row=lastRow + 1, column=49).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=49).border = border
    # ssg
    ws.cell(row=lastRow + 1, column=50).value = f'=sum(ax{firstRow}:ax{lastRow})'
    ws.cell(row=lastRow + 1, column=50).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=50).border = border
    ws.cell(row=lastRow + 1, column=51).value = f'=sum(ay{firstRow}:ay{lastRow})'
    ws.cell(row=lastRow + 1, column=51).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=51).border = border
    ws.cell(row=lastRow + 1, column=52).value = f'=average(az{firstRow}:az{lastRow})'
    ws.cell(row=lastRow + 1, column=52).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=52).border = border
    ws.cell(row=lastRow + 1, column=53).value = f'=sum(ba{firstRow}:ba{lastRow})'
    ws.cell(row=lastRow + 1, column=53).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=53).border = border
    ws.cell(row=lastRow + 1, column=54).value = f'=sum(bb{firstRow}:bb{lastRow})'
    ws.cell(row=lastRow + 1, column=54).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=54).border = border
    ws.cell(row=lastRow + 1, column=55).value = f'=sum(bc{firstRow}:bc{lastRow})'
    ws.cell(row=lastRow + 1, column=55).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=55).border = border
    # g9
    ws.cell(row=lastRow + 1, column=56).value = f'=sum(bd{firstRow}:bd{lastRow})'
    ws.cell(row=lastRow + 1, column=56).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=56).border = border
    ws.cell(row=lastRow + 1, column=57).value = f'=sum(be{firstRow}:be{lastRow})'
    ws.cell(row=lastRow + 1, column=57).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=57).border = border
    ws.cell(row=lastRow + 1, column=58).value = f'=average(bf{firstRow}:bf{lastRow})'
    ws.cell(row=lastRow + 1, column=58).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=58).border = border
    ws.cell(row=lastRow + 1, column=59).value = f'=sum(bg{firstRow}:bg{lastRow})'
    ws.cell(row=lastRow + 1, column=59).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=59).border = border
    ws.cell(row=lastRow + 1, column=60).value = f'=sum(bh{firstRow}:bh{lastRow})'
    ws.cell(row=lastRow + 1, column=60).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=60).border = border
    ws.cell(row=lastRow + 1, column=61).value = f'=sum(bi{firstRow}:bi{lastRow})'
    ws.cell(row=lastRow + 1, column=61).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=61).border = border
    # tmon
    ws.cell(row=lastRow + 1, column=62).value = f'=sum(bj{firstRow}:bj{lastRow})'
    ws.cell(row=lastRow + 1, column=62).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=62).border = border
    ws.cell(row=lastRow + 1, column=63).value = f'=sum(bk{firstRow}:bk{lastRow})'
    ws.cell(row=lastRow + 1, column=63).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=63).border = border
    ws.cell(row=lastRow + 1, column=64).value = f'=average(bl{firstRow}:bl{lastRow})'
    ws.cell(row=lastRow + 1, column=64).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=64).border = border
    ws.cell(row=lastRow + 1, column=65).value = f'=sum(bm{firstRow}:bm{lastRow})'
    ws.cell(row=lastRow + 1, column=65).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=65).border = border
    ws.cell(row=lastRow + 1, column=66).value = f'=sum(bn{firstRow}:bn{lastRow})'
    ws.cell(row=lastRow + 1, column=66).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=66).border = border
    ws.cell(row=lastRow + 1, column=67).value = f'=sum(bo{firstRow}:bo{lastRow})'
    ws.cell(row=lastRow + 1, column=67).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=lastRow + 1, column=67).border = border

    ws.cell(row=lastRow + 3, column=1).value = '일자'
    ws.cell(row=lastRow + 3, column=1).style = toImportant
    ws.cell(row=lastRow + 3, column=1).border = border
    ws.cell(row=lastRow + 3, column=2).value = '전체 거래액'
    ws.cell(row=lastRow + 3, column=2).style = toImportant
    ws.cell(row=lastRow + 3, column=2).border = border
    ws.cell(row=lastRow + 3, column=3).value = '전체 판매수'
    ws.cell(row=lastRow + 3, column=3).style = toImportant
    ws.cell(row=lastRow + 3, column=3).border = border
    ws.cell(row=lastRow + 3, column=4).value = '전체 객단가'
    ws.cell(row=lastRow + 3, column=4).style = toImportant
    ws.cell(row=lastRow + 3, column=4).border = border
    ws.cell(row=lastRow + 3, column=5).value = '전체 매출'
    ws.cell(row=lastRow + 3, column=5).style = toImportant
    ws.cell(row=lastRow + 3, column=5).border = border
    ws.cell(row=lastRow + 3, column=6).value = '전체 매출원가'
    ws.cell(row=lastRow + 3, column=6).style = toImportant
    ws.cell(row=lastRow + 3, column=6).border = border
    ws.cell(row=lastRow + 3, column=7).value = '전체 마진'
    ws.cell(row=lastRow + 3, column=7).style = toImportant
    ws.cell(row=lastRow + 3, column=7).border = border

    ws.cell(row=lastRow + 3, column=8).value = '브리치 거래액'
    ws.cell(row=lastRow + 3, column=8).style = toImportant
    ws.cell(row=lastRow + 3, column=8).border = border
    ws.cell(row=lastRow + 3, column=9).value = '브리치 판매수'
    ws.cell(row=lastRow + 3, column=9).style = toImportant
    ws.cell(row=lastRow + 3, column=9).border = border
    ws.cell(row=lastRow + 3, column=10).value = '브리치 객단가'
    ws.cell(row=lastRow + 3, column=10).style = toImportant
    ws.cell(row=lastRow + 3, column=10).border = border
    ws.cell(row=lastRow + 3, column=11).value = '브리치 매출'
    ws.cell(row=lastRow + 3, column=11).style = toImportant
    ws.cell(row=lastRow + 3, column=11).border = border
    ws.cell(row=lastRow + 3, column=12).value = '브리치 매출원가'
    ws.cell(row=lastRow + 3, column=12).style = toImportant
    ws.cell(row=lastRow + 3, column=12).border = border
    ws.cell(row=lastRow + 3, column=13).value = '브리치 마진'
    ws.cell(row=lastRow + 3, column=13).style = toImportant
    ws.cell(row=lastRow + 3, column=13).border = border

    ws.cell(row=lastRow + 3, column=14).value = '지마켓 거래액'
    ws.cell(row=lastRow + 3, column=14).style = toImportant
    ws.cell(row=lastRow + 3, column=14).border = border
    ws.cell(row=lastRow + 3, column=15).value = '지마켓 판매수'
    ws.cell(row=lastRow + 3, column=15).style = toImportant
    ws.cell(row=lastRow + 3, column=15).border = border
    ws.cell(row=lastRow + 3, column=16).value = '지마켓 객단가'
    ws.cell(row=lastRow + 3, column=16).style = toImportant
    ws.cell(row=lastRow + 3, column=16).border = border
    ws.cell(row=lastRow + 3, column=17).value = '지마켓 매출'
    ws.cell(row=lastRow + 3, column=17).style = toImportant
    ws.cell(row=lastRow + 3, column=17).border = border
    ws.cell(row=lastRow + 3, column=18).value = '지마켓 매출원가'
    ws.cell(row=lastRow + 3, column=18).style = toImportant
    ws.cell(row=lastRow + 3, column=18).border = border
    ws.cell(row=lastRow + 3, column=19).value = '지마켓 마진'
    ws.cell(row=lastRow + 3, column=19).style = toImportant
    ws.cell(row=lastRow + 3, column=19).border = border

    ws.cell(row=lastRow + 3, column=20).value = '옥션 거래액'
    ws.cell(row=lastRow + 3, column=20).style = toImportant
    ws.cell(row=lastRow + 3, column=20).border = border
    ws.cell(row=lastRow + 3, column=21).value = '옥션 판매수'
    ws.cell(row=lastRow + 3, column=21).style = toImportant
    ws.cell(row=lastRow + 3, column=21).border = border
    ws.cell(row=lastRow + 3, column=22).value = '옥션 객단가'
    ws.cell(row=lastRow + 3, column=22).style = toImportant
    ws.cell(row=lastRow + 3, column=22).border = border
    ws.cell(row=lastRow + 3, column=23).value = '옥션 매출'
    ws.cell(row=lastRow + 3, column=23).style = toImportant
    ws.cell(row=lastRow + 3, column=23).border = border
    ws.cell(row=lastRow + 3, column=24).value = '옥션 매출원가'
    ws.cell(row=lastRow + 3, column=24).style = toImportant
    ws.cell(row=lastRow + 3, column=24).border = border
    ws.cell(row=lastRow + 3, column=25).value = '옥션 마진'
    ws.cell(row=lastRow + 3, column=25).style = toImportant
    ws.cell(row=lastRow + 3, column=25).border = border

    ws.cell(row=lastRow + 3, column=26).value = '11번가 거래액'
    ws.cell(row=lastRow + 3, column=26).style = toImportant
    ws.cell(row=lastRow + 3, column=26).border = border
    ws.cell(row=lastRow + 3, column=27).value = '11번가 판매수'
    ws.cell(row=lastRow + 3, column=27).style = toImportant
    ws.cell(row=lastRow + 3, column=27).border = border
    ws.cell(row=lastRow + 3, column=28).value = '11번가 객단가'
    ws.cell(row=lastRow + 3, column=28).style = toImportant
    ws.cell(row=lastRow + 3, column=28).border = border
    ws.cell(row=lastRow + 3, column=29).value = '11번가 매출'
    ws.cell(row=lastRow + 3, column=29).style = toImportant
    ws.cell(row=lastRow + 3, column=29).border = border
    ws.cell(row=lastRow + 3, column=30).value = '11번가 매출원가'
    ws.cell(row=lastRow + 3, column=30).style = toImportant
    ws.cell(row=lastRow + 3, column=30).border = border
    ws.cell(row=lastRow + 3, column=31).value = '11번가 마진'
    ws.cell(row=lastRow + 3, column=31).style = toImportant
    ws.cell(row=lastRow + 3, column=31).border = border

    ws.cell(row=lastRow + 3, column=32).value = '위메프 거래액'
    ws.cell(row=lastRow + 3, column=32).style = toImportant
    ws.cell(row=lastRow + 3, column=32).border = border
    ws.cell(row=lastRow + 3, column=33).value = '위메프 판매수'
    ws.cell(row=lastRow + 3, column=33).style = toImportant
    ws.cell(row=lastRow + 3, column=33).border = border
    ws.cell(row=lastRow + 3, column=34).value = '위메프 객단가'
    ws.cell(row=lastRow + 3, column=34).style = toImportant
    ws.cell(row=lastRow + 3, column=34).border = border
    ws.cell(row=lastRow + 3, column=35).value = '위메프 매출'
    ws.cell(row=lastRow + 3, column=35).style = toImportant
    ws.cell(row=lastRow + 3, column=35).border = border
    ws.cell(row=lastRow + 3, column=36).value = '위메프 매출원가'
    ws.cell(row=lastRow + 3, column=36).style = toImportant
    ws.cell(row=lastRow + 3, column=36).border = border
    ws.cell(row=lastRow + 3, column=37).value = '위메프 마진'
    ws.cell(row=lastRow + 3, column=37).style = toImportant
    ws.cell(row=lastRow + 3, column=37).border = border

    ws.cell(row=lastRow + 3, column=38).value = '인터파크 거래액'
    ws.cell(row=lastRow + 3, column=38).style = toImportant
    ws.cell(row=lastRow + 3, column=38).border = border
    ws.cell(row=lastRow + 3, column=39).value = '인터파크 판매수'
    ws.cell(row=lastRow + 3, column=39).style = toImportant
    ws.cell(row=lastRow + 3, column=39).border = border
    ws.cell(row=lastRow + 3, column=40).value = '인터파크 객단가'
    ws.cell(row=lastRow + 3, column=40).style = toImportant
    ws.cell(row=lastRow + 3, column=40).border = border
    ws.cell(row=lastRow + 3, column=41).value = '인터파크 매출'
    ws.cell(row=lastRow + 3, column=41).style = toImportant
    ws.cell(row=lastRow + 3, column=41).border = border
    ws.cell(row=lastRow + 3, column=42).value = '인터파크 매출원가'
    ws.cell(row=lastRow + 3, column=42).style = toImportant
    ws.cell(row=lastRow + 3, column=42).border = border
    ws.cell(row=lastRow + 3, column=43).value = '인터파크 마진'
    ws.cell(row=lastRow + 3, column=43).style = toImportant
    ws.cell(row=lastRow + 3, column=43).border = border

    ws.cell(row=lastRow + 3, column=44).value = '쿠팡 거래액'
    ws.cell(row=lastRow + 3, column=44).style = toImportant
    ws.cell(row=lastRow + 3, column=44).border = border
    ws.cell(row=lastRow + 3, column=45).value = '쿠팡 판매수'
    ws.cell(row=lastRow + 3, column=45).style = toImportant
    ws.cell(row=lastRow + 3, column=45).border = border
    ws.cell(row=lastRow + 3, column=46).value = '쿠팡 객단가'
    ws.cell(row=lastRow + 3, column=46).style = toImportant
    ws.cell(row=lastRow + 3, column=46).border = border
    ws.cell(row=lastRow + 3, column=47).value = '쿠팡 매출'
    ws.cell(row=lastRow + 3, column=47).style = toImportant
    ws.cell(row=lastRow + 3, column=47).border = border
    ws.cell(row=lastRow + 3, column=48).value = '쿠팡 매출원가'
    ws.cell(row=lastRow + 3, column=48).style = toImportant
    ws.cell(row=lastRow + 3, column=48).border = border
    ws.cell(row=lastRow + 3, column=49).value = '쿠팡 마진'
    ws.cell(row=lastRow + 3, column=49).style = toImportant
    ws.cell(row=lastRow + 3, column=49).border = border

    ws.cell(row=lastRow + 3, column=50).value = 'SSG 거래액'
    ws.cell(row=lastRow + 3, column=50).style = toImportant
    ws.cell(row=lastRow + 3, column=50).border = border
    ws.cell(row=lastRow + 3, column=51).value = 'SSG 판매수'
    ws.cell(row=lastRow + 3, column=51).style = toImportant
    ws.cell(row=lastRow + 3, column=51).border = border
    ws.cell(row=lastRow + 3, column=52).value = 'SSG 객단가'
    ws.cell(row=lastRow + 3, column=52).style = toImportant
    ws.cell(row=lastRow + 3, column=52).border = border
    ws.cell(row=lastRow + 3, column=53).value = 'SSG 매출'
    ws.cell(row=lastRow + 3, column=53).style = toImportant
    ws.cell(row=lastRow + 3, column=53).border = border
    ws.cell(row=lastRow + 3, column=54).value = 'SSG 매출원가'
    ws.cell(row=lastRow + 3, column=54).style = toImportant
    ws.cell(row=lastRow + 3, column=54).border = border
    ws.cell(row=lastRow + 3, column=55).value = 'SSG 마진'
    ws.cell(row=lastRow + 3, column=55).style = toImportant
    ws.cell(row=lastRow + 3, column=55).border = border

    ws.cell(row=lastRow + 3, column=56).value = 'G9 거래액'
    ws.cell(row=lastRow + 3, column=56).style = toImportant
    ws.cell(row=lastRow + 3, column=56).border = border
    ws.cell(row=lastRow + 3, column=57).value = 'G9 판매수'
    ws.cell(row=lastRow + 3, column=57).style = toImportant
    ws.cell(row=lastRow + 3, column=57).border = border
    ws.cell(row=lastRow + 3, column=58).value = 'G9 객단가'
    ws.cell(row=lastRow + 3, column=58).style = toImportant
    ws.cell(row=lastRow + 3, column=58).border = border
    ws.cell(row=lastRow + 3, column=59).value = 'G9 매출'
    ws.cell(row=lastRow + 3, column=59).style = toImportant
    ws.cell(row=lastRow + 3, column=59).border = border
    ws.cell(row=lastRow + 3, column=60).value = 'G9 매출원가'
    ws.cell(row=lastRow + 3, column=60).style = toImportant
    ws.cell(row=lastRow + 3, column=60).border = border
    ws.cell(row=lastRow + 3, column=61).value = 'G9 마진'
    ws.cell(row=lastRow + 3, column=61).style = toImportant
    ws.cell(row=lastRow + 3, column=61).border = border

    ws.cell(row=lastRow + 3, column=62).value = '티몬 거래액'
    ws.cell(row=lastRow + 3, column=62).style = toImportant
    ws.cell(row=lastRow + 3, column=62).border = border
    ws.cell(row=lastRow + 3, column=63).value = '티몬 판매수'
    ws.cell(row=lastRow + 3, column=63).style = toImportant
    ws.cell(row=lastRow + 3, column=63).border = border
    ws.cell(row=lastRow + 3, column=64).value = '티몬 객단가'
    ws.cell(row=lastRow + 3, column=64).style = toImportant
    ws.cell(row=lastRow + 3, column=64).border = border
    ws.cell(row=lastRow + 3, column=65).value = '티몬 매출'
    ws.cell(row=lastRow + 3, column=65).style = toImportant
    ws.cell(row=lastRow + 3, column=65).border = border
    ws.cell(row=lastRow + 3, column=66).value = '티몬 매출원가'
    ws.cell(row=lastRow + 3, column=66).style = toImportant
    ws.cell(row=lastRow + 3, column=66).border = border
    ws.cell(row=lastRow + 3, column=67).value = '티몬 마진'
    ws.cell(row=lastRow + 3, column=67).style = toImportant
    ws.cell(row=lastRow + 3, column=67).border = border

weekSql = f'''
    SELECT 
    week,
    min(date),
    max(date),
    sum(brich_total_amount),
    sum(brich_qty),
    sum(brich_sales),
    sum(brich_cogs),
    sum(gmarket_total_amount),
    sum(gmarket_qty),
    sum(gmarket_sales),
    sum(gmarket_cogs),
    sum(auction_total_amount),
    sum(auction_qty),
    sum(auction_sales),
    sum(auction_cogs),
    sum(11st_total_amount),
    sum(11st_qty),
    sum(11st_sales),
    sum(11st_cogs),
    sum(wemakeprice_total_amount),
    sum(wemakeprice_qty),
    sum(wemakeprice_sales),
    sum(wemakeprice_cogs),
    sum(interpark_total_amount),
    sum(interpark_qty),
    sum(interpark_sales),
    sum(interpark_cogs),
    sum(coupang_total_amount),
    sum(coupang_qty),
    sum(coupang_sales),
    sum(coupang_cogs),
    sum(ssg_total_amount),
    sum(ssg_qty),
    sum(ssg_sales),
    sum(ssg_cogs),
    sum(g9_total_amount),
    sum(g9_qty),
    sum(g9_sales),
    sum(g9_cogs),
    sum(tmon_total_amount),
    sum(tmon_qty),
    sum(tmon_sales),
    sum(tmon_cogs)
    FROM sell_to_channel where month >= {startMonth} and year = {year} GROUP BY week
'''
weekStartRow = ws.max_row + 1

cursor.execute(weekSql)
weekRows = cursor.fetchall()

for weekRow in weekRows:
    week = weekRow[0]
    weekstr = datetime.strftime(weekRow[1], '%Y-%m-%d') + "~" + datetime.strftime(weekRow[2], '%Y-%m-%d')
    brich_total_amount = weekRow[3]
    brich_qty = weekRow[4]
    brich_CT = noZerodiv(brich_qty, brich_total_amount)
    brich_sales = weekRow[5]
    brich_cogs = weekRow[6]
    brich_margin = noneCheck(brich_sales, brich_cogs)
    gmarket_total_amount = weekRow[7]
    gmarket_qty = weekRow[8]
    gmarket_CT = noZerodiv(gmarket_qty, gmarket_total_amount)
    gmarket_sales = weekRow[9]
    gmarket_cogs = weekRow[10]
    gmarket_margin = noneCheck(gmarket_sales, gmarket_cogs)
    auction_total_amount = weekRow[11]
    auction_qty = weekRow[12]
    auction_CT = noZerodiv(auction_qty, auction_total_amount)
    auction_sales = weekRow[13]
    auction_cogs = weekRow[14]
    auction_margin = noneCheck(auction_sales, auction_cogs)
    st_total_amount = weekRow[15]
    st_qty = weekRow[16]
    st_CT = noZerodiv(st_qty, st_total_amount)
    st_sales = weekRow[17]
    st_cogs = weekRow[18]
    st_margin = noneCheck(st_sales, st_cogs)
    wemakeprice_total_amount = weekRow[19]
    wemakeprice_qty = weekRow[20]
    wemakeprice_CT = noZerodiv(wemakeprice_qty, wemakeprice_total_amount)
    wemakeprice_sales = weekRow[21]
    wemakeprice_cogs = weekRow[22]
    wemakeprice_margin = noneCheck(wemakeprice_sales, wemakeprice_cogs)
    interpark_total_amount = weekRow[23]
    interpark_qty = weekRow[24]
    interpark_CT = noZerodiv(interpark_qty, interpark_total_amount)
    interpark_sales = weekRow[25]
    interpark_cogs = weekRow[26]
    interpark_margin = noneCheck(interpark_sales, interpark_cogs)
    coupnag_total_amount = weekRow[27]
    coupnag_qty = weekRow[28]
    coupnag_CT = noZerodiv(coupnag_qty, coupnag_total_amount)
    coupnag_sales = weekRow[29]
    coupnag_cogs = weekRow[30]
    coupnag_margin = noneCheck(coupnag_sales, coupnag_cogs)
    ssg_total_amount = weekRow[31]
    ssg_qty = weekRow[32]
    ssg_CT = noZerodiv(ssg_qty, ssg_total_amount)
    ssg_sales = weekRow[33]
    ssg_cogs = weekRow[34]
    ssg_margin = noneCheck(ssg_sales, ssg_cogs)
    g9_total_amount = weekRow[35]
    g9_qty = weekRow[36]
    g9_CT = noZerodiv(g9_qty, g9_total_amount)
    g9_sales = weekRow[37]
    g9_cogs = weekRow[38]
    g9_margin = noneCheck(g9_sales, g9_cogs)
    tmon_total_amount = weekRow[39]
    tmon_qty = weekRow[40]
    tmon_CT = noZerodiv(tmon_qty, tmon_total_amount)
    tmon_sales = weekRow[41]
    tmon_cogs = weekRow[42]
    tmon_margin = noneCheck(tmon_sales, tmon_cogs)

    ws.cell(row=weekStartRow, column=1).value = weekstr
    ws.cell(row=weekStartRow, column=1).border = border
    ws.cell(row=weekStartRow, column=2).value = f'=sum(h{weekStartRow}+n{weekStartRow}+t{weekStartRow}+z{weekStartRow}+af{weekStartRow}+al{weekStartRow}+ar{weekStartRow}+ax{weekStartRow}+bd{weekStartRow}+bj{weekStartRow})'
    ws.cell(row=weekStartRow, column=2).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=2).border = border
    ws.cell(row=weekStartRow, column=3).value = f'=sum(i{weekStartRow}+o{weekStartRow}+u{weekStartRow}+aa{weekStartRow}+ag{weekStartRow}+am{weekStartRow}+as{weekStartRow}+ay{weekStartRow}+be{weekStartRow}+bk{weekStartRow})'
    ws.cell(row=weekStartRow, column=3).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=3).border = border
    ws.cell(row=weekStartRow, column=4).value = f'=(b{weekStartRow}/c{weekStartRow})'
    ws.cell(row=weekStartRow, column=4).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=4).border = border
    ws.cell(row=weekStartRow, column=5).value = f'=sum(k{weekStartRow}+q{weekStartRow}+w{weekStartRow}+ac{weekStartRow}+ai{weekStartRow}+ao{weekStartRow}+au{weekStartRow}+ba{weekStartRow}+bg{weekStartRow}+bm{weekStartRow})'
    ws.cell(row=weekStartRow, column=5).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=5).border = border
    ws.cell(row=weekStartRow, column=6).value = f'=sum(l{weekStartRow}+r{weekStartRow}+x{weekStartRow}+ad{weekStartRow}+aj{weekStartRow}+ap{weekStartRow}+av{weekStartRow}+bb{weekStartRow}+bh{weekStartRow}+bn{weekStartRow})'
    ws.cell(row=weekStartRow, column=6).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=6).border = border
    ws.cell(row=weekStartRow, column=7).value = f'=sum(m{weekStartRow}+s{weekStartRow}+y{weekStartRow}+ae{weekStartRow}+ak{weekStartRow}+aq{weekStartRow}+aw{weekStartRow}+bc{weekStartRow}+bi{weekStartRow}+bo{weekStartRow})'
    ws.cell(row=weekStartRow, column=7).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=7).border = border

    ws.cell(row=weekStartRow, column=8).value = brich_total_amount
    ws.cell(row=weekStartRow, column=8).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=8).border = border
    ws.cell(row=weekStartRow, column=9).value = brich_qty
    ws.cell(row=weekStartRow, column=9).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=9).border = border
    ws.cell(row=weekStartRow, column=10).value = brich_CT
    ws.cell(row=weekStartRow, column=10).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=10).border = border
    ws.cell(row=weekStartRow, column=11).value = brich_sales
    ws.cell(row=weekStartRow, column=11).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=11).border = border
    ws.cell(row=weekStartRow, column=12).value = brich_cogs
    ws.cell(row=weekStartRow, column=12).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=12).border = border
    ws.cell(row=weekStartRow, column=13).value = brich_margin
    ws.cell(row=weekStartRow, column=13).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=13).border = border

    ws.cell(row=weekStartRow, column=14).value = gmarket_total_amount
    ws.cell(row=weekStartRow, column=14).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=14).border = border
    ws.cell(row=weekStartRow, column=15).value = gmarket_qty
    ws.cell(row=weekStartRow, column=15).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=15).border = border
    ws.cell(row=weekStartRow, column=16).value = gmarket_CT
    ws.cell(row=weekStartRow, column=16).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=16).border = border
    ws.cell(row=weekStartRow, column=17).value = gmarket_sales
    ws.cell(row=weekStartRow, column=17).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=17).border = border
    ws.cell(row=weekStartRow, column=18).value = gmarket_cogs
    ws.cell(row=weekStartRow, column=18).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=18).border = border
    ws.cell(row=weekStartRow, column=19).value = gmarket_margin
    ws.cell(row=weekStartRow, column=19).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=19).border = border

    ws.cell(row=weekStartRow, column=20).value = auction_total_amount
    ws.cell(row=weekStartRow, column=20).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=20).border = border
    ws.cell(row=weekStartRow, column=21).value = auction_qty
    ws.cell(row=weekStartRow, column=21).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=21).border = border
    ws.cell(row=weekStartRow, column=22).value = auction_CT
    ws.cell(row=weekStartRow, column=22).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=22).border = border
    ws.cell(row=weekStartRow, column=23).value = auction_sales
    ws.cell(row=weekStartRow, column=23).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=23).border = border
    ws.cell(row=weekStartRow, column=24).value = auction_cogs
    ws.cell(row=weekStartRow, column=24).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=24).border = border
    ws.cell(row=weekStartRow, column=25).value = auction_margin
    ws.cell(row=weekStartRow, column=25).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=25).border = border

    ws.cell(row=weekStartRow, column=26).value = st_total_amount
    ws.cell(row=weekStartRow, column=26).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=26).border = border
    ws.cell(row=weekStartRow, column=27).value = st_qty
    ws.cell(row=weekStartRow, column=27).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=27).border = border
    ws.cell(row=weekStartRow, column=28).value = st_CT
    ws.cell(row=weekStartRow, column=28).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=28).border = border
    ws.cell(row=weekStartRow, column=29).value = st_sales
    ws.cell(row=weekStartRow, column=29).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=29).border = border
    ws.cell(row=weekStartRow, column=30).value = st_cogs
    ws.cell(row=weekStartRow, column=30).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=30).border = border
    ws.cell(row=weekStartRow, column=31).value = st_margin
    ws.cell(row=weekStartRow, column=31).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=31).border = border

    ws.cell(row=weekStartRow, column=32).value = wemakeprice_total_amount
    ws.cell(row=weekStartRow, column=32).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=32).border = border
    ws.cell(row=weekStartRow, column=33).value = wemakeprice_qty
    ws.cell(row=weekStartRow, column=33).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=33).border = border
    ws.cell(row=weekStartRow, column=34).value = wemakeprice_CT
    ws.cell(row=weekStartRow, column=34).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=34).border = border
    ws.cell(row=weekStartRow, column=35).value = wemakeprice_sales
    ws.cell(row=weekStartRow, column=35).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=35).border = border
    ws.cell(row=weekStartRow, column=36).value = wemakeprice_cogs
    ws.cell(row=weekStartRow, column=36).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=36).border = border
    ws.cell(row=weekStartRow, column=37).value = wemakeprice_margin
    ws.cell(row=weekStartRow, column=37).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=37).border = border

    ws.cell(row=weekStartRow, column=38).value = interpark_total_amount
    ws.cell(row=weekStartRow, column=38).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=38).border = border
    ws.cell(row=weekStartRow, column=39).value = interpark_qty
    ws.cell(row=weekStartRow, column=39).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=39).border = border
    ws.cell(row=weekStartRow, column=40).value = interpark_CT
    ws.cell(row=weekStartRow, column=40).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=40).border = border
    ws.cell(row=weekStartRow, column=41).value = interpark_sales
    ws.cell(row=weekStartRow, column=41).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=41).border = border
    ws.cell(row=weekStartRow, column=42).value = interpark_cogs
    ws.cell(row=weekStartRow, column=42).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=42).border = border
    ws.cell(row=weekStartRow, column=43).value = interpark_margin
    ws.cell(row=weekStartRow, column=43).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=43).border = border

    ws.cell(row=weekStartRow, column=44).value = coupnag_total_amount
    ws.cell(row=weekStartRow, column=44).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=44).border = border
    ws.cell(row=weekStartRow, column=45).value = coupnag_qty
    ws.cell(row=weekStartRow, column=45).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=45).border = border
    ws.cell(row=weekStartRow, column=46).value = coupnag_CT
    ws.cell(row=weekStartRow, column=46).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=46).border = border
    ws.cell(row=weekStartRow, column=47).value = coupnag_sales
    ws.cell(row=weekStartRow, column=47).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=47).border = border
    ws.cell(row=weekStartRow, column=48).value = coupnag_cogs
    ws.cell(row=weekStartRow, column=48).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=48).border = border
    ws.cell(row=weekStartRow, column=49).value = coupnag_margin
    ws.cell(row=weekStartRow, column=49).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=49).border = border

    ws.cell(row=weekStartRow, column=50).value = ssg_total_amount
    ws.cell(row=weekStartRow, column=50).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=50).border = border
    ws.cell(row=weekStartRow, column=51).value = ssg_qty
    ws.cell(row=weekStartRow, column=51).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=51).border = border
    ws.cell(row=weekStartRow, column=52).value = ssg_CT
    ws.cell(row=weekStartRow, column=52).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=52).border = border
    ws.cell(row=weekStartRow, column=53).value = ssg_sales
    ws.cell(row=weekStartRow, column=53).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=53).border = border
    ws.cell(row=weekStartRow, column=54).value = ssg_cogs
    ws.cell(row=weekStartRow, column=54).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=54).border = border
    ws.cell(row=weekStartRow, column=55).value = ssg_margin
    ws.cell(row=weekStartRow, column=55).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=55).border = border

    ws.cell(row=weekStartRow, column=56).value = g9_total_amount
    ws.cell(row=weekStartRow, column=56).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=56).border = border
    ws.cell(row=weekStartRow, column=57).value = g9_qty
    ws.cell(row=weekStartRow, column=57).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=57).border = border
    ws.cell(row=weekStartRow, column=58).value = g9_CT
    ws.cell(row=weekStartRow, column=58).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=58).border = border
    ws.cell(row=weekStartRow, column=59).value = g9_sales
    ws.cell(row=weekStartRow, column=59).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=59).border = border
    ws.cell(row=weekStartRow, column=60).value = g9_cogs
    ws.cell(row=weekStartRow, column=60).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=60).border = border
    ws.cell(row=weekStartRow, column=61).value = g9_margin
    ws.cell(row=weekStartRow, column=61).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=61).border = border

    ws.cell(row=weekStartRow, column=62).value = tmon_total_amount
    ws.cell(row=weekStartRow, column=62).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=62).border = border
    ws.cell(row=weekStartRow, column=63).value = tmon_qty
    ws.cell(row=weekStartRow, column=63).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=63).border = border
    ws.cell(row=weekStartRow, column=64).value = tmon_CT
    ws.cell(row=weekStartRow, column=64).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=64).border = border
    ws.cell(row=weekStartRow, column=65).value = tmon_sales
    ws.cell(row=weekStartRow, column=65).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=65).border = border
    ws.cell(row=weekStartRow, column=66).value = tmon_cogs
    ws.cell(row=weekStartRow, column=66).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=66).border = border
    ws.cell(row=weekStartRow, column=67).value = tmon_margin
    ws.cell(row=weekStartRow, column=67).number_format = '#,##0;[red]-#,##0'
    ws.cell(row=weekStartRow, column=67).border = border

    weekStartRow += 1

# # 샐 너비 변
# for col in ws.columns:
#     max_length = 0
#     columnIndex = col[0].column
#     column = get_column_letter(columnIndex)
#     for cell in col:
#         if max_length < len(str(cell.value)) < 30:
#             max_length = len(str(cell.value))
#         else:
#             pass
#     ws.column_dimensions[column].width = (max_length + 1) * 1.2

makeToday = datetime.today()

now = makeToday.strftime("%m%d_%H%M")
result = "2019_운영지표" + "_" + now + ".xlsx"
print(result)
wb.save(result)
cursor.close()
db.close()