from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import config
import time
import os
from datetime import date
from datetime import datetime
import pyexcel as p
import pymysql
from openpyxl import load_workbook
from tqdm import tqdm
import dateutil.relativedelta
import re
from openpyxl import Workbook



def replacedate(text):
    if text is None:
        return
    else:
        text = text[0:10]
        return text.strip()


def replacenone(text):
    if text is None:
        return
    else:
        text = str(text)
        return text.strip()


def replaceint(text):
    if text is None:
        return
    else:
        text = int(text)
        return text


makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")


def startCrawler(downloadPath, driverPath):
    options = Options()
    # options.add_argument('--headless')
    # options.add_argument("disable-gpu")
    prefs = {
        "download.default_directory": downloadPath,
        "directory_upgrade": True
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(executable_path=driverPath, options=options)

    # driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    # params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "/path/to/download/dir"}}
    # command_result = driver.execute("send_command", params)

    driver.get('https://login.11st.co.kr/auth/front/selleroffice/login.tmall')

    driver.find_element_by_id('user-id').send_keys(config.ST_LOGIN['id'])
    driver.find_element_by_id('passWord').send_keys(config.ST_LOGIN['password'])
    driver.find_element_by_xpath('/html/body/div/form[1]/fieldset/button').click()
    time.sleep(5)
    driver.get('https://soffice.11st.co.kr/escrow/OrderCancelManageList2nd.tmall')
    time.sleep(5)

    def findSelect(xpath, value):
        el = Select(driver.find_element_by_xpath(xpath))
        el.select_by_value(value)

    findSelect('//select[@id="key"]', '02')
    findSelect('//select[@id="shDateType"]', '07')
    findSelect('//select[@id="sltDuration"]', 'TODAY')
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="search_area"]/form/div[1]/div[2]/div[2]/div[2]/div/button[1]').click()
    time.sleep(3)

    driver.find_element_by_xpath('//*[@id="search_area"]/form/div[3]/div/a').click()
    time.sleep(2)
    driver.get('https://soffice.11st.co.kr/escrow/OrderingLogistics.tmall')
    time.sleep(2)
    findUp = driver.find_element_by_id('goDlvTmpltPopup')
    driver.switch_to.frame(findUp.find_element_by_tag_name('iframe'))
    driver.find_element_by_xpath('//*[@id="ext-gen6"]/div/button').click()
    time.sleep(1)
    driver.switch_to.default_content()
    driver.find_element_by_xpath('//*[@id="order_good_301"]').click()
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="searchform"]/div/div[1]/div[6]/div/a[2]').click()
    time.sleep(2)

    print(driver.window_handles)
    driver.switch_to.window(driver.window_handles[1])
    driver.find_element_by_xpath('/html/body/div/div[2]/div[4]/div/a[1]').click()
    time.sleep(5)
    driver.close()
    print(driver.window_handles)
    driver.switch_to.window(driver.window_handles[0])
    driver.get('https://partner.brich.co.kr/login')
    driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/button[2]').click()
    time.sleep(2)
    driver.find_element_by_xpath(
        '//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[1]').send_keys(
        config.BFLOW_LOGIN['id'])
    driver.find_element_by_xpath(
        '//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div/input[2]').send_keys(
        config.BFLOW_LOGIN['password'])
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div/div/div[2]/div[2]/button[1]').click()
    time.sleep(4)
    driver.minimize_window()
    # driver.get('https://partner.brich.co.kr/order/all#/')
    # time.sleep(5)
    # driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div/div[1]/div[2]/div/div[5]/div/div[5]/input').click()
    # driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div/div[1]/div[2]/div/div[10]/div/span/button[5]').click()
    # driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div/div[1]/div[2]/div/div[11]/div/button[1]').click()
    # time.sleep(5)
    # driver.find_element_by_xpath('//*[@id="app"]/div[1]/div/div/div[2]/div[2]/div[2]/button[2]').click()

    driver.get(f'''
        https://partner.brich.co.kr/api/orders-excel-download?type=order&start={endNow}
        &end={totalNow}&condition=code&content=&channel%5B%5D=11st&period=orders.created_at&orderby=orders.created_at&
        per_page=10&selectedProviderOptimusId=&selectedBrandOptimusId=&selectedCrawlingTarget=&productName=&
        productId=&isTodayDelivery=&isHold=&coupon_optimus_id=&refererDomain=''')
    time.sleep(60)
    driver.quit()


def changeFileToXlsx(originalName, resultName):
    stOriExcel = config.ST_LOGIN['excelPath'] + originalName
    stResultExcel = config.ST_LOGIN['excelPath'] + resultName + now + '.xls'
    stResultXlsx = config.ST_LOGIN['excelPath'] + resultName + now + '.xlsx'

    os.rename(stOriExcel, stResultExcel)

    p.save_book_as(file_name=stResultExcel, dest_file_name=stResultXlsx)
    return stResultXlsx


def bflowChangeFile(originalName, resultName):
    bflowOriExcel = config.ST_LOGIN['excelPath'] + originalName + endNow + "_" + totalNow + ".xlsx"
    bflowResultExcel = config.ST_LOGIN['excelPath'] + resultName + now + '.xlsx'

    os.rename(bflowOriExcel, bflowResultExcel)
    return bflowResultExcel


startCrawler(config.ST_LOGIN['excelPath'], '../py_option/chromedriver')
cancelFile = changeFileToXlsx('39731068_sellListlistType.xls', '11st_Cancel_')
logiFile = changeFileToXlsx('39731068_logistics.xls', '11st_logi_')
bflowFile = bflowChangeFile('orders_', '11st_Cancel_order')

cancelList = (
        'state',
        'channel_order_number',
        'channel_order_list',
        'claim_request',
        'claim_complete',
        'product_name',
        'product_option',
        'quantity',
        'order_amount',
        'cancel_reason',
        'cancel_detail_reason',
        'cancel_response',
        'add_delivery_fees',
        'cancel_complete_date',
        'payment_at',
        'product_amount',
        'product_option_amount',
        'cancel_complete_user')


def insertXlsxtoDb(fileName, tableName, columnLists):
    path = fileName
    wb = load_workbook(path)

    ws = wb.active

    db = pymysql.connect(
        host=config.DATABASE_CONFIG['host'],
        user=config.DATABASE_CONFIG['user'],
        password=config.DATABASE_CONFIG['password'],
        db=config.DATABASE_CONFIG['db'],
        charset=config.DATABASE_CONFIG['charset'],
        autocommit=True)
    cursor = db.cursor()
    print(columnLists)
    sql = f'INSERT INTO `excel`.`{tableName}`'
    value = 'VALUES (' + '%s,' * len(columnLists) + ')ON DUPLICATE KEY UPDATE'
    ''' state = %s, claim_request = %s, claim_complete = %s, cancel_reason = %s,
            cancel_detail_reason = %s, cancel_response =%s, add_delivery_fees = %s, cancel_complete_date = %s,
            cancel_complete_user = %s
             '''
    print(sql, columnLists, value)
    resultSql = sql + str(columnLists) + value
    print(resultSql)

    maxRow = ws.max_row - 2

    for row in ws.iter_rows(min_row=7, max_row=maxRow):
        state = replacenone(row[1].value)
        channel_order_number = replacenone(row[2].value)
        channel_order_list = replaceint(row[3].value)
        claim_request = row[4].value
        claim_complete = row[5].value
        product_name = replacenone(row[6].value)
        product_option = replacenone(row[7].value)
        quantity = replaceint(row[8].value)
        order_amount = replaceint(row[13].value)
        cancel_reason = replacenone(row[16].value)
        cancel_detail_reason = replacenone(row[17].value)
        cancel_response = replacenone(row[18].value)
        add_delivery_fees = replaceint(row[19].value)
        cancel_complete_date = replacedate(row[20].value)
        payment_at = row[31].value
        product_amount = replaceint(row[33].value)
        product_option_amount = replaceint(row[34].value)
        cancel_complete_user = replacenone(row[37].value)

        values = (
            state,
            channel_order_number,
            channel_order_list,
            claim_request,
            claim_complete,
            product_name,
            product_option,
            quantity,
            order_amount,
            cancel_reason,
            cancel_detail_reason,
            cancel_response,
            add_delivery_fees,
            cancel_complete_date,
            payment_at,
            product_amount,
            product_option_amount,
            cancel_complete_user,
            state,
            claim_request,
            claim_complete,
            cancel_reason,
            cancel_detail_reason,
            cancel_response,
            add_delivery_fees,
            cancel_complete_date,
            cancel_complete_user
        )
        cursor.execute(resultSql, values)
        print(resultSql, values)


path = logiFile

wb = load_workbook(path)

ws = wb.active

orderSql = '''
        INSERT INTO `excel`.`11st_order` (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE state = %s
        '''
maxRow = ws.max_row - 2

for row in ws.iter_rows(min_row=3):
    state = replacenone(row[1].value)
    channel_order_number = replacenone(row[2].value)
    channel_order_list = replaceint(row[3].value)
    product_name = replacenone(row[6].value)
    product_option = replacenone(row[7].value)
    quantity = replaceint(row[8].value)
    payment_at = row[5].value

    orderValues = (
        state,
        channel_order_number,
        channel_order_list,
        product_name,
        product_option,
        quantity,
        payment_at,
        state
    )
    cursor.execute(orderSql, orderValues)
    print(orderSql, orderValues)
db.close()




path = bflowResultExcel

wb = load_workbook(path)

ws = wb.active

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()


sql = '''INSERT INTO `excel`.`sell` (
        product_order_number,
        order_number,
        payment_at,
        order_state,
        claim,
        provider_name,
        product_name,
        product_option,
        channel,
        product_number,
        product_amount,
        option_amount,
        seller_discount,
        quantity,
        total_amount,
        delivery_at,
        delivery_complete,
        order_complete_at,
        auto_complete_at,
        category_number,
        buyer_email,
        buyer_gender,
        buyer_age,
        crawler,
        provider_number,
        channel_order_number,
        week,
        month
        ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE payment_at = %s, order_state = %s, claim = %s, delivery_at = %s, delivery_complete = %s,
        order_complete_at =%s, auto_complete_at = %s, channel_order_number = %s, week = %s, month = %s
         '''

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)

for row in rows:
    product_order_number = replacenone(row[0].value)
    order_number = replacenone(row[1].value)
    payment_at = replacedate(row[4].value)
    order_state = replacenone(row[5].value)
    claim = replacenone(row[6].value)
    provider_name = replacenone(row[7].value)
    product_name = replacenone(row[8].value)
    product_option = replacenone(row[9].value)
    channel = replacenone(row[10].value)
    product_number = replacenone(row[20].value)
    product_amount = replaceint(row[21].value)
    option_amount = replaceint(row[22].value)
    seller_discount = replaceint(row[23].value)
    quantity = replaceint(row[24].value)
    total_amount = replaceint(row[25].value)
    delivery_at = replacedate(row[26].value)
    delivery_complete = replacedate(row[27].value)
    order_complete_at = replacedate(row[28].value)
    auto_complete_at = replacedate(row[29].value)
    category_number = replacenone(row[41].value)
    buyer_email = replacenone(row[42].value)
    buyer_gender = replacenone(row[43].value)
    buyer_age = replacenone(row[44].value)
    crawler = replacenone(row[45].value)
    provider_number = replacenone(row[46].value)
    channel_order_number = replacenone(row[3].value)
    if payment_at is not None:
        monthStr = datetime.strptime(payment_at, '%Y-%m-%d')
        week = monthStr.isocalendar()[1]
        month = monthStr.month
    else:
        week = None
        month = None

    values = (
        product_order_number, order_number, payment_at, order_state, claim, provider_name, product_name, product_option,
        channel, product_number, product_amount, option_amount, seller_discount, quantity, total_amount, delivery_at,
        delivery_complete, order_complete_at, auto_complete_at, category_number, buyer_email, buyer_gender, buyer_age,
        crawler, provider_number, channel_order_number, week, month, payment_at, order_state, claim, delivery_at,
        delivery_complete, order_complete_at, auto_complete_at, channel_order_number, week, month,
    )
    cursor.execute(sql, values)

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
totalNow = makeToday.strftime("%Y-%m-%d")
makeLastMonth = makeToday - dateutil.relativedelta.relativedelta(months=1)
endNow = makeLastMonth.strftime("%Y-%m-%d")


cancelList = f'''
            select s.`product_order_number`, c.`channel_order_number`, s.`product_option` 
            from `11st_cancel` as c join `sell` as s on c.`channel_order_number` = s.`channel_order_number`
            where c.`claim_complete` >= {totalNow} group by s.`product_order_number`;
            '''
cursor.execute(cancelList)
cancelRows = cursor.fetchall()

p = re.compile('_F[0-9]+_')

wb = Workbook()

ws = wb.active
no = 2
for cancelRow in cancelRows:
    productOrderNumber = cancelRow[0]
    orderNumber = cancelRow[1]
    if cancelRow[2] is None:
        productOption = None
    else:
        productOption = p.search(cancelRow[2]).group()
    print(productOrderNumber)
    print(orderNumber)
    print(productOption)
    cancelState = f'''
            select s.`product_order_number`, s.`order_number`,
            c.`channel_order_number`,c.`product_name`,c.`product_option`,s.`claim`,
            s.`order_state`, c.`state`, c.`cancel_reason`, c.`cancel_detail_reason`
            from sell as s join `11st_cancel` as c on s.`channel_order_number` = c.`channel_order_number`
            where s.`product_order_number` = {productOrderNumber} and c.`product_option` like '%{productOption}%'
            '''
    cursor.execute(cancelState)
    cancelNowTotal = cursor.fetchall()
    for cancelNow in cancelNowTotal:
        product_order_number = cancelNow[0]
        order_number = cancelNow[1]
        channel_order_number = cancelNow[2]
        product_name = cancelNow[3]
        product_option = cancelNow[4]
        claim = cancelNow[5]
        orderState = cancelNow[6]
        state = cancelNow[7]
        cancelReason = cancelNow[8]
        cancelDetailReason = cancelNow[9]

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '주문번호'
        ws.cell(row=1, column=3).value = '외부채널주문번호'
        ws.cell(row=1, column=4).value = '상품명'
        ws.cell(row=1, column=5).value = '상품옵션'
        ws.cell(row=1, column=6).value = '브리치 클레임상태'
        ws.cell(row=1, column=7).value = '브리치 주문상태'
        ws.cell(row=1, column=8).value = '11번가 상태'
        ws.cell(row=1, column=9).value = '11번가 클레임이'
        ws.cell(row=1, column=10).value = '11번가 클레임상세이유'
        print(orderState, state)
        if orderState == '결제취소' and state == '취소완료':
            print('skip')
            continue
        else:
            ws.cell(row=no, column=1).value = product_order_number
            ws.cell(row=no, column=2).value = order_number
            ws.cell(row=no, column=3).value = channel_order_number
            ws.cell(row=no, column=4).value = product_name
            ws.cell(row=no, column=5).value = product_option
            ws.cell(row=no, column=6).value = claim
            ws.cell(row=no, column=7).value = orderState
            ws.cell(row=no, column=8).value = state
            ws.cell(row=no, column=9).value = cancelReason
            ws.cell(row=no, column=10).value = cancelDetailReason

            no += 1

result = config.ST_LOGIN['excelPath'] + '11stCancelResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()

orderList = f'''
            select s.`product_order_number`, c.`channel_order_number`, s.`product_option` 
            from `11st_order` as c join `sell` as s on c.`channel_order_number` = s.`channel_order_number`
            where c.`payment_at` >= {endNow} group by s.`product_order_number`;
            '''

cursor.execute(orderList)
orderRows = cursor.fetchall()

p = re.compile('_F[0-9]+_')

wb = Workbook()

ws = wb.active
no = 2
for orderRow in orderRows:
    productOrderNumber = orderRow[0]
    orderNumber = orderRow[1]
    if orderRow[2] is None:
        productOption = None
    else:
        productOption = p.search(orderRow[2]).group()
    print(productOrderNumber)
    print(orderNumber)
    print(productOption)
    orderState = f'''
            select s.`product_order_number`, s.`order_number`,
            c.`channel_order_number`,c.`product_name`,c.`product_option`,s.`claim`,
            s.`order_state`, c.`state`
            from sell as s join `11st_order` as c on s.`channel_order_number` = c.`channel_order_number`
            where s.`product_order_number` = {productOrderNumber} and c.`product_option` like '%{productOption}%'
            '''
    cursor.execute(orderState)
    orderNowTotal = cursor.fetchall()
    print(orderNowTotal)
    for orderNow in orderNowTotal:
        product_order_number = orderNow[0]
        order_number = orderNow[1]
        channel_order_number = orderNow[2]
        product_name = orderNow[3]
        product_option = orderNow[4]
        claim = orderNow[5]
        orderState = orderNow[6]
        state = orderNow[7]

        ws.cell(row=1, column=1).value = '상품주문번호'
        ws.cell(row=1, column=2).value = '주문번호'
        ws.cell(row=1, column=3).value = '외부채널주문번호'
        ws.cell(row=1, column=4).value = '상품명'
        ws.cell(row=1, column=5).value = '상품옵션'
        ws.cell(row=1, column=6).value = '브리치 주문상태'
        ws.cell(row=1, column=7).value = '브리치 클레임상태'
        ws.cell(row=1, column=8).value = '11번가 상태'
        print(orderState, state)
        if orderState == '배송준비' or orderState == '결제확인' or orderState == '배송지연' and state == '배송준비중':
            print('skip')
            continue
        else:
            ws.cell(row=no, column=1).value = product_order_number
            ws.cell(row=no, column=2).value = order_number
            ws.cell(row=no, column=3).value = channel_order_number
            ws.cell(row=no, column=4).value = product_name
            ws.cell(row=no, column=5).value = product_option
            ws.cell(row=no, column=6).value = claim
            ws.cell(row=no, column=7).value = orderState
            ws.cell(row=no, column=8).value = state

            no += 1

result = config.ST_LOGIN['excelPath'] + '11stOrderResult_' + totalNow + "_" + now + '.xlsx'
wb.save(result)
wb.close()
print(result)
db.close()



