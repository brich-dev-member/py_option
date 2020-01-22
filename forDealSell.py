import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tqdm import tqdm
import config

path = '/var/www/works/py_option/2019_딜운영.xlsx'

wb = load_workbook(path)

ws = wb.active

db = pymysql.connect(
    host=config.DATABASE_CONFIG['host'],
    user=config.DATABASE_CONFIG['user'],
    password=config.DATABASE_CONFIG['password'],
    db=config.DATABASE_CONFIG['db'],
    charset=config.DATABASE_CONFIG['charset'],
    autocommit=True)
cursor = db.cursor()

iter_row = iter(ws.rows)
next(iter_row)

rows = tqdm(iter_row)


def replacedate(text):
    if text is None:
        return
    else:
        text = str(text)[0:10]
        return text.strip()

no = 2

for row in rows:
    date = replacedate(row[0].value)
    channelName = row[1].value
    product_number = row[5].value
    add_product_number = row[6].value
    if product_number is None:
        no += 1
    else:
        sql = f'''
            select 
            sum(s.`total_amount`), 
            sum(s.`quantity`),
            p.`provider_name`,
            (select {channelName}_fees from product where product_number = {product_number}),
            sum(c.`channel_calculate`),
            sum(c.`margin`)
            from 
            sell as s inner join product as p on s.`product_number` = p.`product_number` 
            left outer join `calculate` as c on s.`product_order_number` = c.`product_order_number`
            where
            s.product_number in ('{product_number}')
            and s.channel = '{channelName}'
            and s.payment_at >= date_add('{date}', interval -1 day) and s.payment_at <= date_add('{date}', interval 7 day )
        '''

        cursor.execute(sql)
        results = cursor.fetchall()


        for result in results:
            deal_total_amount = result[0]
            deal_total_qty = result[1]
            deal_provider = result[2]
            deal_fees = result[3]
            deal_calculate = result[4]
            deal_margin = result[5]
            if deal_fees is not None:
                deal_fees = (result[3] / 100)
            if deal_total_amount is None or deal_total_qty is None or deal_margin is None:
                deal_ct = None
                deal_profit = None
            else:
                deal_ct = round(deal_total_amount / deal_total_qty, 2)
                deal_profit = round(deal_margin / deal_total_amount, 2)

            ws.cell(row=no, column=8).value = deal_provider
            ws.cell(row=no, column=9).value = deal_fees
            ws.cell(row=no, column=9).number_format = '0.00%;[red]-0.00%'
            ws.cell(row=no, column=10).value = deal_total_amount
            ws.cell(row=no, column=10).number_format = '#,##0;[red]-#,##0'
            ws.cell(row=no, column=11).value = deal_total_qty
            ws.cell(row=no, column=11).number_format = '#,##0;[red]-#,##0'
            ws.cell(row=no, column=12).value = deal_ct
            ws.cell(row=no, column=12).number_format = '#,##0;[red]-#,##0'
            ws.cell(row=no, column=13).value = deal_calculate
            ws.cell(row=no, column=13).number_format = '#,##0;[red]-#,##0'
            ws.cell(row=no, column=14).value = deal_margin
            ws.cell(row=no, column=14).number_format = '#,##0;[red]-#,##0'
            ws.cell(row=no, column=15).value = deal_profit
            ws.cell(row=no, column=15).number_format = '0.00%;[red]-0.00%'
            no += 1


wa = wb.create_sheet('주간통계')

newRow = 1

startWeek = 1
endWeek = startWeek + 5
year = 2020
for week in range(startWeek, endWeek):
    weekSql = f'''
            select s.`week`, min(s.payment_at), max(s.payment_at), s.`channel`,
            count(DISTINCT(s.`product_number`)), sum(s.`total_amount`), sum(s.`quantity`),
            sum(c.`brich_calculate`), sum(c.`channel_calculate`), sum(c.`margin`)
            from sell as s left join `product` as p on s.`product_number` = p.`product_number` 
            left join `calculate` as c on s.`product_order_number` = c.`product_order_number`
            where p.`is_deal` = 1 and s.week = {week} and s.year = {year} group by s.`channel`, s.`week`;
            '''
    cursor.execute(weekSql)
    weekDataSet = cursor.fetchall()
    for weekData in weekDataSet:
        weekNum = weekData[0]
        minDate = datetime.strftime(weekData[1], '%Y-%m-%d')
        maxDate = datetime.strftime(weekData[2], '%Y-%m-%d')
        channel = weekData[3]
        dealCount = weekData[4]
        dealTotalAmount = weekData[5]
        dealTotalQty = weekData[6]
        if dealTotalAmount is not None:
            dealCt = dealTotalAmount / dealTotalQty
        else:
            dealCt = 0
        dealEachCt = dealTotalAmount / dealCount
        dealCalculate = weekData[7]
        dealChannelCalculate = weekData[8]
        dealMargin = weekData[9]

        wa.cell(row=1, column=1).value = '주차'
        wa.cell(row=1, column=2).value = '일'
        wa.cell(row=1, column=3).value = '채널'
        wa.cell(row=1, column=4).value = '딜수'
        wa.cell(row=1, column=5).value = '딜 거래량'
        wa.cell(row=1, column=6).value = '딜 판매량'
        wa.cell(row=1, column=7).value = '딜 객단가'
        wa.cell(row=1, column=8).value = '딜별 거래액'
        wa.cell(row=1, column=9).value = '정산대상금액'
        wa.cell(row=1, column=10).value = '채널정산액'
        wa.cell(row=1, column=11).value = '마진'

        wa.cell(row=newRow, column=1).value = week
        wa.cell(row=newRow, column=2).value = minDate + "~" + maxDate
        wa.cell(row=newRow, column=3).value = channel
        wa.cell(row=newRow, column=4).value = dealCount
        wa.cell(row=newRow, column=4).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=5).value = dealTotalAmount
        wa.cell(row=newRow, column=5).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=6).value = dealTotalQty
        wa.cell(row=newRow, column=6).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=7).value = dealCt
        wa.cell(row=newRow, column=7).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=8).value = dealEachCt
        wa.cell(row=newRow, column=8).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=9).value = dealCalculate
        wa.cell(row=newRow, column=9).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=10).value = dealChannelCalculate
        wa.cell(row=newRow, column=10).number_format = '#,##0;[red]-#,##0'
        wa.cell(row=newRow, column=11).value = dealMargin
        wa.cell(row=newRow, column=11).number_format = '#,##0;[red]-#,##0'

        newRow += 1
    newRow += 2

# for col in ws.columns:
#     max_length = 0
#     columnIndex = col[0].column
#     column = get_column_letter(columnIndex)
#     for cell in col:
#         if max_length < len(str(cell.value)) < 30:
#             max_length = len(str(cell.value))
#         else:
#             pass
#     ws.column_dimensions[column].width = (max_length + 1) * 1.2

# for col in wa.columns:
#     max_length = 0
#     columnIndex = col[0].column
#     column = get_column_letter(columnIndex)
#     for cell in col:
#         if max_length < len(str(cell.value)) < 30:
#             max_length = len(str(cell.value))
#         else:
#             pass
#     wa.column_dimensions[column].width = (max_length + 1) * 1.2

makeToday = datetime.today()
now = makeToday.strftime("%m%d_%H%M")
result = "2020_딜운영_" + now + ".xlsx"
wb.save(result)

cursor.close()
db.close()